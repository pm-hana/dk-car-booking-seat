import streamlit as st
import datetime

# 1. 페이지 기본 설정 및 다크 테마 고정
st.set_page_config(page_title="DK CAR BOOKING", page_icon="🚗", layout="wide")

# 2. 시스템 버전 및 새로고침 업데이트 카운터 연산 (00h 기준 초기화 및 mmdd ver.N 포맷, app.py 수정 저장 시 자동 감지 갱신)
import json
import os
import re
import html
import hmac
import hashlib

# ─────────────────────────────────────────────────────────────
# 🔒 [보안 1] 사용자 입력을 HTML/SVG에 넣기 직전 이스케이프하는 공통 헬퍼
#   이 앱은 좌석 배치도·예약 카드·팝업을 unsafe_allow_html=True로 직접 그린다.
#   신청자 이름·출발지·목적지를 그대로 끼워 넣으면, 한 사람이 넣은 태그가
#   현황판을 여는 '전원의 브라우저'에서 실행된다(저장형 XSS). 전 인원이 매일 쓰면
#   예약 1건으로 전 직원 화면이 영향을 받는다.
#   → 저장값은 원본 그대로 두고, '화면에 출력할 때만' 여기서 막는다.
#   [벤치마킹] NGUYỄN VĂN HẢI(FAE Engineer Staff) "Daekhon Vina Vision"
#              — 사내 PWA에서 XSS 취약점을 찾아 패치한 사례를 참고. 상세: docs/BENCHMARK.md
# ─────────────────────────────────────────────────────────────
def esc(v):
    """HTML/SVG 출력용 이스케이프. None·숫자도 안전하게 문자열로 변환."""
    return html.escape("" if v is None else str(v), quote=True)

# ─────────────────────────────────────────────────────────────
# 🔒 [보안 2] 관리자 PIN — 원문·해시 모두 소스에 두지 않는다
#   INNOVA·SEDONA '운전석'을 클릭하면 뜨는 관리자 로그인 팝업의 숫자 PIN.
#   인증 성공 시 '전체 예약 초기화' 등 관리자 기능이 잠금 해제된다.
#
#   ⚠️ 소스에 PIN을 적으면 배포 저장소와 '커밋 이력'에 남는다. 나중에 값을 바꿔도
#      과거 커밋에서 조회할 수 있어, 값 교체만으로는 절대 닫히지 않는다.
#      → Streamlit Cloud Secrets에 해시로만 넣는다(Firebase 자격증명과 같은 방식).
#
#   설정 방법 (Streamlit Cloud → App settings → Secrets):
#       [admin]
#       salt = "임의 문자열 (예: dkvina-2026)"
#       pin_hash = "sha256(salt + PIN) 16진 문자열"
#   해시 만들기:
#       python -c "import hashlib;print(hashlib.sha256('dkvina-2026' '9137'.encode()).hexdigest())"
#   (해시를 미리 못 만들면 pin_hash 대신 pin = "9137" 로 넣어도 된다 — 값은 Secrets 안에만 존재)
#
#   Secrets 미설정 배포에서는 앱이 멈추지 않도록 기존 PIN으로 계속 동작하되,
#   로그인 팝업에 경고를 띄워 '아직 막히지 않은 상태'임을 드러낸다.
#   [벤치마킹] Đào Văn Bảo(FAE Leader) "K-Pulse" — 평문 비밀번호 저장·DB 키 노출을
#              해시 + 서버측 권한 검사로 해결한 사례.
#              NGUYỄN TRỌNG CHƯƠNG(FAE) "AOI Log Analyzer" — 비밀 키를 .gitignore로 커밋 차단.
#              상세: docs/BENCHMARK.md
# ─────────────────────────────────────────────────────────────
_LEGACY_ADMIN_PIN = "1234"   # Secrets 미설정 배포용 임시 폴백(경고 표시). 설정하면 더 이상 쓰이지 않는다.
ADMIN_MAX_TRIES = 5          # 연속 실패 허용 횟수 — 4자리 PIN은 무제한 시도면 해시로 옮겨도 의미가 없다.

def _admin_pin_config():
    """(기대 해시, salt, Secrets 설정 여부) 반환. 미설정이면 폴백 해시 + configured=False."""
    salt, pin_hash = "", ""
    try:
        sec = st.secrets["admin"]
        salt = str(sec.get("salt", ""))
        pin_hash = str(sec.get("pin_hash", "")).strip().lower()
        if not pin_hash and sec.get("pin"):
            pin_hash = hashlib.sha256((salt + str(sec["pin"])).encode("utf-8")).hexdigest()
    except Exception:
        pass   # secrets.toml 자체가 없는 로컬 개발 등 → 폴백
    if pin_hash:
        return pin_hash, salt, True
    return hashlib.sha256((salt + _LEGACY_ADMIN_PIN).encode("utf-8")).hexdigest(), salt, False

def verify_admin_pin(pin):
    """입력 PIN 검증. 해시 비교는 타이밍 차로 값을 추정당하지 않도록 hmac.compare_digest 사용."""
    pin = (pin or "").strip()
    if not pin.isdigit() or not (4 <= len(pin) <= 8):
        return False
    expected, salt, _ = _admin_pin_config()
    given = hashlib.sha256((salt + pin).encode("utf-8")).hexdigest()
    return hmac.compare_digest(given, expected)

COUNTER_FILE = "version_counter.json"

# [버전 표시 — 읽기 전용]
#   버전 카운트는 "Claude Code가 app.py 코드 작업을 완료할 때"만 1 증가한다.
#   실제 증가·백업은 version_bump.py(Claude Code Stop 훅)가 담당하고,
#   앱은 version_counter.json에 저장된 값을 화면에 "표시만" 한다.
#   → 브라우저 새로고침·사용자 저장·수동 rerun 으로는 절대 증가하지 않는다.
def read_display_version():
    now = datetime.datetime.now()
    if os.path.exists(COUNTER_FILE):
        try:
            with open(COUNTER_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            count = max(int(data.get("count", 1) or 1), 1)
            # 저장된 날짜(YYYY-MM-DD)를 mmdd 접두사로 변환
            try:
                display_date = datetime.datetime.strptime(
                    str(data.get("date", "")), "%Y-%m-%d").strftime("%m%d")
            except Exception:
                display_date = now.strftime("%m%d")
            return f"{display_date} ver.{count}"
        except Exception:
            pass
    return now.strftime("%m%d") + " ver.1"

date_version_str = read_display_version()
st.session_state.current_version_str = date_version_str

# UI 고도화 및 PM님 요청 간격 비율을 100% 매칭하기 위한 프리미엄 커스텀 CSS
st.markdown("""
    <style>
    /* 메인 앱 배경 전체를 럭셔리 다크 블랙 테마로 동기화 */
    .stApp {
        background-color: #0f1014 !important;
    }

    /* 우측 상단 Streamlit 기본 UI(Fork 배지·GitHub·⋮ 메뉴·상단 데코바·푸터) 전부 숨김
       + 우하단 'Hosted with Streamlit' 배지(Community Cloud)까지 신·구 셀렉터 모두 차단 */
    #MainMenu,
    header[data-testid="stHeader"],
    [data-testid="stToolbar"],
    [data-testid="stMainMenu"],
    [data-testid="stDecoration"],
    [data-testid="stStatusWidget"],
    [data-testid="stAppViewerBadge"],
    [class*="viewerBadge"],
    [class*="_viewerBadge"],
    [class*="_profileContainer"],
    [class*="stAppHostedMenu"],
    a[href*="streamlit.io"],
    a[href*="share.streamlit.app"],
    .viewerBadge_container__1QSob,
    .stAppDeployButton,
    footer {
        display: none !important;
        visibility: hidden !important;
    }

    /* ===== 팝업(다이얼로그) 공통: 좁은 화면(폰)에서도 웹과 동일한 구성 유지 ===== */
    /* 날짜/시간 3칸·버튼 2칸 등 폼 컬럼이 폰에서 세로로 쌓이지 않고 한 줄 가로 유지되게
       (Streamlit 모바일 반응형이 컬럼에 min-width:100%를 걸어 wrap 되는 것을 해제) */
    [role="dialog"] [data-testid="stHorizontalBlock"] {
        flex-wrap: nowrap !important;
        gap: 8px !important;
    }
    [role="dialog"] [data-testid="stHorizontalBlock"] > [data-testid="stColumn"] {
        min-width: 0 !important;
        width: auto !important;
        flex: 1 1 0% !important;
    }
    /* 팝업 닫기(X) 아이콘을 다이얼로그 제목(24px)과 동일 크기로 확대 */
    [role="dialog"] button[aria-label="Close"] svg {
        width: 24px !important;
        height: 24px !important;
    }
    /* 좌석 선택/신청 폼 팝업: 크롬 title은 공백(' ')이라 본문 최상단에 단계별 제목을 직접 그린다.
       크롬 제목 자리(빈 공백)가 차지하는 여백을 줄여, 커스텀 제목이 X와 같은 줄 높이에 오게 한다. */
    [role="dialog"] .dlg-step-title {
        font-size: 24px !important;
        font-weight: 700 !important;
        color: #ffffff !important;
        line-height: 1.2 !important;
        margin: -44px 0 12px 0 !important;   /* 빈 크롬 헤더 높이만큼 끌어올려 X와 같은 줄에 정렬 */
        padding-right: 40px !important;       /* 우측 X 버튼과 겹치지 않도록 여백 */
    }

    /* 관리자 '좌석 신청 현황' 표 — 운전석 제외 전체 좌석(빈 좌석 포함) 신청 현황 */
    .seat-status-wrap { width: 100%; overflow-x: auto; margin: 4px 0 14px 0; }
    .seat-status-table {
        width: 100%; border-collapse: collapse; font-size: 13px; color: #e9ecef;
    }
    .seat-status-table th, .seat-status-table td {
        padding: 8px 10px; text-align: left; border-bottom: 1px solid #2b2f38;
        white-space: nowrap;
    }
    .seat-status-table thead th {
        background: #1b1f27; color: #fab005; font-weight: 700;
        border-bottom: 2px solid #3a3f4a; position: sticky; top: 0;
    }
    .seat-status-table td.ss-seat { font-weight: 700; color: #ffffff; text-align: center; width: 60px; }
    .seat-status-table td.ss-time { font-variant-numeric: tabular-nums; color: #adb5bd; font-size: 12px; }
    .seat-status-table tbody tr:hover { background: rgba(250,176,5,0.06); }
    .seat-status-table tbody tr.seat-status-empty td { color: #6c757d; }
    .seat-status-table tbody tr.seat-status-empty td.ss-seat { color: #868e96; }
    /* 오늘 완료된 탑승(누적 이력) 행: 초록 톤으로 현재 예약과 구분 */
    .seat-status-table tbody tr.seat-status-done td { color: #63b365; }
    /* 좌석 현황 하단 버튼: 닫기(75%)·로그아웃(25%) — 다이얼로그 컬럼 1:1 강제 CSS를 스코프로 덮어써 3:1 유지 */
    .st-key-admin_status_btns [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:nth-child(1) { flex: 3 1 0% !important; }
    .st-key-admin_status_btns [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:nth-child(2) { flex: 1 1 0% !important; }
    /* 로그아웃 버튼: 눈에 띄되 과하지 않게(붉은 톤) */
    .st-key-admin_status_logout_btn button { background: #3a1e1e !important; border-color: #7e2a2a !important; color: #ffc9c9 !important; font-weight: 700 !important; }
    .st-key-admin_status_logout_btn button:hover { background: #522727 !important; border-color: #a83232 !important; color: #ffffff !important; }

    /* 타이틀을 화면 최상단부터 시작 — 메인 컨테이너 상단 여백 축소 */
    [data-testid="stMainBlockContainer"],
    [data-testid="stAppViewBlockContainer"],
    .block-container {
        padding-top: 0.4rem !important;
    }

    /* 상단 헤더: 왼쪽 DAEKHON VINA 로고 · 가운데 큰 타이틀 · 오른쪽 버전/시계 */
    .top-header-container {
        display: flex !important;
        flex-direction: row !important;
        align-items: center !important;
        gap: 16px !important;
        width: 100% !important;
        margin: 0 0 2px 0 !important;
        padding: 0 4px !important;
    }
    /* 왼쪽 브랜드 로고 락업(🐋 + DAEKHON VINA) */
    .brand-lockup {
        display: flex !important;
        align-items: center !important;
        gap: 8px !important;
        flex: 0 0 auto !important;
    }
    .brand-mark { font-size: 36px; line-height: 1; }
    .brand-logo-img { height: 42px; width: auto; display: block; }
    .brand-name {
        font-size: 15px;
        font-weight: 800;
        color: #38bdf8;
        letter-spacing: 0.5px;
        line-height: 1.1;
        white-space: nowrap;
    }
    /* 오른쪽 실시간 시계(세로 스택, 우측 정렬) */
    .header-meta {
        display: flex !important;
        flex-direction: column !important;
        align-items: flex-end !important;
        flex: 0 0 auto !important;
        gap: 2px !important;
    }
    /* 버전(0706 ver.N)을 메인 타이틀 오른쪽에 SEAT(4글자)만큼 띄워 배치 */
    .brand-version { margin-left: 4ch !important; }
    /* 실시간 시계: 오른쪽 컬럼에서 우측 정렬(언어 토글 위에 세로 스택) */
    .header-clock { flex: 0 0 auto !important; white-space: nowrap !important; text-align: right !important; }
    /* 언어 선택 라디오를 전체 프레임 오른쪽 끝선에 붙여 우측 정렬(여러 계층 커버) */
    .st-key-lang_toggle { display: flex !important; flex-direction: column !important; align-items: flex-end !important; }
    /* 언어 토글 라벨을 실시간 시계와 동일 크기(16px)로 */
    .st-key-lang_toggle div[data-testid="stRadio"] label { font-size: 16px !important; }
    .st-key-lang_toggle div[data-testid="stRadio"] { width: 100% !important; }
    .st-key-lang_toggle div[data-testid="stRadio"] > div { display: flex !important; justify-content: flex-end !important; }
    .st-key-lang_toggle div[role="radiogroup"] { justify-content: flex-end !important; margin-left: auto !important; }
    /* 오른쪽 헤더 묶음: 시계(위)+토글(아래)을 컴팩트 세로 스택으로 프레임 오른쪽 끝선에 정렬·수직 중앙 */
    .st-key-hdr_right {
        display: flex !important;
        flex-direction: column !important;
        align-items: flex-end !important;   /* 오른쪽 끝선 정렬 */
        justify-content: center !important;
        gap: 2px !important;
        padding-right: 10% !important;      /* 차량 박스(80% 가운데정렬)의 우측 10% 여백과 동일 끝선 */
    }
    .st-key-hdr_right div[data-testid="stVerticalBlock"] { gap: 2px !important; }
    .st-key-hdr_right [data-testid="stElementContainer"] { margin: 0 !important; }
    .st-key-hdr_right .st-key-lang_toggle { margin: 0 !important; padding: 0 !important; }
    /* 예약 이력 버튼: TAXI 박스(width 80% 가운데정렬)와 동일 끝선·폭으로 → 박스 바로 아래 한 줄 정렬 */
    .st-key-csv_inset { padding: 0 10% !important; }
    /* 엑셀 내보내기 팝업의 다운로드 버튼: 엑셀 그린 풀폭 버튼 */
    .st-key-export_dl button { background: #21a366 !important; border-color: #21a366 !important; color: #ffffff !important; font-weight: 700 !important; min-height: 46px !important; }
    .st-key-export_dl button:hover { background: #1a8551 !important; border-color: #1a8551 !important; color: #ffffff !important; }
    /* 예약 현황 카드: 한 줄에 '항상 2개'. Streamlit의 모바일 컬럼 세로적층을 이기기 위해
       자식결합자(>)로 선택자 특이도를 (0,3,0)까지 올려 nowrap + 균등분배를 강제한다. + 가로 스크롤 차단. */
    .st-key-booking_board { overflow-x: hidden !important; max-width: 100% !important; }
    .st-key-booking_board [data-testid="stHorizontalBlock"] { flex-wrap: nowrap !important; flex-direction: row !important; gap: 6px !important; }
    /* (0,3,0) → Streamlit 모바일 적층(컬럼 flex-basis:100%)을 확실히 덮어씀.
       바깥 2열은 50%씩(2개가 화면에 딱), 안쪽 버튼 3열은 1/3씩 (같은 규칙으로 각 레벨 균등 분배). */
    .st-key-booking_board [data-testid="stHorizontalBlock"] > [data-testid="stColumn"] { flex: 1 1 0% !important; width: auto !important; min-width: 0 !important; padding: 0 1px !important; }
    /* 가로 3분할 버튼(수정·취소·도착완료): word-break:keep-all로 '예약/수정'처럼 단어 사이(공백)에서만 줄바꿈 → 2줄 표시.
       버튼 내부 텍스트 요소(p/div/span)까지 직접 적용해야 CJK 글자 중간 끊김('예약 수'/'정')을 확실히 막는다. */
    .st-key-booking_board .stButton { margin-bottom: 0 !important; }
    .st-key-booking_board .stButton button,
    .st-key-booking_board .stButton button * { word-break: keep-all !important; overflow-wrap: normal !important; white-space: normal !important; }
    .st-key-booking_board .stButton button { padding: 2px 1px !important; font-size: 11px !important; line-height: 1.1 !important; min-height: 32px !important; width: 100% !important; }
    .main-title {
        flex: 0 0 auto;                      /* 크기 고정(title-group이 flex 담당) */
        font-size: 40px !important;          /* 다른 문구보다 확실히 크게(메인 타이틀 강조) */
        font-weight: bold;
        color: #ffffff;
        margin: 0 !important;
        text-align: center;
        white-space: nowrap;
        letter-spacing: 1px;
    }
    /* 메인 타이틀 + 버전을 한 묶음으로 가운데 배치, 버전은 타이틀 오른쪽에 붙는다 */
    .title-group {
        flex: 1 1 auto;
        display: flex;
        justify-content: center;
        align-items: baseline;
        min-width: 0;
    }
    
    /* 노란색 외곽 테두리 박스를 완전히 없앤 미니멀 폰트 스타일링 */
    .clean-timestamp-stamp {
        color: #fcc419 !important;
        font-family: inherit !important;   /* 타이틀과 동일한 단일 폰트로 통일 */
        font-size: 16px !important;
        font-weight: bold !important;
        padding: 0 !important;
        line-height: 1 !important;
        white-space: nowrap !important;    /* 버전·시계가 두 줄로 접히지 않게 */
    }
    /* 현재 시각은 헤더 오른쪽에 정렬 */
    #live-digital-clock {
        text-align: right !important;
    }
    .sub-title {
        font-size: 13px;
        color: #8e929e;
        margin-bottom: 20px;
    }
    /* 예약 현황판 제목(expander 제거 후 헤더 행에서 사용) */
    .board-title {
        font-size: 15px;
        font-weight: bold;
        color: #fafafa;
        /* 상하 0 → 오른쪽 '예약 이력' 버튼과 수직 중심선 일치. 왼쪽 5% → INNOVA 박스(80%가운데) 왼쪽 끝선 정렬 */
        padding: 0 0 0 5% !important;
        min-height: 38px;           /* download_button 높이에 맞춰 라인 정렬 */
        display: flex;
        align-items: center;
        white-space: nowrap;
    }
    
    /* 일반 차량 제목 영역 가이드 (가로 중앙 정렬) */
    .car-header-center {
        text-align: center;
        margin-top: 10px !important;
        margin-bottom: 5px !important;
        min-height: 35px;
        display: flex;
        justify-content: center;
        align-items: center;
    }
    .car-title-text {
        font-size: clamp(13px, 1.5vw, 19px);   /* 창 폭에 따라 자동 축소 → 좁은 창에서 옆 칸 침범 방지 */
        font-weight: bold;
        color: #ffffff;
        margin: 0 !important;
        white-space: nowrap;                    /* 브랜드명 줄바꿈 금지 → 4칸 헤더 높이 균일 유지 */
        overflow: hidden;
        text-overflow: ellipsis;
        max-width: 100%;
    }
    /* 차량명 사각 프레임: 배경색은 각 차량 외관색(이노바=화이트/세도나=블랙/VF5=레드/택시=옐로우) 기준.
       폭은 아래 차량 사진 박스(.car-layout-container 80%)와 동일하게 80% 가운데 정렬로 좌우 끝선 일치. */
    .car-name-frame {
        display: flex;
        align-items: center;
        justify-content: center;
        width: 80%;
        margin: 0 auto;
        box-sizing: border-box;
        padding: 7px 12px;
        border-radius: 9px;
        box-shadow: 0 2px 7px rgba(0, 0, 0, 0.4);
    }
    
    /* 라디오 버튼(현재는 언어 토글 한국어/ENG) 컨테이너 기본 플렉스 정렬 */
    div[data-testid="stRadio"] {
        margin: 5px auto 0 auto !important;
        padding: 0 !important;
        display: flex !important;
        justify-content: center !important;
        width: 100% !important;
    }
    div[data-testid="stRadio"] > div[role="radiogroup"] {
        display: flex !important;
        flex-direction: row !important;
        justify-content: center !important;
        gap: 16px !important; 
        width: auto !important;
    }
    div[data-testid="stRadio"] label {
        color: #ffffff !important;
        font-size: 14px !important;
        font-weight: bold !important;
        padding: 0px 2px !important;
        margin: 0 !important;
    }

    /* 언어 선택 토글(한국어/Tiếng Việt/ENG)은 항상 한 줄에 나란히 우측 정렬.
       3종으로 늘면서 폭이 커졌으므로 항목 간격을 12px → 10px로 좁힌다. */
    .st-key-lang_toggle div[role="radiogroup"] {
        flex-wrap: nowrap !important;
        justify-content: flex-end !important;
        gap: 10px !important;
    }
    .st-key-lang_toggle div[role="radiogroup"] label {
        white-space: nowrap !important;
    }

    /* 차량 박스: 컬럼 폭(=아래 배차 현황 카드 폭)에 꽉 채우고, 높이는 세로 이미지 비율로 자동 산출
       → 상단 차량 폭과 하단 배차 카드 폭이 정확히 일치, 사진도 왜곡/여백 없이 채워짐 */
    .car-layout-container {
        background-color: #1a1c23;
        border: 2px solid #3f4452;
        border-radius: 12px;
        padding: 10px;
        display: flex;
        justify-content: center;
        align-items: center;
        width: 80% !important;        /* 컬럼 폭의 80%로 축소 */
        aspect-ratio: 160 / 250;
        height: auto !important;      /* 높이는 폭에서 자동 산출(세로 비율 유지) */
        margin: 0 auto !important;    /* 컬럼 안에서 가운데 정렬 */
        box-shadow: 0 8px 24px rgba(0, 0, 0, 0.5);
    }
    
    /* 좌석 배치도 사각형 프레임과 하단 선택 토글 간격을 정확히 20px로 격리 */
    .dropdown-spacing-wrapper {
        margin-top: 20px !important; 
    }
    
    /* 하단 임베디드 신청서 양식 박스 스타일 */
    .booking-form-box {
        background-color: #15161a;
        border: 2px dashed #38bdf8;
        border-radius: 10px;
        padding: 20px;
        margin-top: 20px;
        box-shadow: 0 4px 15px rgba(56, 189, 248, 0.15);
    }
    
    /* 중복 경고 알림창 커스텀 스타일 */
    .custom-error-box {
        background-color: rgba(224, 49, 49, 0.15);
        border: 1px solid #e03131;
        border-radius: 8px;
        padding: 12px 16px;
        display: flex;
        justify-content: space-between;
        align-items: center;
        width: 100%;
        margin-bottom: 5px;
    }
    .custom-error-text {
        color: #ffc9c9 !important;
        font-size: 14px !important;
        font-weight: bold !important;
        margin: 0 !important;
        line-height: 1.4 !important;
    }
    .custom-error-close-btn {
        cursor: pointer !important;
        font-size: 14px !important;
        margin-left: 15px !important;
        user-select: none !important;
        color: #ffc9c9 !important;
        text-decoration: none !important;
        transition: opacity 0.2s ease;
    }
    .custom-error-close-btn:hover {
        opacity: 0.7;
        color: #ffffff !important;
    }
    
    /* 좌석 마우스 오버 시 클릭 가능한 포인터 핸들 스타일 주입 */
    .clickable-seat-rect {
        cursor: pointer !important;
        transition: transform 0.1s ease;
    }
    .clickable-seat-rect:hover {
        filter: brightness(1.25);
    }
    /* 빈 좌석 그룹 전체를 클릭 가능한 포인터로 표시 */
    .seat-clickable {
        cursor: pointer !important;
    }
    /* INNOVA·SEDONA 운전석: 관리자 로그인용 클릭 가능 표시 */
    .admin-login-seat {
        cursor: pointer !important;
    }
    .admin-login-seat:hover { filter: brightness(1.25); }

    /* 좌석 클릭을 부드러운 rerun으로 처리하기 위한 숨김 버튼 (화면 밖 배치, JS가 대신 클릭) */
    div[class*="st-key-seatsel_"],
    div[class*="st-key-adminlogin_"],
    div[class*="st-key-restore_admin"],
    div[class*="st-key-carnavclick_"] {
        position: fixed !important;
        left: -9999px !important;
        top: -9999px !important;
        width: 1px !important;
        height: 1px !important;
        overflow: hidden !important;
        margin: 0 !important;
        padding: 0 !important;
    }
    /* 앱: 클릭 가능한 차량 이름 바(로고+이름 프레임 전체가 버튼처럼). 바 사이 세로 간격 확보(앱·웹 동일). */
    .car-nav-click { cursor: pointer !important; margin: 6px 0 !important; }
    .car-nav-click .car-name-frame { width: 100% !important; transition: transform 0.08s ease, box-shadow 0.08s ease; }
    .car-nav-click:hover .car-name-frame { transform: translateY(-1px); box-shadow: 0 4px 12px rgba(0,0,0,0.5); }
    /* 이름 바 바로 아래 숨김 CARNAV 버튼 컨테이너가 차지하던 빈 세로 공간 제거(간격 이중 발생 방지) */
    div[class*="st-key-carnavclick_"] { height: 0 !important; }

    /* ===== 메인 차량 선택: 정사각형 타일 2×2 =====
       가로로 긴 바는 탭 영역이 화면 폭 전체라 옆 차량까지 잘못 눌리기 쉬웠다(오클릭).
       정사각형으로 줄이고 2열로 배치해 각 타일의 경계를 분명히 한다. */
    .st-key-car_nav_grid { max-width: 520px !important; margin: 0 auto 4px auto !important; }
    /* 모바일에서도 2열 유지 — Streamlit의 컬럼 세로적층(flex-basis:100%)을 자식결합자 특이도(0,3,0)로 덮어씀 */
    .st-key-car_nav_grid [data-testid="stHorizontalBlock"] { flex-wrap: nowrap !important; flex-direction: row !important; gap: 10px !important; }
    .st-key-car_nav_grid [data-testid="stHorizontalBlock"] > [data-testid="stColumn"] { flex: 1 1 0% !important; width: auto !important; min-width: 0 !important; }
    .car-nav-tile { margin: 0 0 10px 0 !important; }
    /* 정사각형 본체: 로고(위) + 이름(아래) 세로 스택 */
    .car-nav-tile .car-name-frame {
        width: 100% !important;
        max-width: 240px !important;
        margin: 0 auto !important;
        aspect-ratio: 1 / 1;
        flex-direction: column;
        gap: 10px;
        padding: 12px !important;
        border-radius: 14px;
    }
    .car-nav-tile:hover .car-name-frame { transform: translateY(-2px); box-shadow: 0 6px 16px rgba(0,0,0,0.55); }
    /* 로고는 가로 바 시절의 우측 여백(inline margin-right)을 지우고 크게 — 세로 스택이라 필요 없다.
       (로고·차량명은 운행 정보보다 확실히 커야 해서 이전 대비 50% 확대: 54×34 → 81×51) */
    .car-nav-tile .car-nav-logo { display: flex; align-items: center; justify-content: center; font-size: 51px; line-height: 1; }
    .car-nav-tile .car-nav-logo svg,
    .car-nav-tile .car-nav-logo img { width: 81px !important; height: 51px !important; margin-right: 0 !important; }
    /* 이름은 타일 폭에 맞춰 줄바꿈 허용(단어 사이에서만). 이전 대비 50% 확대 */
    .car-nav-tile .car-title-text {
        white-space: normal !important;
        text-align: center;
        line-height: 1.2;
        font-size: clamp(18px, 2.25vw, 25px) !important;
        word-break: keep-all;
        overflow: visible;
        text-overflow: clip;
    }
    /* 운행 정보 패널: 어두운 스크림(alpha 0.55) 위 흰 글자 — 실버·블랙·옐로우 어느 타일에서도 최저 7.0:1 */
    .car-nav-tile .car-nav-info {
        background: rgba(0, 0, 0, 0.55);
        border-radius: 8px;
        padding: 5px 9px;
        text-align: center;
        color: #ffffff;
        line-height: 1.3;
        max-width: 100%;
    }
    /* 운전자 이름은 정보 3줄 중 대표값이라 조금 크고 굵게, 번호·연락처는 한 단계 작게 */
    .car-nav-tile .cni-driver { font-size: 14px; font-weight: 800; letter-spacing: 0.3px; }
    .car-nav-tile .cni-line { font-size: 12px; font-variant-numeric: tabular-nums; white-space: nowrap; }
    
    /* 드래그 대상 마우스 커서 grab/grabbing 형태 지정 */
    [draggable="true"] {
        cursor: grab !important;
    }
    [draggable="true"]:active {
        cursor: grabbing !important;
    }

    /* ===== 중간 폭(창 축소·태블릿) 자동 최적화 (769px~1200px) =====
       4개 차량 컬럼을 그대로 유지하되, 제목·간격·배치도를 창 폭에 맞춰 자동 축소해
       좁은 창에서 옆 칸과 겹치는(오버랩) 현상을 방지. 폰(≤768px)·풀와이드에는 영향 없음. */
    @media (min-width: 769px) and (max-width: 1200px) {
        /* 가로 넘침 원천 차단 */
        html, body, .stApp { overflow-x: hidden !important; }
        .block-container { padding-left: 0.5rem !important; padding-right: 0.5rem !important; }
        /* 차량 컬럼 사이 간격 축소 → 각 칸에 여유 폭 확보 */
        div[data-testid="stHorizontalBlock"] { gap: 0.4rem !important; }
        /* 제목·헤더 축소로 브랜드명이 옆 칸을 침범하지 않게 */
        .car-title-text { font-size: clamp(11px, 1.4vw, 16px) !important; }
        .car-header-center { min-height: 30px !important; }
        /* 배치도 박스는 컬럼 폭을 더 채우되(90%) 넘치지 않도록 패딩 축소 */
        .car-layout-container { width: 90% !important; padding: 6px !important; }
    }

    /* ===== 모바일/좁은 화면 자동 최적화 (max-width 768px) =====
       화면 폭 기준이라 ?m=1 파라미터가 없어도(안드로이드 홈화면 앱 등) 폰에서 자동 적용.
       PC(넓은 화면)에는 전혀 영향 없음. */
    @media (max-width: 768px) {
        /* 좌우 여백 최소화 + 가로 스크롤 방지 */
        .block-container { padding: 0.3rem 0.6rem 2.5rem !important; max-width: 100% !important; }
        html, body, .stApp { overflow-x: hidden !important; }

        /* 헤더를 세로 가운데 스택으로: 브랜드(위) → 타이틀+버전 → 시계 → 토글 */
        .top-header-container { flex-wrap: wrap !important; justify-content: center !important; gap: 4px !important; }
        .brand-lockup { order: 0 !important; width: 100% !important; justify-content: center !important; }
        .title-group { order: 1 !important; width: 100% !important; justify-content: center !important; align-items: baseline !important; flex-wrap: wrap !important; }
        /* 타이틀: 화면 폭에 맞춰 자동 축소(넘치면 줄바꿈) → 폰에서 양옆 잘림 방지 */
        .main-title { font-size: clamp(22px, 7.5vw, 40px) !important; text-align: center !important; letter-spacing: 0.2px !important; white-space: normal !important; max-width: 100% !important; }
        /* 시계 가운데: base의 #live-digital-clock(우측정렬 ID규칙)을 눌러야 하므로 ID로 지정 */
        .header-clock, #live-digital-clock { text-align: center !important; font-size: 12px !important; }
        /* 시계+토글 묶음: PC용 우측정렬·10% 인셋 해제 → 가운데 정렬 */
        .st-key-hdr_right { align-items: center !important; padding-right: 0 !important; }
        .brand-logo-img { height: 28px !important; }
        .brand-name { font-size: 12px !important; }
        .brand-version { font-size: 12px !important; margin-left: 6px !important; }
        .sub-title { font-size: 11px !important; text-align: center !important; margin-bottom: 8px !important; }

        /* 언어 선택은 모바일에선 가운데 정렬 + 3종(한국어/Tiếng Việt/ENG)이 한 줄에 들어가도록 축소 */
        .st-key-lang_toggle, .st-key-lang_toggle [role="radiogroup"] { justify-content: center !important; align-items: center !important; }
        .st-key-lang_toggle div[role="radiogroup"] { gap: 8px !important; }
        .st-key-lang_toggle div[data-testid="stRadio"] label { font-size: 13px !important; }

        /* 차량/예약 컬럼을 한 화면에 하나씩 세로로 강제 스택 */
        div[data-testid="stHorizontalBlock"] { flex-wrap: wrap !important; }
        div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"],
        div[data-testid="stHorizontalBlock"] > div[data-testid="column"] {
            flex: 1 1 100% !important;
            width: 100% !important;
            min-width: 100% !important;
        }

        /* 차량 선택 타일: 폰에서도 2열 유지하되 화면 폭을 꽉 채워 탭 영역을 키운다 */
        .st-key-car_nav_grid { max-width: 100% !important; }
        .st-key-car_nav_grid [data-testid="stHorizontalBlock"] { gap: 8px !important; }
        .car-nav-tile .car-name-frame { max-width: none !important; padding: 8px !important; gap: 6px; border-radius: 12px; }
        .car-nav-tile .car-nav-logo { font-size: 28px; }
        .car-nav-tile .car-nav-logo svg,
        .car-nav-tile .car-nav-logo img { width: 44px !important; height: 28px !important; }
        .car-nav-tile .car-title-text { font-size: 13px !important; }

        /* 차량 박스: 화면 폭 88%·세로비율(160:250)로, 가운데 */
        .car-layout-container { width: 88% !important; height: auto !important; aspect-ratio: 160 / 250 !important; max-height: 62vh; margin: 2px auto 6px !important; padding: 6px !important; }
        .car-title-text { font-size: 16px !important; }
        .car-header-center { min-height: 26px !important; }
        /* 모바일에선 PC용 좌/우 인셋(제목 5%·버튼 10%) 해제 → 전체폭 기준 */
        .board-title { font-size: 14px !important; padding-left: 0 !important; }
        .st-key-csv_inset { padding: 0 !important; }
    }
    </style>
""", unsafe_allow_html=True)

# 2-b. 모바일 전용 주소(?m=1 또는 ?view=mobile) 감지 → 모바일 최적화 CSS만 추가 적용.
#      파라미터가 없으면 기존 PC 화면 그대로(불변). 폰에서 이 주소를 북마크해 사용.
try:
    _qp = st.query_params
    IS_MOBILE = (_qp.get("m") == "1") or (str(_qp.get("view", "")).lower() == "mobile")
except Exception:
    IS_MOBILE = False

if IS_MOBILE:
    st.markdown("""
    <style>
    /* ===== 모바일 전용(?m=1) 화면 최적화 — PC 화면에는 영향 없음 ===== */
    /* 좌우 여백 최소화 + 가로 스크롤 방지 */
    .block-container { padding: 0.3rem 0.6rem 2.5rem !important; max-width: 100% !important; }
    html, body, .stApp { overflow-x: hidden !important; }

    /* 헤더를 세로 가운데 스택으로: 브랜드(위) → 타이틀+버전 → 시계 → 토글 */
    .top-header-container { flex-wrap: wrap !important; justify-content: center !important; gap: 4px !important; }
    .brand-lockup { order: 0 !important; width: 100% !important; justify-content: center !important; }
    .title-group { order: 1 !important; width: 100% !important; justify-content: center !important; align-items: baseline !important; flex-wrap: wrap !important; }
    /* 타이틀: 화면 폭에 맞춰 자동 축소(넘치면 줄바꿈) → 폰에서 양옆 잘림 방지 */
    .main-title { font-size: clamp(22px, 7.5vw, 40px) !important; text-align: center !important; letter-spacing: 0.2px !important; white-space: normal !important; max-width: 100% !important; }
    /* 시계 가운데: base의 #live-digital-clock(우측정렬 ID규칙)을 눌러야 하므로 ID로 지정 */
    .header-clock, #live-digital-clock { text-align: center !important; font-size: 13px !important; }
    /* 시계+토글 묶음: PC용 우측정렬·10% 인셋 해제 → 가운데 정렬 */
    .st-key-hdr_right { align-items: center !important; padding-right: 0 !important; }
    .brand-logo-img { height: 30px !important; }
    .brand-name { font-size: 13px !important; }
    .brand-version { font-size: 13px !important; margin-left: 6px !important; }
    .sub-title { font-size: 11px !important; text-align: center !important; margin-bottom: 10px !important; }

    /* 언어 선택은 모바일에선 가운데 정렬 + 3종(한국어/Tiếng Việt/ENG)이 한 줄에 들어가도록 축소 */
    .st-key-lang_toggle, .st-key-lang_toggle [role="radiogroup"] { justify-content: center !important; align-items: center !important; }
    .st-key-lang_toggle div[role="radiogroup"] { gap: 8px !important; }
    .st-key-lang_toggle div[data-testid="stRadio"] label { font-size: 13px !important; }

    /* 차량 선택 타일: 폰에서도 2열 유지하되 화면 폭을 꽉 채워 탭 영역을 키운다 */
    .st-key-car_nav_grid { max-width: 100% !important; }
    .st-key-car_nav_grid [data-testid="stHorizontalBlock"] { gap: 8px !important; }
    .car-nav-tile .car-name-frame { max-width: none !important; padding: 8px !important; gap: 6px; border-radius: 12px; }
    .car-nav-tile .car-nav-logo { font-size: 42px; }
    .car-nav-tile .car-nav-logo svg,
    .car-nav-tile .car-nav-logo img { width: 66px !important; height: 42px !important; }
    .car-nav-tile .car-title-text { font-size: 20px !important; }
    .car-nav-tile .car-nav-info { padding: 4px 7px; border-radius: 7px; }
    .car-nav-tile .cni-driver { font-size: 12px; }
    .car-nav-tile .cni-line { font-size: 10.5px; }

    /* 차량 박스: 화면 높이 기준 적당한 세로 크기로 고정(폭은 세로비율로 자동), 가운데 */
    .car-layout-container { width: auto !important; height: 58vh !important; max-height: 470px !important; aspect-ratio: 160 / 250 !important; margin: 2px auto 6px !important; padding: 6px !important; }
    /* 차량명 프레임: 모바일에선 차량 박스와 동일 폭(58vh*160/250, 최대 301px)으로 가운데 정렬 → 좌우 끝선 일치 */
    .car-name-frame { width: min(calc(58vh * 0.64), 301px) !important; max-width: 100% !important; }
    /* 예약 현황 카드: 앱에서도 한 줄에 '항상 2개'. 자식결합자(>)로 특이도(0,3,0) → Streamlit 모바일 적층을 확실히 이김.
       (홀수면 오른쪽 칸 빈 채 유지. 버튼은 정보 아래 가로 3분할로 카드 높이 최소화. 가로 스크롤 차단.) */
    .st-key-booking_board { overflow-x: hidden !important; max-width: 100% !important; }
    .st-key-booking_board [data-testid="stHorizontalBlock"] { flex-wrap: nowrap !important; flex-direction: row !important; gap: 6px !important; }
    .st-key-booking_board [data-testid="stHorizontalBlock"] > [data-testid="stColumn"] { flex: 1 1 0% !important; width: auto !important; min-width: 0 !important; margin: 0 !important; padding: 0 1px !important; }
    .st-key-booking_board .stButton { margin-bottom: 0 !important; }
    .st-key-booking_board .stButton button,
    .st-key-booking_board .stButton button * { word-break: keep-all !important; overflow-wrap: normal !important; white-space: normal !important; }
    .st-key-booking_board .stButton button { min-height: 32px !important; font-size: 11px !important; padding: 2px 1px !important; line-height: 1.1 !important; }
    .car-title-text { font-size: 16px !important; }
    .car-header-center { min-height: 26px !important; }

    /* 예약 현황판 제목·카드 폰트 소폭 축소 + PC용 좌/우 인셋 해제 */
    .board-title { font-size: 14px !important; padding-left: 0 !important; }
    .st-key-csv_inset { padding: 0 !important; }
    </style>
    """, unsafe_allow_html=True)

# 2-c. PWA(홈 화면 추가) 설정 주입 — 안드로이드 크롬 메뉴 '홈 화면에 추가' 시 아이콘·전체화면(standalone) 앱으로 실행.
#      Streamlit은 <head>를 직접 못 건드리므로, 컴포넌트(iframe)에서 부모 document의 <head>에 manifest·메타를 주입한다.
#      manifest/아이콘은 static/ 폴더(enableStaticServing) → /app/static/ 로 서빙된다.
import streamlit.components.v1 as _pwa_components
_pwa_components.html("""
<script>
(function () {
  try {
    var head = window.parent.document.head;
    if (head.querySelector('link[rel="manifest"]')) return;   // 중복 주입 방지
    var tags = [
      ['link', {rel: 'manifest', href: '/app/static/manifest.json'}],
      ['link', {rel: 'apple-touch-icon', href: '/app/static/icon-192.png'}],
      ['meta', {name: 'theme-color', content: '#0e1117'}],
      ['meta', {name: 'mobile-web-app-capable', content: 'yes'}],
      ['meta', {name: 'apple-mobile-web-app-capable', content: 'yes'}],
      ['meta', {name: 'apple-mobile-web-app-status-bar-style', content: 'black-translucent'}],
      ['meta', {name: 'apple-mobile-web-app-title', content: 'DK CAR'}]
    ];
    tags.forEach(function (t) {
      var el = window.parent.document.createElement(t[0]);
      for (var k in t[1]) el.setAttribute(k, t[1][k]);
      head.appendChild(el);
    });
  } catch (e) { /* 크로스오리진 등 예외는 무시 */ }
})();
</script>
""", height=0)

# 3. 세션 상태 레지스트리 저장소 선언 및 영속화 로직
#    - 배포(Streamlit Cloud 등): Firestore에 저장 → 서버 재시작·다중 사용자에도 예약 유지
#    - 로컬 개발: 자격증명이 없으면 자동으로 bookings.json 파일 방식으로 대체 동작
DB_FILE = "bookings.json"
COLLECTION = "bookings"
# 탑승 이력 아카이브: 도착완료(또는 이후 확장)된 예약을 월/일별 통계·엑셀 내보내기용으로 영속 보관.
#  현재 예약(bookings)은 완료·취소 시 삭제되므로, 이력 조회의 근거 데이터는 이 아카이브에만 남는다.
HISTORY_FILE = "history.json"
HISTORY_COLLECTION = "history"

# 베트남(UTC+7) 실시간 — 상단 시계·출발시간 기본값과 동일 기준. 서버가 UTC라도 현지시각으로 기록.
def now_vn():
    return datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7)))

# ─────────────────────────────────────────────────────────────
# 👤 '내 정보 기억' — 매일 아침 이름·출발지를 다시 타이핑하는 마찰 제거
#   저장 위치: 브라우저 쿠키(dk_profile). 서버에는 남기지 않는다(개인정보를 앱 DB에 쌓지 않음).
#   · 읽기: st.context.cookies (Streamlit 1.42+) — 요청과 함께 오므로 별도 왕복이 필요 없다.
#   · 쓰기: '신청 완료를 누른 순간'에만 1회성 컴포넌트로 기록한다.
#     ⚠️ 매 실행 도는 JS 브릿지에서는 절대 쓰지 않는다 — 관리자 '로그인 유지'에서 브릿지가 값을
#        되살려 재로그인 사고를 냈던 것과 같은 함정이다(0710~0711에 여러 번 겪음).
#   [벤치마킹] NGUYỄN THẾ ANH(AE Engineer) "Vision Inspection V2" — Session Persistence(Auto Save/Load)로
#              매일 반복되던 설정 입력 시간을 제거한 사례. 상세: docs/BENCHMARK.md
# ─────────────────────────────────────────────────────────────
import urllib.parse

PROFILE_COOKIE = "dk_profile"
PROFILE_MAX_NAME = 40      # 쿠키 비대·장난 입력 방지용 길이 상한
PROFILE_MAX_DEP = 60

def load_user_profile():
    """브라우저에 저장해 둔 '내 정보'(이름·기본 출발지)를 읽는다. 없거나 손상되면 빈 dict."""
    try:
        raw = st.context.cookies.get(PROFILE_COOKIE)
    except Exception:
        return {}
    if not raw:
        return {}
    try:
        data = json.loads(urllib.parse.unquote(raw))
    except Exception:
        return {}
    if not isinstance(data, dict):
        return {}
    return {
        "name": str(data.get("n", ""))[:PROFILE_MAX_NAME],
        "departure": str(data.get("d", ""))[:PROFILE_MAX_DEP],
    }

def _profile_cookie_script(payload=None):
    """쿠키 기록/삭제용 1회성 <script>. payload=None이면 삭제.
    값은 전부 퍼센트 인코딩하므로 따옴표·세미콜론이 섞여도 쿠키·JS 문자열이 깨지지 않는다."""
    if payload is None:
        body = PROFILE_COOKIE + "=; path=/; max-age=0; SameSite=Lax"
    else:
        enc = urllib.parse.quote(json.dumps(payload, ensure_ascii=False), safe="")
        body = PROFILE_COOKIE + "=" + enc + "; path=/; max-age=31536000; SameSite=Lax"
    return "<script>try{window.parent.document.cookie='" + body + "';}catch(e){}</script>"

_db_cache = "uninit"  # "uninit" | None(파일모드) | Firestore client

def _get_db():
    """Firestore 클라이언트를 1회 초기화해 반환. 자격증명이 없으면 None(파일 모드)."""
    global _db_cache
    if _db_cache != "uninit":
        return _db_cache
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore

        cred = None
        # 1) Streamlit Cloud 시크릿(st.secrets["firebase"])에 서비스 계정 키가 있으면 사용
        try:
            if "firebase" in st.secrets:
                cred = credentials.Certificate(dict(st.secrets["firebase"]))
        except Exception:
            cred = None
        # 2) 로컬에 serviceAccountKey.json 파일이 있으면 사용
        if cred is None and os.path.exists("serviceAccountKey.json"):
            cred = credentials.Certificate("serviceAccountKey.json")

        if cred is None:
            _db_cache = None  # 자격증명 없음 → 파일 모드
            return _db_cache

        if not firebase_admin._apps:
            firebase_admin.initialize_app(cred)
        _db_cache = firestore.client()
    except Exception:
        _db_cache = None  # firebase-admin 미설치/오류 시에도 안전하게 파일 모드로
    return _db_cache

def load_bookings():
    db = _get_db()
    if db is not None:
        try:
            bookings = {}
            for doc in db.collection(COLLECTION).stream():
                parts = doc.id.split("||")
                if len(parts) == 2:
                    car_name, seat_num = parts
                    bookings[(car_name, int(seat_num))] = doc.to_dict()
            return bookings
        except Exception:
            pass  # Firestore 오류 시 아래 파일 모드로 대체 시도
    # 파일 모드(로컬 개발)
    if os.path.exists(DB_FILE):
        try:
            with open(DB_FILE, "r", encoding="utf-8") as f:
                return _decode_bookings(json.load(f))
        except Exception:
            return {}
    return {}

def _decode_bookings(data):
    """{'차량||좌석': {...}} 형태를 {(차량, 좌석번호): {...}}로 변환.
    파일·백업·업로드 등 '외부에서 온 JSON'을 읽는 모든 경로가 이 한 곳을 쓴다.
    형식이 어긋난 항목은 통째로 버리지 않고 그 항목만 건너뛴다(일부 손상 시 나머지는 살린다)."""
    out = {}
    if not isinstance(data, dict):
        return out
    for key, val in data.items():
        parts = str(key).split("||")
        if len(parts) != 2 or not isinstance(val, dict):
            continue
        try:
            out[(parts[0], int(parts[1]))] = val
        except (TypeError, ValueError):
            continue
    return out

def _encode_bookings(bookings):
    """{(차량, 좌석): {...}} → {'차량||좌석': {...}} (저장·백업 공통 형식)."""
    return {f"{car_name}||{seat_num}": val for (car_name, seat_num), val in bookings.items()}

def save_bookings(bookings):
    """예약 저장. 성공하면 True.
    ⚠️ 실패를 조용히 삼키면 사용자는 저장된 줄 안다 → 실패 시 플래그를 세워 화면에 경고를 띄운다.
    [벤치마킹] Đào Văn Bảo "K-Pulse" — "Hiển thị cảnh báo khi có lỗi kèm nguyên nhân"(오류 시 경고 표시),
               "Thao tác xóa không có hiệu lực"(삭제가 반영 안 되던 문제). 상세: docs/BENCHMARK.md"""
    db = _get_db()
    desired = _encode_bookings(bookings)
    if db is not None:
        try:
            col = db.collection(COLLECTION)
            existing_ids = {doc.id for doc in col.stream()}
            batch = db.batch()
            # 추가/갱신
            for doc_id, val in desired.items():
                batch.set(col.document(doc_id), val)
            # 취소·초기화로 사라진 예약 삭제
            for doc_id in existing_ids - set(desired.keys()):
                batch.delete(col.document(doc_id))
            batch.commit()
            return True
        except Exception:
            pass  # Firestore 오류 시 아래 파일 모드로 대체 저장
    # 파일 모드(로컬 개발): 원자적 저장으로 동시 쓰기 중 파일 손상 방지
    try:
        tmp_file = DB_FILE + ".tmp"
        with open(tmp_file, "w", encoding="utf-8") as f:
            json.dump(desired, f, ensure_ascii=False, indent=4)
        os.replace(tmp_file, DB_FILE)
        return True
    except Exception:
        st.session_state.save_failed = True   # 다음 렌더에서 화면 상단에 경고 표시
        return False

def archive_booking(car_name, seat_num, info, status="완료"):
    """완료 처리된 예약 1건을 탑승 이력 아카이브에 적재(2단계 월/일별 통계·엑셀 조회 근거).
    현황판 예약과 달리 삭제되지 않고 누적된다. Firestore 우선, 실패 시 history.json 폴백."""
    record = {
        "created_at": info.get("created_at", ""),   # 신청일시(신청완료 클릭 시각)
        "car": car_name,
        "seat": seat_num,
        "name": info.get("name", ""),
        "date": info.get("date", ""),
        "departure": info.get("departure", ""),
        "destination": info.get("destination", ""),
        "time": info.get("time", ""),
        "arrive": info.get("arrive", ""),
        "status": status,
        "completed_at": now_vn().strftime("%Y-%m-%d %H:%M:%S"),
    }
    db = _get_db()
    if db is not None:
        try:
            db.collection(HISTORY_COLLECTION).add(record)
            return
        except Exception:
            pass  # Firestore 오류 시 파일 폴백
    # 파일 모드: history.json에 append(원자적 저장)
    try:
        history = []
        if os.path.exists(HISTORY_FILE):
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                history = json.load(f)
        history.append(record)
        tmp_file = HISTORY_FILE + ".tmp"
        with open(tmp_file, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=4)
        os.replace(tmp_file, HISTORY_FILE)
    except Exception:
        pass

def load_history():
    """탑승 이력 아카이브 전체를 리스트로 반환(월/일별 통계·엑셀 조회용). Firestore 우선, 실패 시 history.json."""
    db = _get_db()
    if db is not None:
        try:
            return [doc.to_dict() for doc in db.collection(HISTORY_COLLECTION).stream()]
        except Exception:
            pass
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data if isinstance(data, list) else []
        except Exception:
            return []
    return []

# ─────────────────────────────────────────────────────────────
# 💾 백업 · 복원 — 파괴적 동작 옆에는 반드시 되돌릴 길을 둔다
#   '전체 예약 초기화'는 PIN만 통과하면 그날 전 직원 배차를 한 번에 지우는데 복구 수단이 없었다.
#   · 초기화·복원 '직전' 상태를 스냅샷 1개로 보관 → 되돌리기 1회 제공
#   · 관리자 화면에서 원클릭 JSON 내보내기 / 파일로 복원
#   [벤치마킹] Hà Văn Lượng(PM Part Leader) "Concept 3D 시뮬레이션" —
#              "Export/Import JSON giúp sao lưu toàn bộ... nút Reset khôi phục ngay trạng thái ban đầu chỉ với 1 click"
#              (원클릭 JSON 백업/복원 + 원클릭 초기화). 상세: docs/BENCHMARK.md
# ─────────────────────────────────────────────────────────────
BACKUP_FILE = "bookings_backup.json"
BACKUP_COLLECTION = "backups"
BACKUP_DOC = "last_snapshot"

def save_snapshot(bookings, reason=""):
    """파괴적 동작 직전 상태를 스냅샷 1개로 덮어써 보관. 성공하면 True."""
    payload = {
        "at": now_vn().strftime("%Y-%m-%d %H:%M:%S"),
        "reason": reason,                 # reset / import
        "count": len(bookings),
        "data": _encode_bookings(bookings),
    }
    db = _get_db()
    if db is not None:
        try:
            db.collection(BACKUP_COLLECTION).document(BACKUP_DOC).set(payload)
            return True
        except Exception:
            pass
    try:
        tmp_file = BACKUP_FILE + ".tmp"
        with open(tmp_file, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        os.replace(tmp_file, BACKUP_FILE)
        return True
    except Exception:
        return False

def load_snapshot():
    """보관된 직전 상태 스냅샷을 반환. 없으면 빈 dict."""
    db = _get_db()
    if db is not None:
        try:
            doc = db.collection(BACKUP_COLLECTION).document(BACKUP_DOC).get()
            if doc.exists:
                return doc.to_dict() or {}
        except Exception:
            pass
    if os.path.exists(BACKUP_FILE):
        try:
            with open(BACKUP_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}
    return {}

# ─────────────────────────────────────────────────────────────
# 🧾 감사 로그(Audit Log) — 누가 언제 어떤 예약을 바꿨는지 기록
#   전 인원이 매일 쓰면 "내 예약이 사라졌다" 문의는 반드시 생긴다. 그때 답할 수 있는 유일한 근거다.
#   ⚠️ 탑승 이력(history)과 '분리된' 저장소를 쓴다.
#      엑셀 내보내기(excel_export_dialog)가 history를 status 필터 없이 전량 내보내므로,
#      같은 곳에 취소·수정 기록을 섞으면 월간 탑승 실적 엑셀이 오염된다.
#   · action은 언어와 무관한 고정 토큰(cancel/edit/done/reset)으로 저장하고 화면에서만 번역한다.
#   [벤치마킹] Đào Văn Bảo(FAE Leader) "K-Pulse" — 확산 계획의
#              "Nhật ký thao tác phục vụ kiểm toán (Audit Log)". 상세: docs/BENCHMARK.md
# ─────────────────────────────────────────────────────────────
AUDIT_FILE = "audit.json"
AUDIT_COLLECTION = "audit"
AUDIT_KEEP = 1000     # 파일 모드 보관 상한(오래된 것부터 버림) — 무한 증가 방지

def current_actor():
    """지금 조작하는 사람의 식별값. 관리자면 'ADMIN', '내 정보 기억'이 있으면 그 이름, 없으면 빈 문자열."""
    if st.session_state.get("admin_unlocked"):
        return "ADMIN"
    return (load_user_profile().get("name") or "").strip()

def log_action(action, car, seat, info=None, note=""):
    """예약 변경 1건을 감사 로그에 남긴다. 기록에 실패해도 예약 처리 자체는 막지 않는다."""
    info = info or {}
    rec = {
        "at": now_vn().strftime("%Y-%m-%d %H:%M:%S"),
        "action": action,                        # cancel / edit / done / reset
        "actor": current_actor(),                # 조작자(ADMIN 또는 기억된 이름, 없으면 "")
        "car": car, "seat": seat,
        "target": str(info.get("name", "")),     # 대상 예약의 신청자
        "date": str(info.get("date", "")),
        "note": note,
    }
    db = _get_db()
    if db is not None:
        try:
            db.collection(AUDIT_COLLECTION).add(rec)
            return
        except Exception:
            pass
    try:
        rows = []
        if os.path.exists(AUDIT_FILE):
            with open(AUDIT_FILE, "r", encoding="utf-8") as f:
                rows = json.load(f)
            if not isinstance(rows, list):
                rows = []
        rows.append(rec)
        rows = rows[-AUDIT_KEEP:]
        tmp_file = AUDIT_FILE + ".tmp"
        with open(tmp_file, "w", encoding="utf-8") as f:
            json.dump(rows, f, ensure_ascii=False, indent=2)
        os.replace(tmp_file, AUDIT_FILE)
    except Exception:
        pass

def load_audit(limit=30):
    """최근 감사 로그를 최신순으로 반환(관리자 화면 표시용)."""
    rows = []
    db = _get_db()
    if db is not None:
        try:
            rows = [d.to_dict() for d in db.collection(AUDIT_COLLECTION).stream()]
        except Exception:
            rows = []
    if not rows and os.path.exists(AUDIT_FILE):
        try:
            with open(AUDIT_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            rows = data if isinstance(data, list) else []
        except Exception:
            rows = []
    rows.sort(key=lambda r: str(r.get("at", "")), reverse=True)
    return rows[:limit]

# ─────────────────────────────────────────────────────────────
# ✅ 승인 워크플로우 — 신청(대기) → 승인 → 도착 완료
#   신청 = 즉시 확정이던 것을 관리자 승인 단계로 나눈다. 아침 배차는 실제로는 차량·기사 사정에 따라
#   조정이 필요하고, 지금은 그 조정이 앱 바깥(구두·메신저)에서만 일어나 화면과 실제가 어긋난다.
#   ⚠️ 기존 예약에는 status 필드가 없다 → '승인됨'으로 간주한다.
#      그렇지 않으면 배포 순간 기존 예약이 전부 '대기'로 바뀌어 아침에 혼란이 생긴다.
#   [벤치마킹] Larry Nguyen(FAE Manager) "창고 관리 SW" —
#              "thêm các chức năng approval, cảnh báo quá hạn trả"(승인 기능 + 기한 초과 경고).
#              상세: docs/BENCHMARK.md
# ─────────────────────────────────────────────────────────────
STATUS_PENDING = "pending"
STATUS_APPROVED = "approved"
URGENT_MINUTES = 30      # 출발까지 이 시간 이내인데 아직 대기면 '임박' 경고

def booking_status(info):
    """예약의 승인 상태. 필드가 없거나 알 수 없는 값이면 기존 예약으로 보고 '승인됨' 처리."""
    s = str((info or {}).get("status", "")).strip().lower()
    return STATUS_PENDING if s == STATUS_PENDING else STATUS_APPROVED

def _departure_dt(info):
    """예약의 출발 일시(date + time)를 datetime으로. 형식이 어긋나면 None."""
    d = str((info or {}).get("date", "")).strip()
    tm = str((info or {}).get("time", "")).strip()
    try:
        y, mo, dd = (int(x) for x in d.split("-"))
        hh, mi = (int(x) for x in tm.split(":")[:2])
        return datetime.datetime(y, mo, dd, hh % 24, mi % 60,
                                 tzinfo=datetime.timezone(datetime.timedelta(hours=7)))
    except Exception:
        return None

def pending_urgency(info):
    """승인 대기 건의 시급도 — 'over'(출발시각 지남) / 'soon'(임박) / ''(여유·승인됨)."""
    if booking_status(info) != STATUS_PENDING:
        return ""
    dep = _departure_dt(info)
    if dep is None:
        return ""
    left = (dep - now_vn()).total_seconds() / 60.0
    if left < 0:
        return "over"
    if left <= URGENT_MINUTES:
        return "soon"
    return ""

def pending_bookings():
    """승인 대기 예약을 출발이 급한 순(날짜·시간 오름차순)으로 반환."""
    items = [(k, v) for k, v in st.session_state.bookings.items()
             if booking_status(v) == STATUS_PENDING]
    return sorted(items, key=lambda kv: (str(kv[1].get("date", "")), str(kv[1].get("time", "")), kv[0][1]))

# ─────────────────────────────────────────────────────────────
# 🔐 예약 소유 확인 — 남의 예약을 실수로 수정·취소·완료 처리하는 사고 방지
#   전 인원이 매일 쓰면 오터치 한 번으로 남의 출근 배차가 사라진다. 그 앞에 확인 한 단계를 둔다.
#   ⚠️ 한계를 분명히: 이름은 카드에 그대로 보이므로 '악의'는 막지 못한다(실수를 막는 가드레일).
#      진짜 차단은 사용자별 로그인이 필요하며, 그때까지는 감사 로그가 사후 추적을 담당한다.
# ─────────────────────────────────────────────────────────────
def is_my_booking(info):
    """'내 정보 기억'에 저장된 이름과 예약자 이름이 같은가(앞뒤 공백·대소문자 무시)."""
    me = (load_user_profile().get("name") or "").strip().casefold()
    return bool(me) and me == str((info or {}).get("name", "")).strip().casefold()

def can_manage_booking(info):
    """이 예약을 수정·취소·도착완료 할 수 있는가. 관리자는 항상 허용."""
    return bool(st.session_state.get("admin_unlocked")) or is_my_booking(info)

def purge_stale_bookings():
    """출발날짜가 '오늘'(베트남 기준, 00:00~24:00)보다 이전인 미완료 예약을 자동 삭제.
    도착완료된 예약은 이미 bookings에서 빠져 이력에만 남으므로, 여기 남은 '지난 날짜' 예약 = 미완료 →
    익일 00:00부터 정리 대상. (Streamlit은 상시 실행이 아니므로 '앱이 다음에 열릴 때' 지난 예약을 청소한다.)"""
    today_str = now_vn().strftime("%Y-%m-%d")
    stale = []
    for key, info in list(st.session_state.bookings.items()):
        d = str((info or {}).get("date", "")).strip()
        # ISO(YYYY-MM-DD) 형식만 안전하게 사전식 비교. 형식이 다르거나 비어 있으면 삭제하지 않음(데이터 보존).
        if len(d) == 10 and d[4] == "-" and d[7] == "-" and d < today_str:
            stale.append(key)
    if stale:
        for key in stale:
            st.session_state.bookings.pop(key, None)
        save_bookings(st.session_state.bookings)
    return len(stale)

# 파일로부터 기존 예약 정보 상시 로딩
st.session_state.bookings = load_bookings()
# 출발날짜가 지난(미완료) 예약은 익일 00:00부터 자동 삭제(앱 로드 시 정리)
purge_stale_bookings()

if "duplicate_error_msg" not in st.session_state:
    st.session_state.duplicate_error_msg = None

if "editing_booking" not in st.session_state:
    st.session_state.editing_booking = None

# 각 차량별 현재 클릭 및 선택된 좌석의 백엔드 동기화 상태 저장소 셋업 (양방향 하이브리드 제어 핵심)
if "selected_seat_state" not in st.session_state:
    st.session_state.selected_seat_state = {
        "TAXI (4 SEAT)": "-- 선택 --",
        "TAXI (7 SEAT)": "-- 선택 --",
        "TOYOTA INNOVA (7 SEAT)": "-- 선택 --",
        "HYUNDAI SEDONA (6 SEAT)": "-- 선택 --",
        "VINFAST VF5 (4 SEAT)": "-- 선택 --"
    }

# ─────────────────────────────────────────────────────────────
# 🌐 다국어(i18n): 한국어 / Tiếng Việt / English 전환
#   · 언어는 배너의 라디오 위젯 값(lang_toggle)에서 매 실행 최상단에서 도출한다.
#   · 내부 상태 토큰("좌석 N", "-- 선택 --", "완료" 등)은 그대로 두고,
#     화면 표시만 t()/format_func로 번역해 로직 호환성을 유지한다.
#     → 언어를 바꿔도 저장된 예약·이력 데이터는 전혀 영향을 받지 않는다.
#   · 베트남어(vi): 매일 아침 실제 사용 인원 대다수가 베트남 직원이라 모국어가 필수.
#     [벤치마킹] Nguyễn Huy Hoàng(AE Staff) "Button Design 2.0" — 사내 도구는 VI/EN 이중언어가 기본이라는 전제.
#                컨테스트 플랫폼(AI Work Booster 2026)도 한국어·Tiếng Việt·English 3종 토글 제공.
#                상세: docs/BENCHMARK.md
# ─────────────────────────────────────────────────────────────
LANG_OPTIONS = {"한국어": "ko", "Tiếng Việt": "vi", "ENG": "en"}   # 표시 라벨 → 내부 언어 코드
lang = LANG_OPTIONS.get(st.session_state.get("lang_toggle"), "ko")

TR = {
    "ko": {
        "app_title": "DK CAR BOOKING SEAT",
        "subtitle": "대곤 비나 직원 여러분, 차량 신청은 카시트 배치도를 보고 빈 자리를 선택해 주세요!",
        "legend_empty": "빈 자리", "legend_booked": "예약됨", "legend_selected": "선택 중",
        "legend_driver": "운전석", "legend_drag": "· 예약된 좌석은 드래그해서 빈 자리로 옮길 수 있어요",
        "seat_driver": "운전석", "seat_n": "좌석 {n}",
        "badge_seats": "{n}인승", "taxi_4": "4인승", "taxi_7": "7인승", "taxi_count": "TAXI 대수",
        "seats_left": "{n}자리 있음",
        "select_ph": "-- 선택 --", "seat_select": "{car} 좌석 선택", "full": "❌ 만차 (잔여 좌석 없음)",
        "seatmap_title": "🚗 좌석 선택", "seatmap_hint": "빈 좌석을 클릭하면 차량 신청 창이 열립니다.",
        "dialog_title": "📝 차량 신청 정보 입력", "form_step_title": "📝 신청 정보 입력",
        "form_edit": "[{car}] 좌석 {seat} · 차량 예약 수정", "form_new": "[{car}] 좌석 {seat} · 차량 신청",
        "dup_error": "⚠️ 중복 신청 거부: [{name}]님은 이미 다른 차량에 배차되어 있습니다!",
        "f_name": "1. 신청자 이름", "f_name_ph": "예: 홍길동 PM",
        "f_dep": "2. 출발지", "f_dep_ph": "예: 본사 오피스",
        "f_dest": "3. 목적지 (위치)", "f_dest_ph": "예: 하노이 박닌 공장",
        "f_date": "4. 출발 날짜", "f_time": "5. 출발 시간", "f_arrive": "6. 도착 시간",
        "remember_me": "이 기기에 내 정보 기억 (다음부터 이름·출발지 자동 입력)",
        "btn_update": "수정 완료", "btn_submit": "신청 완료", "btn_cancel": "취소",
        "err_name_dest": "이름과 목적지를 정확히 입력해 주세요!",
        "toast_booked": "🎉 [{name}]님 좌석 {seat} 신청(수정) 완료!",
        "toast_moved": "🔄 [{name}]님의 예약이 [{car}] 좌석 {seat}(으)로 이동되었습니다!",
        "list_title": "📋 실시간 차량 예약 현황 · {n}건",
        "csv_btn": "📄 예약 이력",
        "search_ph": "🔍 신청자 이름 · 차량 · 목적지로 검색",
        "csv_headers": ["신청일시", "차량", "좌석", "신청자", "출발날짜", "출발지", "목적지", "출발시간", "도착시간"],
        "csv_file": "예약 이력_{date}.csv",
        "export_title": "📥 엑셀 데이터 내보내기", "export_year": "연도", "export_month": "월", "export_day": "일",
        "export_all": "전체", "export_btn": "⬇️ 엑셀 다운로드",
        "export_caption": "선택한 연도·월·일의 탑승 이력을 파이어베이스에서 조회하여 엑셀(XLSX) 파일로 내보냅니다.",
        "export_file": "탑승 이력_{ym}.xlsx", "export_empty": "선택한 기간에 해당하는 탑승 이력이 없습니다.",
        "arrive_title": "🏁 도착 완료 처리", "arrive_done": "완료",
        "arrive_desc": "[{car}] 좌석 {seat} · {name}\n도착 시간을 입력하고 완료를 누르면 탑승 이력에 기록됩니다.",
        "no_result": "🔍 [{q}] 검색 결과가 없습니다.",
        "c_applicant": "신청자:", "c_departure": "출발지:", "c_destination": "목적지:",
        "c_date": "출발날짜:", "c_time": "출발시간:", "c_arrive": "도착시간:", "edit_tip": "예약 수정하기",
        "btn_edit_bk": "예약 수정", "btn_cancel_bk": "예약 취소", "btn_done_bk": "도착 완료",
        "owner_warn": "⚠️ [{name}]님의 예약입니다. 본인이 맞으면 신청자 이름을 입력해 주세요.",
        "owner_ask": "신청자 이름 확인", "owner_ok": "확인",
        "owner_err": "이름이 일치하지 않습니다. 본인 예약만 변경할 수 있습니다.",
        "cancel_title": "🗑️ 예약 취소",
        "cancel_confirm": "[{car}] 좌석 {seat} · {name}\n이 예약을 취소할까요? 되돌릴 수 없습니다.",
        "cancel_yes": "네, 취소합니다",
        "toast_cancelled": "🗑️ [{name}]님 좌석 {seat} 예약이 취소되었습니다.",
        "audit_title": "🧾 최근 활동 기록", "audit_empty": "기록이 없습니다.",
        "audit_at": "일시", "audit_action": "동작", "audit_actor": "조작자", "audit_target": "대상",
        "audit_unknown": "(미확인)",
        "audit_act_cancel": "예약 취소", "audit_act_edit": "예약 수정",
        "audit_act_done": "도착 완료", "audit_act_reset": "전체 초기화",
        "toast_done": "🏁 [{name}]님 좌석 {seat} 도착 완료로 처리되었습니다.",
        "btn_reset_all": "🗑️ 전체 예약 초기화",
        "reset_warn": "⚠️ 지금 등록된 예약 {n}건을 모두 삭제합니다. 삭제 직전 상태는 자동 보관되어 '백업·복원'에서 되돌릴 수 있습니다.",
        "backup_title": "💾 백업 · 복원",
        "backup_export": "⬇️ 현재 예약 백업 내려받기 ({n}건)",
        "backup_file": "예약 백업_{date}.json",
        "backup_import": "백업 파일로 복원 (.json)",
        "backup_bad_file": "백업 파일을 읽을 수 없습니다. 이 앱에서 내려받은 JSON이 맞는지 확인해 주세요.",
        "backup_import_warn": "⚠️ 현재 예약 {cur}건이 백업 파일의 {n}건으로 교체됩니다. 교체 직전 상태는 자동 보관됩니다.",
        "backup_import_do": "네, 복원합니다",
        "backup_restored": "♻️ 예약 {n}건이 복원되었습니다.",
        "backup_undo": "↩️ 마지막 초기화·복원 되돌리기",
        "backup_snap_info": "보관된 직전 상태: {at} · {n}건",
        "backup_no_snap": "되돌릴 수 있는 직전 상태가 없습니다.",
        "save_failed": "⚠️ 저장에 실패했습니다. 방금 변경한 내용이 서버에 반영되지 않았을 수 있습니다. 새로고침 후 다시 확인해 주세요.",
        "audit_act_restore": "백업 복원", "audit_act_undo": "되돌리기", "audit_act_approve": "배차 승인",
        "status_pending": "승인 대기", "status_approved": "승인 완료",
        "status_soon": "출발 임박 · 미승인", "status_over": "출발 시각 초과 · 미승인",
        "approve_title": "✅ 승인 대기 ({n}건)", "approve_none": "승인 대기 중인 신청이 없습니다.",
        "approve_btn": "승인",
        "toast_approved": "✅ [{name}]님 좌석 {seat} 배차가 승인되었습니다.",
        "pending_banner": "⏳ 승인 대기 {n}건",
        "pending_banner_urgent": "⏳ 승인 대기 {n}건 · 이 중 출발 임박·초과 {u}건",
        "stats_title": "📊 기간 요약", "stats_total": "탑승 건수", "stats_cars": "이용 차량",
        "stats_top_dest": "최다 목적지", "stats_by_car": "차량별 탑승",
        "stats_top_dests": "목적지 TOP 5", "stats_by_hour": "출발 시간대", "stats_hour": "{h}시",
        "btn_reset_yes": "네, 전체 삭제", "toast_reset": "🧹 모든 예약이 초기화되었습니다.",
        "admin_title": "🔑 관리자 로그인", "admin_pw_label": "PASSWORD (숫자)", "admin_pw_ph": "****",
        "admin_hint": "0~9 숫자 4~8자리를 입력하세요.", "admin_ok": "확인",
        "admin_err": "PIN이 올바르지 않습니다.", "admin_unlocked_toast": "🔓 관리자 잠금이 해제되었습니다.",
        "admin_locked_out": "🚫 입력 실패가 많아 잠겼습니다. 페이지를 새로고침한 뒤 다시 시도하세요.",
        "admin_pin_unset": "⚠️ 관리자 PIN이 아직 Secrets에 설정되지 않아 기본값으로 동작 중입니다. "
                           "Streamlit Cloud → App settings → Secrets 에 [admin] pin_hash 를 등록하세요.",
        "admin_lock": "🔒 관리자 잠금", "admin_locked_toast": "🔒 관리자 잠금 상태로 돌아갔습니다.",
        "no_bookings": "접수된 배차 신청 내역이 없습니다.",
        "tip_from": "📍 출발: {v}", "tip_to": "🎯 목적지: {v}",
        "admin_status_title": "📋 좌석 신청 현황",
        "st_seat": "좌석 번호", "st_name": "신청자", "st_dep": "출발지",
        "st_dest": "목적지", "st_reqtime": "신청 시간", "st_deptime": "출발 시간", "st_arrive": "도착 시간",
        "st_empty": "미신청", "btn_close": "닫기", "btn_logout": "로그아웃",
        "admin_keep_login": "로그인 유지 (재접속 시 비밀번호 없이 자동 로그인)",
    },
    "vi": {
        "app_title": "DK CAR BOOKING SEAT",
        "subtitle": "Anh/chị CBNV DAEKHON VINA, vui lòng xem sơ đồ ghế và chọn chỗ trống để đăng ký xe!",
        "legend_empty": "Chỗ trống", "legend_booked": "Đã đặt", "legend_selected": "Đang chọn",
        "legend_driver": "Ghế lái", "legend_drag": "· Có thể kéo ghế đã đặt sang chỗ trống",
        "seat_driver": "Ghế lái", "seat_n": "Ghế {n}",
        "badge_seats": "{n} chỗ", "taxi_4": "4 chỗ", "taxi_7": "7 chỗ", "taxi_count": "Số xe TAXI",
        "seats_left": "Còn {n} chỗ",
        "select_ph": "-- Chọn --", "seat_select": "Chọn ghế {car}", "full": "❌ Hết chỗ",
        "seatmap_title": "🚗 Chọn ghế", "seatmap_hint": "Nhấn vào ghế trống để mở form đăng ký xe.",
        "dialog_title": "📝 Nhập thông tin đăng ký xe", "form_step_title": "📝 Nhập thông tin",
        "form_edit": "[{car}] Ghế {seat} · Sửa đăng ký", "form_new": "[{car}] Ghế {seat} · Đăng ký xe",
        "dup_error": "⚠️ Từ chối đăng ký trùng: [{name}] đã được xếp cho xe khác!",
        "f_name": "1. Tên người đăng ký", "f_name_ph": "VD: Nguyễn Văn A (PM)",
        "f_dep": "2. Điểm đi", "f_dep_ph": "VD: Văn phòng trụ sở",
        "f_dest": "3. Điểm đến (vị trí)", "f_dest_ph": "VD: Nhà máy Bắc Ninh, Hà Nội",
        "f_date": "4. Ngày đi", "f_time": "5. Giờ đi", "f_arrive": "6. Giờ đến",
        "remember_me": "Ghi nhớ thông tin trên thiết bị này (tự động điền tên·điểm đi lần sau)",
        "btn_update": "Cập nhật", "btn_submit": "Hoàn tất", "btn_cancel": "Hủy",
        "err_name_dest": "Vui lòng nhập chính xác tên và điểm đến!",
        "toast_booked": "🎉 [{name}] đã đăng ký ghế {seat} thành công!",
        "toast_moved": "🔄 Đăng ký của [{name}] đã chuyển sang [{car}] ghế {seat}!",
        "list_title": "📋 Tình trạng đặt xe theo thời gian thực · {n}",
        "csv_btn": "📄 Lịch sử đặt xe",
        "search_ph": "🔍 Tìm theo tên · xe · điểm đến",
        "csv_headers": ["Thời gian đăng ký", "Xe", "Ghế", "Người đăng ký", "Ngày đi", "Điểm đi", "Điểm đến", "Giờ đi", "Giờ đến"],
        "csv_file": "Lich su dat xe_{date}.csv",
        "export_title": "📥 Xuất dữ liệu Excel", "export_year": "Năm", "export_month": "Tháng", "export_day": "Ngày",
        "export_all": "Tất cả", "export_btn": "⬇️ Tải Excel",
        "export_caption": "Truy xuất lịch sử di chuyển của năm·tháng·ngày đã chọn từ Firebase và xuất ra file Excel (XLSX).",
        "export_file": "Lich su di chuyen_{ym}.xlsx", "export_empty": "Không có lịch sử di chuyển trong khoảng thời gian đã chọn.",
        "arrive_title": "🏁 Xử lý hoàn tất chuyến", "arrive_done": "Hoàn tất",
        "arrive_desc": "[{car}] Ghế {seat} · {name}\nNhập giờ đến rồi nhấn Hoàn tất để ghi vào lịch sử di chuyển.",
        "no_result": "🔍 Không có kết quả cho [{q}].",
        "c_applicant": "Người ĐK:", "c_departure": "Điểm đi:", "c_destination": "Điểm đến:",
        "c_date": "Ngày đi:", "c_time": "Giờ đi:", "c_arrive": "Giờ đến:", "edit_tip": "Sửa đăng ký",
        "btn_edit_bk": "Sửa ĐK", "btn_cancel_bk": "Hủy ĐK", "btn_done_bk": "Đã đến",
        "owner_warn": "⚠️ Đây là đăng ký của [{name}]. Nếu đúng là bạn, vui lòng nhập tên người đăng ký.",
        "owner_ask": "Xác nhận tên người đăng ký", "owner_ok": "Xác nhận",
        "owner_err": "Tên không khớp. Bạn chỉ có thể thay đổi đăng ký của chính mình.",
        "cancel_title": "🗑️ Hủy đăng ký",
        "cancel_confirm": "[{car}] Ghế {seat} · {name}\nBạn có muốn hủy đăng ký này không? Không thể hoàn tác.",
        "cancel_yes": "Vâng, hủy đăng ký",
        "toast_cancelled": "🗑️ Đã hủy đăng ký ghế {seat} của [{name}].",
        "audit_title": "🧾 Nhật ký thao tác gần đây", "audit_empty": "Chưa có ghi nhận nào.",
        "audit_at": "Thời gian", "audit_action": "Thao tác", "audit_actor": "Người thực hiện", "audit_target": "Đối tượng",
        "audit_unknown": "(chưa xác định)",
        "audit_act_cancel": "Hủy đăng ký", "audit_act_edit": "Sửa đăng ký",
        "audit_act_done": "Hoàn tất chuyến", "audit_act_reset": "Xóa toàn bộ",
        "toast_done": "🏁 [{name}] ghế {seat} đã được xử lý hoàn tất.",
        "btn_reset_all": "🗑️ Xóa toàn bộ đăng ký",
        "reset_warn": "⚠️ Sẽ xóa toàn bộ {n} đăng ký hiện có. Trạng thái ngay trước khi xóa được lưu tự động và có thể hoàn tác ở mục 'Sao lưu · Khôi phục'.",
        "backup_title": "💾 Sao lưu · Khôi phục",
        "backup_export": "⬇️ Tải bản sao lưu hiện tại ({n} đăng ký)",
        "backup_file": "Sao luu dat xe_{date}.json",
        "backup_import": "Khôi phục từ file sao lưu (.json)",
        "backup_bad_file": "Không đọc được file sao lưu. Vui lòng kiểm tra đây có đúng là file JSON tải từ ứng dụng này không.",
        "backup_import_warn": "⚠️ {cur} đăng ký hiện tại sẽ được thay bằng {n} đăng ký trong file sao lưu. Trạng thái trước khi thay được lưu tự động.",
        "backup_import_do": "Vâng, khôi phục",
        "backup_restored": "♻️ Đã khôi phục {n} đăng ký.",
        "backup_undo": "↩️ Hoàn tác lần xóa·khôi phục gần nhất",
        "backup_snap_info": "Trạng thái đã lưu: {at} · {n} đăng ký",
        "backup_no_snap": "Không có trạng thái nào để hoàn tác.",
        "save_failed": "⚠️ Lưu thất bại. Thay đổi vừa rồi có thể chưa được ghi lên máy chủ. Vui lòng tải lại trang và kiểm tra.",
        "audit_act_restore": "Khôi phục sao lưu", "audit_act_undo": "Hoàn tác", "audit_act_approve": "Duyệt xe",
        "status_pending": "Chờ duyệt", "status_approved": "Đã duyệt",
        "status_soon": "Sắp khởi hành · chưa duyệt", "status_over": "Quá giờ đi · chưa duyệt",
        "approve_title": "✅ Chờ duyệt ({n})", "approve_none": "Không có đăng ký nào đang chờ duyệt.",
        "approve_btn": "Duyệt",
        "toast_approved": "✅ Đã duyệt xe ghế {seat} cho [{name}].",
        "pending_banner": "⏳ {n} đăng ký đang chờ duyệt",
        "pending_banner_urgent": "⏳ {n} đăng ký chờ duyệt · trong đó {u} sắp/đã quá giờ đi",
        "stats_title": "📊 Tổng quan kỳ", "stats_total": "Số chuyến", "stats_cars": "Số xe sử dụng",
        "stats_top_dest": "Điểm đến nhiều nhất", "stats_by_car": "Chuyến theo xe",
        "stats_top_dests": "TOP 5 điểm đến", "stats_by_hour": "Khung giờ đi", "stats_hour": "{h}h",
        "btn_reset_yes": "Vâng, xóa tất cả", "toast_reset": "🧹 Toàn bộ đăng ký đã được xóa.",
        "admin_title": "🔑 Đăng nhập quản trị", "admin_pw_label": "PASSWORD (số)", "admin_pw_ph": "****",
        "admin_hint": "Nhập 4~8 chữ số (0-9).", "admin_ok": "Xác nhận",
        "admin_err": "PIN không đúng.", "admin_unlocked_toast": "🔓 Đã mở khóa quản trị.",
        "admin_locked_out": "🚫 Nhập sai quá nhiều lần nên đã bị khóa. Vui lòng tải lại trang và thử lại.",
        "admin_pin_unset": "⚠️ PIN quản trị chưa được đặt trong Secrets nên đang dùng giá trị mặc định. "
                           "Hãy thêm [admin] pin_hash tại Streamlit Cloud → App settings → Secrets.",
        "admin_lock": "🔒 Khóa quản trị", "admin_locked_toast": "🔒 Đã trở lại trạng thái khóa quản trị.",
        "no_bookings": "Chưa có đăng ký xe nào.",
        "tip_from": "📍 Đi: {v}", "tip_to": "🎯 Đến: {v}",
        "admin_status_title": "📋 Tình trạng đăng ký ghế",
        "st_seat": "Số ghế", "st_name": "Người ĐK", "st_dep": "Điểm đi",
        "st_dest": "Điểm đến", "st_reqtime": "Giờ đăng ký", "st_deptime": "Giờ đi", "st_arrive": "Giờ đến",
        "st_empty": "Chưa ĐK", "btn_close": "Đóng", "btn_logout": "Đăng xuất",
        "admin_keep_login": "Duy trì đăng nhập (tự động đăng nhập không cần mật khẩu khi quay lại)",
    },
    "en": {
        "app_title": "DK CAR BOOKING SEAT",
        "subtitle": "DAEKHON VINA staff — to request a vehicle, check the seat map and pick an empty seat!",
        "legend_empty": "Empty", "legend_booked": "Booked", "legend_selected": "Selecting",
        "legend_driver": "Driver", "legend_drag": "· Drag a booked seat to move it to an empty one",
        "seat_driver": "Driver", "seat_n": "Seat {n}",
        "badge_seats": "{n}-seater", "taxi_4": "4-Seat", "taxi_7": "7-Seat", "taxi_count": "TAXI count",
        "seats_left": "{n} SEAT LEFT",
        "select_ph": "-- Select --", "seat_select": "{car} seat select", "full": "❌ Full (no seats left)",
        "seatmap_title": "🚗 Select Seat", "seatmap_hint": "Click an empty seat to open the request form.",
        "dialog_title": "📝 Vehicle Request", "form_step_title": "📝 Request Info",
        "form_edit": "[{car}] Seat {seat} · Edit Request", "form_new": "[{car}] Seat {seat} · New Request",
        "dup_error": "⚠️ Duplicate rejected: [{name}] is already assigned to another vehicle!",
        "f_name": "1. Applicant name", "f_name_ph": "e.g. John Doe (PM)",
        "f_dep": "2. Departure", "f_dep_ph": "e.g. HQ Office",
        "f_dest": "3. Destination", "f_dest_ph": "e.g. Hanoi Bac Ninh Plant",
        "f_date": "4. Departure date", "f_time": "5. Departure time", "f_arrive": "6. Arrival time",
        "remember_me": "Remember me on this device (auto-fill name & departure next time)",
        "btn_update": "Update", "btn_submit": "Submit", "btn_cancel": "Cancel",
        "err_name_dest": "Please enter a valid name and destination!",
        "toast_booked": "🎉 [{name}] — seat {seat} request saved!",
        "toast_moved": "🔄 [{name}]'s booking moved to [{car}] seat {seat}!",
        "list_title": "📋 Live Seat Booking · {n}",
        "csv_btn": "📄 Booking History",
        "search_ph": "🔍 Search by name · vehicle · destination",
        "csv_headers": ["Requested At", "Car", "Seat", "Applicant", "Date", "Departure", "Destination", "Time", "Arrival"],
        "csv_file": "Booking History_{date}.csv",
        "export_title": "📥 Export Excel Data", "export_year": "Year", "export_month": "Month", "export_day": "Day",
        "export_all": "All", "export_btn": "⬇️ Download Excel",
        "export_caption": "Queries the ride history from Firebase for the selected year/month/day and exports an Excel (XLSX) file.",
        "export_file": "Ride History_{ym}.xlsx", "export_empty": "No ride history for the selected period.",
        "arrive_title": "🏁 Mark Arrival", "arrive_done": "Done",
        "arrive_desc": "[{car}] Seat {seat} · {name}\nEnter the arrival time and press Done to save it to the ride history.",
        "no_result": "🔍 No results for [{q}].",
        "c_applicant": "Applicant:", "c_departure": "Departure:", "c_destination": "Destination:",
        "c_date": "Date:", "c_time": "Time:", "c_arrive": "Arrival:", "edit_tip": "Edit booking",
        "btn_edit_bk": "Edit", "btn_cancel_bk": "Cancel", "btn_done_bk": "Arrived",
        "owner_warn": "⚠️ This booking belongs to [{name}]. If that is you, enter the applicant name.",
        "owner_ask": "Confirm applicant name", "owner_ok": "Confirm",
        "owner_err": "Name does not match. You can only change your own booking.",
        "cancel_title": "🗑️ Cancel Booking",
        "cancel_confirm": "[{car}] Seat {seat} · {name}\nCancel this booking? This cannot be undone.",
        "cancel_yes": "Yes, cancel it",
        "toast_cancelled": "🗑️ [{name}]'s seat {seat} booking has been cancelled.",
        "audit_title": "🧾 Recent Activity", "audit_empty": "No records yet.",
        "audit_at": "When", "audit_action": "Action", "audit_actor": "By", "audit_target": "Target",
        "audit_unknown": "(unknown)",
        "audit_act_cancel": "Cancelled", "audit_act_edit": "Edited",
        "audit_act_done": "Arrived", "audit_act_reset": "Reset all",
        "toast_done": "🏁 [{name}] — seat {seat} marked as arrived.",
        "btn_reset_all": "🗑️ Reset all bookings",
        "reset_warn": "⚠️ This deletes all {n} current bookings. The state right before deletion is saved automatically and can be undone in 'Backup · Restore'.",
        "backup_title": "💾 Backup · Restore",
        "backup_export": "⬇️ Download current backup ({n} bookings)",
        "backup_file": "Booking Backup_{date}.json",
        "backup_import": "Restore from a backup file (.json)",
        "backup_bad_file": "Could not read the backup file. Please check that it is a JSON file downloaded from this app.",
        "backup_import_warn": "⚠️ The current {cur} bookings will be replaced with {n} from the backup file. The state before replacing is saved automatically.",
        "backup_import_do": "Yes, restore",
        "backup_restored": "♻️ {n} bookings restored.",
        "backup_undo": "↩️ Undo the last reset / restore",
        "backup_snap_info": "Saved state: {at} · {n} bookings",
        "backup_no_snap": "No saved state available to undo.",
        "save_failed": "⚠️ Save failed. Your latest change may not have reached the server. Please reload and check again.",
        "audit_act_restore": "Restored backup", "audit_act_undo": "Undone", "audit_act_approve": "Approved",
        "status_pending": "Pending", "status_approved": "Approved",
        "status_soon": "Departing soon · not approved", "status_over": "Past departure · not approved",
        "approve_title": "✅ Pending approval ({n})", "approve_none": "No requests are waiting for approval.",
        "approve_btn": "Approve",
        "toast_approved": "✅ [{name}]'s seat {seat} has been approved.",
        "pending_banner": "⏳ {n} request(s) pending approval",
        "pending_banner_urgent": "⏳ {n} pending · {u} departing soon or overdue",
        "stats_title": "📊 Period summary", "stats_total": "Rides", "stats_cars": "Vehicles used",
        "stats_top_dest": "Top destination", "stats_by_car": "Rides by vehicle",
        "stats_top_dests": "Top 5 destinations", "stats_by_hour": "Departure hour", "stats_hour": "{h}h",
        "btn_reset_yes": "Yes, delete all", "toast_reset": "🧹 All bookings have been reset.",
        "admin_title": "🔑 Admin Login", "admin_pw_label": "PASSWORD (digits)", "admin_pw_ph": "****",
        "admin_hint": "Enter 4-8 digits (0-9).", "admin_ok": "OK",
        "admin_err": "Incorrect PIN.", "admin_unlocked_toast": "🔓 Admin unlocked.",
        "admin_locked_out": "🚫 Too many failed attempts. Reload the page and try again.",
        "admin_pin_unset": "⚠️ The admin PIN is not yet set in Secrets, so the default is still in use. "
                           "Add [admin] pin_hash under Streamlit Cloud → App settings → Secrets.",
        "admin_lock": "🔒 Lock admin", "admin_locked_toast": "🔒 Admin locked again.",
        "no_bookings": "No dispatch requests yet.",
        "tip_from": "📍 From: {v}", "tip_to": "🎯 To: {v}",
        "admin_status_title": "📋 Seat Request Status",
        "st_seat": "Seat", "st_name": "Applicant", "st_dep": "From",
        "st_dest": "To", "st_reqtime": "Requested", "st_deptime": "Departure", "st_arrive": "Arrival",
        "st_empty": "—", "btn_close": "Close", "btn_logout": "Logout",
        "admin_keep_login": "Keep me logged in (auto-login without password on return)",
    },
}

def t(key, **kw):
    """현재 언어(lang)의 번역 문자열 반환. {변수}가 있으면 kw로 포맷."""
    s = TR.get(lang, TR["ko"]).get(key, TR["ko"].get(key, key))
    if kw and isinstance(s, str):
        try:
            s = s.format(**kw)
        except Exception:
            pass
    return s

# 상단 배너 디자인 - 깔끔하게 버전 텍스트 및 실시간 자바스크립트 디지털시계 탑재
init_time_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# DAEKHON VINA 로고 이미지 로드 (아래 헤더에서 사용하므로 헤더보다 먼저 정의)
def _load_brand_logo_uri():
    """DAEKHON VINA 로고 이미지를 base64 data URI로 로드. 없으면 빈 문자열(→ 이모지 폴백)."""
    import base64, os
    base_dir = os.path.dirname(os.path.abspath(__file__))
    for fn, mime in (("DAEKHON VINA LOGO.png", "image/png"), ("daekhon_vina_logo.png", "image/png")):
        p = os.path.join(base_dir, fn)
        if os.path.exists(p):
            b64 = base64.b64encode(open(p, "rb").read()).decode()
            return f"data:{mime};base64,{b64}"
    return ""

DAEKHON_LOGO_URI = _load_brand_logo_uri()

# 좌측: 타이틀/버전/실시간 시계 배너  ·  우측: 언어 선택 토글(한국어/ENG)
_bn_l, _bn_r = st.columns([6, 2], vertical_alignment="center")
# 로고 파일이 있으면 이미지, 없으면 이모지 폴백
brand_mark_html = (f'<img class="brand-logo-img" src="{DAEKHON_LOGO_URI}" alt="DAEKHON VINA"/>'
                   if DAEKHON_LOGO_URI else '<span class="brand-mark">🐋</span>')
with _bn_l:
    # 로고+브랜드명(왼쪽) · 메인 타이틀+버전 묶음(가운데). 시계는 오른쪽 컬럼으로 이동.
    st.markdown(f"""
    <div class="top-header-container">
        <div class="brand-lockup">
            {brand_mark_html}
            <span class="brand-name">DAEKHON VINA</span>
        </div>
        <div class="title-group">
            <p class="main-title">{t("app_title")}</p>
            <span class="clean-timestamp-stamp brand-version">{date_version_str}</span>
        </div>
    </div>
    """, unsafe_allow_html=True)
with _bn_r:
    # 실시간 시계(위) + 언어 토글(아래)을 컴팩트 세로 스택으로 묶어 오른쪽 프레임 끝선에 정렬
    with st.container(key="hdr_right"):
        st.markdown(f'<div id="live-digital-clock" class="clean-timestamp-stamp header-clock">{init_time_str}</div>', unsafe_allow_html=True)
        st.radio("Language", list(LANG_OPTIONS.keys()), key="lang_toggle",
                 horizontal=True, label_visibility="collapsed")

st.markdown(f'<div class="sub-title">{t("subtitle")}</div>', unsafe_allow_html=True)

# 좌석 상태 색 — 배치도 좌석 테두리 / 범례 / 예약 카드 배지가 모두 이 두 상수를 참조한다(한 곳에서 관리).
#  ⚠️ 범례가 모듈 실행 시점에 바로 그려지므로, 상수는 반드시 범례보다 먼저 정의돼야 한다.
BOOKED_SEAT_LINE = "#40c057"    # 승인 완료(확정) — 초록 실선
# 승인 대기 — 주황 파선.
#  색 선택 근거(눈대중 아님): 검증 도구로 빈자리(#1c7ed6)·확정(#40c057)과 함께 검사한 결과
#  정상 색각 분리도는 충분(ΔE 27.1)하나, 적록 색각에서 확정과의 분리도가 6.6(deutan)로 경계 구간이다.
#  → 이 구간은 '보조 인코딩을 반드시 동반할 때만' 허용되므로 파선(stroke-dasharray) 테두리를 함께 쓰고,
#    범례·툴팁에도 상태 문구를 넣어 색을 못 읽어도 상태를 알 수 있게 했다.
#  (핑크 #e64980은 검사를 완전히 통과했지만 '대기'라는 의미 전달이 약해 채택하지 않음)
PENDING_SEAT_LINE = "#fd7e14"

# 좌석 색상 의미를 한눈에 알려주는 범례(legend) — 배치도 위 안내
st.markdown(f"""
<div style="display: flex; flex-direction: row; gap: 18px; align-items: center; flex-wrap: wrap; margin: -8px 0 14px 2px; font-size: 12px; color: #c7ccd6;">
    <span style="display:inline-flex; align-items:center; gap:6px;"><span style="width:12px; height:12px; border-radius:3px; background:#1e293b; border:2px solid #1c7ed6; display:inline-block;"></span>{t("legend_empty")}</span>
    <span style="display:inline-flex; align-items:center; gap:6px;"><span style="width:12px; height:12px; border-radius:3px; background:#1b3b22; border:2px solid {BOOKED_SEAT_LINE}; display:inline-block;"></span>{t("status_approved")}</span>
    <span style="display:inline-flex; align-items:center; gap:6px;"><span style="width:12px; height:12px; border-radius:3px; background:#3b2a12; border:2px dashed {PENDING_SEAT_LINE}; display:inline-block;"></span>{t("status_pending")}</span>
    <span style="display:inline-flex; align-items:center; gap:6px;"><span style="width:12px; height:12px; border-radius:3px; background:#3a2f15; border:2px solid #fab005; display:inline-block;"></span>{t("legend_selected")}</span>
    <span style="display:inline-flex; align-items:center; gap:6px;"><span style="width:12px; height:12px; border-radius:3px; background:#2c1a1a; border:2px solid #e03131; display:inline-block;"></span>{t("legend_driver")}</span>
</div>
""", unsafe_allow_html=True)

# ⚡ [OPT1 트리거 지원] 좌석 배치도 내부 프리미엄 가죽 시트 렌더러 (클릭 이벤트 주입)
def render_premium_seat(x, y, w, h, label, seat_id, car_display_name, is_driver=False, is_booked=False, tooltip="", sub_label="", admin_login=False, book_state=""):
    # 예약된 좌석의 테두리 실선/파선 — 승인 대기는 파선으로 그린다.
    #  ⚠️ 색만으로 구분하지 않는다: 적록 색각에서 초록(확정)↔주황(대기)은 구분이 약하다(deutan ΔE 6.6).
    #     파선은 흑백으로 봐도 구분되므로, 색이 안 보여도 상태를 읽을 수 있다.
    dash_attr = ""
    if is_driver:
        stroke_color = "#e03131"
        main_fill = "#2c1a1a"
        inner_fill = "#3b1e1e"
        text_color = "#ffffff" # 글자색 백색으로 통일
    elif is_booked and book_state == STATUS_PENDING:
        stroke_color = PENDING_SEAT_LINE   # 승인 대기 좌석은 주황색 + 파선 테두리
        main_fill = "#3b2a12"
        inner_fill = "#241606"
        text_color = "#ffffff"
        dash_attr = ' stroke-dasharray="3 2"'
    elif is_booked:
        stroke_color = BOOKED_SEAT_LINE # 신청 완료(승인) 좌석은 초록색(예약 카드 배지 배경과 동일)
        main_fill = "#1b3b22"
        inner_fill = "#0b2412"
        text_color = "#ffffff" # 글자색 백색으로 통일
    else:
        # 현재 마우스 클릭(OPT1) 혹은 토글 선택(OPT2)으로 타깃팅된 활성화 좌석 강조 컬러 바인딩
        if st.session_state.selected_seat_state.get(car_display_name) == f"좌석 {seat_id}":
            stroke_color = "#fab005" # 현재 선택 중인 좌석은 럭셔리 골드/옐로우로 구분
            main_fill = "#3a2f15"
            inner_fill = "#261e0b"
            text_color = "#ffffff" # 글자색 백색으로 통일
        else:
            stroke_color = "#1c7ed6"
            main_fill = "#1e293b"    
            inner_fill = "#0f172a"   
            text_color = "#ffffff" # 글자색 백색으로 통일

    svg = []
    
    # 드래그 앤 드롭용 클래스 및 데이터 속성 주입 (운전석 제외 전체 차량의 좌석 적용)
    drag_drop_attrs = ""
    if is_driver:
        if admin_login:
            # INNOVA·SEDONA 운전석: 클릭하면 관리자 로그인 팝업(JS가 숨김 ADMINLOGIN 버튼 대신 클릭)
            drag_drop_attrs = f' class="admin-login-seat" data-car="{car_display_name}"'
    else:
        if is_booked:
            # 이미 예약된 좌석: 드래그하여 이동시킬 출발지(Source)
            drag_drop_attrs = f' class="seat-draggable" data-car="{car_display_name}" data-seat="{seat_id}" draggable="true"'
        else:
            # 예약 안 된 빈 좌석: 클릭 선택(seat-clickable) + 드래그 이동 목적지(seat-droptarget)
            drag_drop_attrs = f' class="seat-droptarget seat-clickable" data-car="{car_display_name}" data-seat="{seat_id}"'

    svg.append(f'<g{drag_drop_attrs}>')

    # 예약된 좌석에 마우스를 올리면 뜨는 예약 현황(SVG 네이티브 <title> 툴팁)
    if tooltip:
        svg.append(f'<title>{esc(tooltip)}</title>')

    # ── 좌석 형상(첨부 이미지 형): 팔걸이(양쪽) + 시트 쿠션(아래) + 등받이(메인) ──
    aw = w * 0.20  # 팔걸이 폭
    svg.append(f'<rect class="clickable-seat-rect" x="{x:.1f}" y="{y+h*0.34:.1f}" width="{aw:.1f}" height="{h*0.52:.1f}" rx="{aw*0.5:.1f}" fill="{inner_fill}" stroke="{stroke_color}" stroke-width="1.3"{dash_attr} />')
    svg.append(f'<rect class="clickable-seat-rect" x="{x+w-aw:.1f}" y="{y+h*0.34:.1f}" width="{aw:.1f}" height="{h*0.52:.1f}" rx="{aw*0.5:.1f}" fill="{inner_fill}" stroke="{stroke_color}" stroke-width="1.3"{dash_attr} />')
    svg.append(f'<rect class="clickable-seat-rect" x="{x+w*0.13:.1f}" y="{y+h*0.58:.1f}" width="{w*0.74:.1f}" height="{h*0.40:.1f}" rx="{w*0.14:.1f}" fill="{inner_fill}" stroke="{stroke_color}" stroke-width="1.3"{dash_attr} />')
    svg.append(f'<rect class="clickable-seat-rect" x="{x+w*0.12:.1f}" y="{y:.1f}" width="{w*0.76:.1f}" height="{h*0.66:.1f}" rx="{w*0.26:.1f}" fill="{main_fill}" stroke="{stroke_color}" stroke-width="1.8"{dash_attr} />')

    cx = x + w/2
    # 글자 크기·textLength 계산은 '원본' 길이 기준(이스케이프하면 &amp; 처럼 길어져 크기가 틀어진다).
    #  화면에 넣을 때만 esc()로 감싼다 — 예약된 좌석의 label은 신청자가 입력한 이름이다.
    if is_driver or is_booked:
        # 운전석 / 예약자 이름: 등받이 중앙 한 줄(길면 자동 축소). 보조라벨(운전자명)은 아랫줄.
        fs = 8.0
        if len(label) > 3: fs = 6.5
        if len(label) > 5: fs = 5.5
        lattr = f' textLength="{w*0.66:.1f}" lengthAdjust="spacingAndGlyphs"' if len(label) >= 4 else ""
        if sub_label:
            svg.append(f'<text x="{cx:.1f}" y="{y+h*0.30:.1f}" font-family="sans-serif" font-size="{fs}" font-weight="bold" fill="{text_color}" text-anchor="middle"{lattr}>{esc(label)}</text>')
            svg.append(f'<text x="{cx:.1f}" y="{y+h*0.48:.1f}" font-family="sans-serif" font-size="6" font-weight="bold" fill="#fab005" text-anchor="middle">{esc(sub_label)}</text>')
        else:
            svg.append(f'<text x="{cx:.1f}" y="{y+h*0.40:.1f}" font-family="sans-serif" font-size="{fs}" font-weight="bold" fill="{text_color}" text-anchor="middle"{lattr}>{esc(label)}</text>')
    else:
        # 빈 좌석: "좌석"(윗줄, 작게) / 숫자(아랫줄, 크게) — label 예: "좌석 5" / "Seat 5"
        _p = label.rsplit(" ", 1)
        _word, _num = (_p[0], _p[1]) if len(_p) == 2 else (label, "")
        svg.append(f'<text x="{cx:.1f}" y="{y+h*0.27:.1f}" font-family="sans-serif" font-size="5.5" font-weight="bold" fill="{text_color}" text-anchor="middle">{esc(_word)}</text>')
        svg.append(f'<text x="{cx:.1f}" y="{y+h*0.52:.1f}" font-family="sans-serif" font-size="11" font-weight="bold" fill="{text_color}" text-anchor="middle">{esc(_num)}</text>')

    svg.append('</g>')
    return "".join(svg)

# ─────────────────────────────────────────────────────────────
# 🚗 차량 모델별 3D 상단뷰(Top-View) 섀시 렌더러
#   · 모델마다 실루엣(길이·폭·코너 라운딩), 도색, 후드/트렁크 비율이 다르다.
#   · 좌석(render_premium_seat)은 이 섀시 위에 그대로 얹혀 클릭·드래그가 유지된다.
#   · viewBox 0 0 160 250 좌표계 고정 → 기존 좌석 좌표와 완전 호환.
# ─────────────────────────────────────────────────────────────
#   hi=하이라이트 / base=기본도장 / lo=음영 / edge=외곽윤곽 / glass=유리색
CAR_MODELS = {
    # MPV(이노바): 길고 완만하게 둥근 실버 메탈릭
    "innova": {"x0": 16, "y0": 6,  "x1": 144, "y1": 244, "rf": 26, "rr": 20,
               "hi": "#f4f7fb", "base": "#c4cad4", "lo": "#7e8590", "edge": "#4c515c", "glass": "#1d2732", "taxi": False},
    # 미니밴(카니발/세도나): 각지고 넓은 어깨의 펄 화이트
    "sedona": {"x0": 13, "y0": 4,  "x1": 147, "y1": 246, "rf": 18, "rr": 14,
               "hi": "#ffffff", "base": "#e4e8ee", "lo": "#aeb4c0", "edge": "#6f7580", "glass": "#1d2732", "taxi": False},
    # 컴팩트 SUV(VF5): 좁고 단단한 실루엣의 빈패스트 블루 메탈릭
    "vf5":    {"x0": 20, "y0": 16, "x1": 140, "y1": 244, "rf": 24, "rr": 24,
               "hi": "#bfeaff", "base": "#4aa6d6", "lo": "#1f5f8c", "edge": "#123f5c", "glass": "#0d2735", "taxi": False},
    # 택시 4인승: 트렁크가 또렷한 세단형 옐로우 메탈릭
    "taxi4":  {"x0": 15, "y0": 8,  "x1": 145, "y1": 246, "rf": 24, "rr": 14,
               "hi": "#fff3b0", "base": "#ffcf1a", "lo": "#c99400", "edge": "#8a6400", "glass": "#1d2732", "taxi": True},
    # 택시 7인승: MPV형 옐로우 메탈릭
    "taxi7":  {"x0": 15, "y0": 6,  "x1": 145, "y1": 244, "rf": 26, "rr": 20,
               "hi": "#fff3b0", "base": "#ffcf1a", "lo": "#c99400", "edge": "#8a6400", "glass": "#1d2732", "taxi": True},
}

def _model_key(car_name):
    """표시명에서 실제 차량 모델 키를 도출한다."""
    n = car_name.upper()
    if "INNOVA" in n:
        return "innova"
    if "SEDONA" in n:
        return "sedona"
    if "VF5" in n or "VINFAST" in n:
        return "vf5"
    if "TAXI" in n:
        # 6·7인승은 MPV형(taxi7), 그 외(4인승)는 세단형(taxi4)
        return "taxi7" if ("7" in n or "6" in n) else "taxi4"
    return "innova"

def _body_path(x0, y0, x1, y1, rf, rr):
    """앞(위)코너는 rf, 뒤(아래)코너는 rr 라운딩을 적용한 차체 외곽 패스."""
    return (f"M {x0+rf} {y0} L {x1-rf} {y0} Q {x1} {y0} {x1} {y0+rf} "
            f"L {x1} {y1-rr} Q {x1} {y1} {x1-rr} {y1} L {x0+rr} {y1} "
            f"Q {x0} {y1} {x0} {y1-rr} L {x0} {y0+rf} Q {x0} {y0} {x0+rf} {y0} Z")

def _checker_row(xs, xe, y, size):
    """택시 식별용 체커(바둑판) 가로 밴드."""
    out, x, i = [], xs, 0
    while x < xe:
        c = "#111111" if i % 2 else "#f2f2f2"
        out.append(f'<rect x="{x}" y="{y}" width="{size}" height="{size}" fill="{c}"/>')
        x += size
        i += 1
    return "".join(out)

@st.cache_data(show_spinner=False)
def _load_car_image_uri():
    """전 차량 공통 외관 배경(세로 실사 상단뷰)을 base64 data URI로 로드.
    세로 원본 car_topview_src.png 우선(뷰박스 160x250 세로에 맞음)."""
    import base64, os
    base_dir = os.path.dirname(os.path.abspath(__file__))
    for fn, mime in (("car_topview_src.png", "image/png"), ("car_topview.jpg", "image/jpeg"), ("car_topview.png", "image/png")):
        p = os.path.join(base_dir, fn)
        if os.path.exists(p):
            b64 = base64.b64encode(open(p, "rb").read()).decode()
            return f"data:{mime};base64,{b64}"
    return ""

CAR_IMAGE_URI = _load_car_image_uri()

@st.cache_data(show_spinner=False)
def _load_taxi_logo_uri():
    """TAXI 브랜드 로고 이미지(taxi_logo.png 등)를 base64 data URI로 로드. 없으면 빈 문자열."""
    import base64, os
    base_dir = os.path.dirname(os.path.abspath(__file__))
    for fn, mime in (("taxi_logo.png", "image/png"), ("taxi_logo.jpg", "image/jpeg"),
                     ("taxi_logo.jpeg", "image/jpeg"), ("taxi_logo.webp", "image/webp")):
        p = os.path.join(base_dir, fn)
        if os.path.exists(p):
            b64 = base64.b64encode(open(p, "rb").read()).decode()
            return f"data:{mime};base64,{b64}"
    return ""

TAXI_LOGO_URI = _load_taxi_logo_uri()

def render_chassis(mk=None):
    """전 차량 공통 외관: 첨부된 실사 상단뷰 사진을 배경으로 깔고,
    사진에 이미 합성돼 있던 좌석을 가리는 불투명 실내 패널을 덮는다.
    실제 좌석(render_premium_seat)은 이 패널 위에 레이아웃별로 얹혀 클릭·드래그가 유지된다.
    (mk 인자는 하위호환용으로 받기만 하고 무시 — 모든 모델이 동일 외관을 사용)"""
    # 모델별 외관 색상: 공통 사진(어두운 청회색)에 CSS 컬러 필터를 씌워 첨부 색상에 근접
    #   이노바=화이트 / 세도나=블랙 / VF5=레드 / 택시=옐로우
    WHITE_FILTER = "grayscale(1) brightness(1.72) contrast(0.82)"
    YELLOW_FILTER = "grayscale(1) sepia(1) saturate(7) hue-rotate(5deg) brightness(1.15)"  # 택시 옐로우
    MODEL_FILTER = {
        "innova": WHITE_FILTER,   # 요청: 이노바 외관을 택시(흰색)와 동일하게
        "sedona": "grayscale(1) brightness(0.42) contrast(1.1)",
        "vf5":    "grayscale(1) sepia(1) saturate(6.5) hue-rotate(-40deg) brightness(1.02)",
        "taxi4":  YELLOW_FILTER,   # 요청: 택시 외관 노란색
        "taxi7":  YELLOW_FILTER,
    }
    img_filter = MODEL_FILTER.get(mk, "")
    s = []
    if CAR_IMAGE_URI:
        # viewBox(160x250)에 사진을 꽉 채워 좌석 좌표계와 정렬 (none = 완전 채움)
        style_attr = f' style="filter: {img_filter}"' if img_filter else ""
        s.append(f'<image href="{CAR_IMAGE_URI}" x="0" y="0" width="160" height="250" preserveAspectRatio="none"{style_attr}/>')
    else:
        s.append('<rect x="0" y="0" width="160" height="250" fill="#1b1e24"/>')
    # 사진 속 합성 좌석 영역을 덮는 실내 패널 (라이브 좌석 안착면)
    s.append('<rect x="29" y="96" width="107" height="146" rx="14" fill="#171a21" stroke="#3a4150" stroke-width="1"/>')
    s.append('<rect x="34" y="100" width="98" height="9" rx="4.5" fill="#20242e" opacity="0.9"/>')
    return "".join(s)

# (참고) 아래는 이전 SVG 손그림 섀시 렌더러 — 현재 미사용, 롤백 대비 보존
def _render_chassis_legacy_unused(mk):
    m = CAR_MODELS.get(mk, CAR_MODELS["innova"])
    x0, y0, x1, y1 = m["x0"], m["y0"], m["x1"], m["y1"]
    rf, rr = m["rf"], m["rr"]
    hi, base, lo, edge = m["hi"], m["base"], m["lo"], m["edge"]
    glass = m["glass"]
    cx = (x0 + x1) / 2
    ib = y1 - 14  # 실내 바닥 하단 경계 (좌석이 얹히는 영역)
    body = _body_path(x0, y0, x1, y1, rf, rr)
    s = []

    # ── 그라디언트/필터 정의 (모델별 고유 id로 SVG 간 충돌 방지) ──
    s.append('<defs>')
    # 금속 도장: 위→아래로 하늘반사·기본·음영·기본·하이라이트 5스톱
    s.append(f'<linearGradient id="paint_{mk}" x1="0" y1="0" x2="0" y2="1">'
             f'<stop offset="0%" stop-color="{hi}"/><stop offset="16%" stop-color="{base}"/>'
             f'<stop offset="50%" stop-color="{lo}"/><stop offset="84%" stop-color="{base}"/>'
             f'<stop offset="100%" stop-color="{hi}"/></linearGradient>')
    # 둥근 볼륨감: 좌우 가장자리만 어둡게
    s.append(f'<linearGradient id="round_{mk}" x1="0" y1="0" x2="1" y2="0">'
             f'<stop offset="0%" stop-color="#000" stop-opacity="0.5"/>'
             f'<stop offset="14%" stop-color="#000" stop-opacity="0"/>'
             f'<stop offset="86%" stop-color="#000" stop-opacity="0"/>'
             f'<stop offset="100%" stop-color="#000" stop-opacity="0.5"/></linearGradient>')
    # 유리: 대각 하늘반사 → 짙은 유리
    s.append(f'<linearGradient id="glass_{mk}" x1="0" y1="0" x2="1" y2="1">'
             f'<stop offset="0%" stop-color="#cfe0ec"/><stop offset="34%" stop-color="{glass}"/>'
             f'<stop offset="62%" stop-color="#0b1016"/><stop offset="100%" stop-color="{glass}"/></linearGradient>')
    # 스튜디오 바닥 (실사 렌더 배경)
    s.append(f'<radialGradient id="floor_{mk}" cx="50%" cy="40%" r="70%">'
             f'<stop offset="0%" stop-color="#2b3140"/><stop offset="100%" stop-color="#0c0e13"/></radialGradient>')
    # 부드러운 블러 (그림자·반사용)
    s.append(f'<filter id="blur_{mk}" x="-60%" y="-60%" width="220%" height="220%"><feGaussianBlur stdDeviation="3.4"/></filter>')
    s.append('</defs>')

    # ── 스튜디오 바닥 + 접지 그림자 ──
    s.append(f'<rect x="0" y="0" width="160" height="250" fill="url(#floor_{mk})"/>')
    s.append(f'<path d="{_body_path(x0+2, y0+11, x1+2, y1+13, rf, rr)}" fill="#000000" opacity="0.55" filter="url(#blur_{mk})"/>')

    # ── 타이어 4개 (차체 아래로 살짝 노출) ──
    for wx, wy in [(x0-3, 50), (x1-6, 50), (x0-3, y1-80), (x1-6, y1-80)]:
        s.append(f'<rect x="{wx}" y="{wy}" width="9" height="30" rx="4.5" fill="#050608"/>')

    # ── 차체 금속 도장 + 볼륨 음영 + 외곽 하이라이트 ──
    s.append(f'<path d="{body}" fill="url(#paint_{mk})"/>')
    s.append(f'<path d="{body}" fill="url(#round_{mk})"/>')
    s.append(f'<path d="{body}" fill="none" stroke="{edge}" stroke-width="1.4"/>')
    s.append(f'<path d="{body}" fill="none" stroke="{hi}" stroke-width="0.7" opacity="0.6"/>')

    # ── 정반사 하이라이트: 후드 하늘반사 + 대각 광택 밴드 ──
    s.append(f'<ellipse cx="{cx}" cy="{y0+18}" rx="{(x1-x0)/2-16}" ry="9" fill="#ffffff" opacity="0.20" filter="url(#blur_{mk})"/>')
    s.append(f'<path d="M {x0+16} {y0} L {x0+34} {y0} L {x0+14} {y1} L {x0} {y1} Z" fill="#ffffff" opacity="0.10" filter="url(#blur_{mk})"/>')

    # ── 사이드미러 (앞 어깨에서 돌출) ──
    s.append(f'<path d="M {x0} 62 l -7 -1 q -3 0 -3 3 l 0 3 q 0 3 3 2 l 7 -2 Z" fill="url(#paint_{mk})" stroke="{edge}" stroke-width="0.7"/>')
    s.append(f'<path d="M {x1} 62 l 7 -1 q 3 0 3 3 l 0 3 q 0 3 -3 2 l -7 -2 Z" fill="url(#paint_{mk})" stroke="{edge}" stroke-width="0.7"/>')

    # ── 헤드라이트 + 후드 캐릭터 라인 ──
    s.append(f'<rect x="{x0+9}" y="{y0+8}" width="15" height="7" rx="3.5" fill="#eef8ff" opacity="0.92"/>')
    s.append(f'<rect x="{x1-24}" y="{y0+8}" width="15" height="7" rx="3.5" fill="#eef8ff" opacity="0.92"/>')
    s.append(f'<path d="M {x0+30} {y0+9} Q 80 {y0+4} {x1-30} {y0+9}" fill="none" stroke="{edge}" stroke-width="0.9" opacity="0.4"/>')

    # ── 앞유리 (반사 하이라이트 포함) ──
    s.append(f'<path d="M {x0+30} 40 L {x1-30} 40 L {x1-18} 54 L {x0+18} 54 Z" fill="url(#glass_{mk})" stroke="{edge}" stroke-width="0.8"/>')
    s.append(f'<path d="M {x0+35} 41 L {x0+58} 41 L {x0+43} 53 L {x0+23} 53 Z" fill="#ffffff" opacity="0.16"/>')

    # ── 실내 바닥 (좌석 안착 영역) ──
    s.append(f'<rect x="{x0+7}" y="54" width="{x1-x0-14}" height="{ib-54}" rx="10" fill="#12141a" stroke="{edge}" stroke-width="0.9"/>')

    # ── 뒷유리 + 테일라이트 ──
    s.append(f'<path d="M {x0+30} {ib} L {x1-30} {ib} L {x1-18} {y1-3} L {x0+18} {y1-3} Z" fill="url(#glass_{mk})" stroke="{edge}" stroke-width="0.8"/>')
    s.append(f'<rect x="{x0+8}" y="{y1-11}" width="14" height="6" rx="2.5" fill="#ff5a5a" opacity="0.92"/>')
    s.append(f'<rect x="{x1-22}" y="{y1-11}" width="14" height="6" rx="2.5" fill="#ff5a5a" opacity="0.92"/>')

    # ── 스티어링 휠 (운전석 방향 표시) ──
    s.append('<circle cx="39" cy="62" r="6" fill="none" stroke="#6b7280" stroke-width="1.6"/>')
    s.append('<line x1="33" y1="62" x2="45" y2="62" stroke="#6b7280" stroke-width="1.3"/>')

    # ── 택시 전용: 지붕 사인 + 뒷범퍼 체커 밴드 ──
    if m["taxi"]:
        s.append('<rect x="66" y="28" width="28" height="12" rx="3" fill="#111111" stroke="#ffd21a" stroke-width="1.4"/>')
        s.append('<text x="80" y="37" font-family="sans-serif" font-size="8" font-weight="bold" fill="#ffd21a" text-anchor="middle">TAXI</text>')
        s.append(_checker_row(x0 + 22, x1 - 22, y1 - 10, 4))

    return "".join(s)

# 실사형 차량 배치도 섀시 렌더러 (지정 간격 수치 정밀 유지 버전)
def render_car_layout(car_name, layout_type, bookings):
    car_bookings = {}
    for (c_name, s_id), b_info in bookings.items():
        if c_name == car_name:
            car_bookings[s_id] = b_info

    svg = []
    svg.append('<svg width="100%" height="100%" viewBox="0 0 160 250" xmlns="http://www.w3.org/2000/svg">')

    # 모델별 3D 상단뷰 섀시 (실루엣·도색·후드/트렁크 비율이 차종마다 다름)
    svg.append(render_chassis(_model_key(car_name)))

    def get_seat_label(seat_id):
        if seat_id in car_bookings:
            return car_bookings[seat_id]["name"]
        return t("seat_n", n=seat_id)

    def get_seat_tip(seat_id):
        # 예약된 좌석 hover 시 보여줄 예약 현황 텍스트 (미예약이면 빈 문자열 → 툴팁 없음)
        if seat_id not in car_bookings:
            return ""
        info = car_bookings[seat_id]
        lines = [f"👤 {info.get('name', '')}"]
        if info.get("departure"):
            lines.append(t("tip_from", v=info["departure"]))
        lines.append(t("tip_to", v=info.get("destination", "")))
        dt = f"{info.get('date', '')} {info.get('time', '')}".strip()
        if dt:
            lines.append(f"🕒 {dt}")
        # 승인 상태를 '글자로도' 남긴다 — 색·파선을 못 읽는 경우에도 상태를 알 수 있어야 한다.
        urg = pending_urgency(info)
        if booking_status(info) == STATUS_APPROVED:
            lines.append("✅ " + t("status_approved"))
        else:
            lines.append("⏳ " + (t("status_over") if urg == "over"
                                  else t("status_soon") if urg == "soon"
                                  else t("status_pending")))
        return "\n".join(lines)

    def get_seat_state(seat_id):
        """좌석 배치도의 예약 상태 — 승인 대기면 주황·파선, 그 외(승인·기존 예약)는 초록 실선."""
        info = car_bookings.get(seat_id)
        if not info:
            return ""
        return STATUS_PENDING if booking_status(info) == STATUS_PENDING else STATUS_APPROVED

    # 좌석 배치: 운전석 + 인승별 승객석 좌표 (실사 사진 차실 x29~136 / y96~242에 맞춤)
    #   3열 X: 좌 35 / 중 66 / 우 97,  3행 Y: 앞 104 / 중 151 / 뒤 198,  좌석 32x32
    #   SW=좌석 폭(기존 32에서 10% 축소한 29 → 열 간격 유지 시 좌우 겹침 방지), SH=좌석 높이(유지)
    LX, MX, RX, R1, R2, R3, SW, SH = 35, 66, 97, 104, 151, 198, 29, 32
    seat_map = {
        "2-3-3": [(1, RX, R1), (2, LX, R2), (3, MX, R2), (4, RX, R2), (5, LX, R3), (6, MX, R3), (7, RX, R3)],
        "2-2-3": [(1, RX, R1), (2, LX, R2), (3, RX, R2), (4, LX, R3), (5, MX, R3), (6, RX, R3)],
        "2-3":   [(1, RX, R1), (2, LX, R3), (3, MX, R3), (4, RX, R3)],
        # TAXI 6인승: 앞(운전석+1)·중(3)·뒤(2), 뒤열은 좌/우로 벌려 배치
        "2-3-2": [(1, RX, R1), (2, LX, R2), (3, MX, R2), (4, RX, R2), (5, LX, R3), (6, RX, R3)],
    }
    if layout_type in seat_map:
        # 운전석 아래에 표시할 운전자 이름 — 메인 타일과 같은 CAR_INFO를 참조한다.
        #  (예전에는 여기에 이름을 따로 적어 둬, 차량이 바뀌어도 옛 이름이 남는 문제가 있었다)
        driver_name = car_info(car_name).get("driver", "")
        # 운전석: 이름이 있으면 'Driver' 아래 줄로 함께 박스 세로 중앙 정렬(이름은 골드 #fab005)
        #   INNOVA·SEDONA 운전석은 클릭 시 관리자 로그인 팝업이 뜨도록 admin_login=True
        _admin_car = ("INNOVA" in car_name) or ("SEDONA" in car_name)
        svg.append(render_premium_seat(LX, R1, SW, SH, t("seat_driver"), 0, car_name, is_driver=True, sub_label=driver_name, admin_login=_admin_car))
        if layout_type == "2-3":
            svg.append('  <line x1="33" y1="150" x2="129" y2="150" stroke="#3a4150" stroke-width="1" stroke-dasharray="3 3" />')
        for sid, sx, sy in seat_map[layout_type]:
            svg.append(render_premium_seat(sx, sy, SW, SH, get_seat_label(sid), sid, car_name,
                                           is_booked=(sid in car_bookings), tooltip=get_seat_tip(sid),
                                           book_state=get_seat_state(sid)))

        # 잔여 좌석 수 배지 — 운전석과의 간격을 좌석 행 간격만큼 벌리기 위해 y=70에 배치
        remaining = sum(1 for _sid, _sx, _sy in seat_map[layout_type] if _sid not in car_bookings)
        # 잔여 좌석 배지(자리 있음/seat left)는 분홍색, 만차일 때만 빨강
        rc_col = "#f783ac" if remaining > 0 else "#ff6b6b"
        rc_bg = "#2c1a24" if remaining > 0 else "#1f1111"
        rc_text = t("seats_left", n=remaining)
        svg.append(
            f'<g><rect x="26" y="70" width="54" height="13" rx="6.5" fill="{rc_bg}" '
            f'stroke="{rc_col}" stroke-width="1"/>'
            f'<text x="53" y="79.5" font-family="sans-serif" font-size="8" font-weight="bold" '
            f'fill="{rc_col}" text-anchor="middle">{rc_text}</text></g>'
        )

    svg.append('</svg>')
    return "".join(svg)

# ⚡ [클릭 및 드래그 파라미터 연동 파싱] SVG 앵커 링크 클릭 또는 드래그앤드롭 신호를 캐치해 처리
query_params = st.query_params

if "clear_error" in query_params:
    st.session_state.duplicate_error_msg = None
    st.query_params.clear()

elif "drag_src_car" in query_params and "drag_src_seat" in query_params and "drag_tgt_car" in query_params and "drag_tgt_seat" in query_params:
    try:
        src_car = query_params["drag_src_car"]
        src_seat = int(query_params["drag_src_seat"])
        tgt_car = query_params["drag_tgt_car"]
        tgt_seat = int(query_params["drag_tgt_seat"])
    except (ValueError, TypeError, KeyError):
        # 비정상 좌석 파라미터(잘못된 URL) 방어: 무시하고 주소창 초기화
        st.query_params.clear()
        st.rerun()
    else:
        src_key = (src_car, src_seat)
        tgt_key = (tgt_car, tgt_seat)

        # 출발지에 예약이 존재하고, 목적지 자리는 비어있을 때만 자리 이동 실행
        if src_key in st.session_state.bookings and tgt_key not in st.session_state.bookings:
            booking_info = st.session_state.bookings[src_key]
            st.session_state.bookings[tgt_key] = booking_info
            del st.session_state.bookings[src_key]

            # 파일 저장 및 Rerun 알림
            save_bookings(st.session_state.bookings)
            st.toast(t("toast_moved", name=booking_info['name'], car=tgt_car, seat=tgt_seat))

        st.query_params.clear()
        st.rerun()

elif "edit_car" in query_params and "edit_seat" in query_params:
    try:
        edit_car = query_params["edit_car"].replace("_", " ")
        edit_seat = int(query_params["edit_seat"])
    except (ValueError, TypeError, KeyError):
        # 비정상 좌석 파라미터(잘못된 URL) 방어: 무시하고 주소창 초기화
        st.query_params.clear()
        st.rerun()
    else:
        edit_key = (edit_car, edit_seat)

        if edit_key in st.session_state.bookings:
            info = st.session_state.bookings[edit_key]
            # 입력 필드의 세션 상태에 기존 예약 정보를 미리 로드 (구버전 데이터 방어를 위해 .get 사용)
            st.session_state.input_user_real_name = info.get("name", "")
            st.session_state.input_user_departure_loc = info.get("departure", "")
            st.session_state.input_user_destination_loc = info.get("destination", "")
            # 예약 날짜 필드도 함께 복원
            try:
                y, mo, d = map(int, info.get("date", "").split("-"))
                st.session_state.input_user_departure_date = datetime.date(y, mo, d)
            except Exception:
                pass
            try:
                h, m = map(int, info.get("time", "").split(":"))
                st.session_state.input_user_departure_time_tick = datetime.time(h, m)
            except Exception:
                pass
            try:
                h, m = map(int, info.get("arrive", "").split(":"))
                st.session_state.input_user_arrival_time_tick = datetime.time(h, m)
            except Exception:
                pass
            # 예약된 차량의 좌석 상태를 타깃팅하여 활성화
            st.session_state.selected_seat_state[edit_car] = f"좌석 {edit_seat}"
            st.session_state.editing_booking = edit_key
            st.session_state.duplicate_error_msg = None

        st.query_params.clear()
        st.rerun()

elif "car" in query_params and "seat" in query_params:
    incoming_car = query_params["car"].replace("_", " ")
    incoming_seat = query_params["seat"]
    # 다른 좌석 클릭 시 기존 중복 경고 메시지 즉시 리셋
    st.session_state.duplicate_error_msg = None
    # 하단 selectbox of 동기화를 위해 전역 상태 데이터 풀에 즉각 주입
    st.session_state.selected_seat_state[incoming_car] = f"좌석 {incoming_seat}"
    # 유격 싱크가 끝났으므로 URL 주소창 파라미터는 즉시 초기화하여 다중 재호출 Rerun 현상 전면 차단
    st.query_params.clear()

# 4. 차량 기본 구성 데이터 명세 수립
#   운용 차량: INNOVA / SEDONA / TAXI1 / TAXI2 (2×2 타일 한 화면에 딱 맞는 4대)
#   · VINFAST VF5는 운용에서 제외하고 그 자리를 TAXI 한 대로 대체했다.
#   · TAXI는 모두 6인승 2-3-2 동일 사양. 대수를 늘리려면 n_taxi만 올리면 된다(TAXI3… 자동 생성).
n_taxi = 2
cars_data = [
    {"name": "TOYOTA INNOVA", "layout": "2-3-3", "seats": 7},
    {"name": "HYUNDAI SEDONA", "layout": "2-2-3", "seats": 6},
]
for _ti in range(1, n_taxi + 1):
    cars_data.append({"name": "TAXI", "layout": "2-3-2", "seats": 6, "taxi_index": _ti})

total_cars = len(cars_data)  # 고정 2종 + TAXI n대

def brand_logo(name):
    """차량명 앞에 붙는 브랜드 로고(인라인 SVG). 공식 로고 파일 대신 식별 가능한 심볼로 근사."""
    n = name.upper()
    S = 'width="28" height="18" viewBox="0 0 28 18" style="vertical-align:middle;margin-right:7px"'
    if "TOYOTA" in n:
        # 토요타 엠블럼: 큰 타원 + 세로 타원 + 가로 타원
        return (f'<svg {S}>'
                '<ellipse cx="14" cy="9" rx="13" ry="8" fill="none" stroke="#EB0A1E" stroke-width="1.6"/>'
                '<ellipse cx="14" cy="7.6" rx="3.2" ry="5.4" fill="none" stroke="#EB0A1E" stroke-width="1.6"/>'
                '<ellipse cx="14" cy="6.4" rx="7.4" ry="2.6" fill="none" stroke="#EB0A1E" stroke-width="1.6"/></svg>')
    if "HYUNDAI" in n:
        # 현대: 타원 안 기울인 H
        return (f'<svg {S}>'
                '<ellipse cx="14" cy="9" rx="13" ry="8" fill="none" stroke="#9aa7b8" stroke-width="1.6"/>'
                '<text x="14" y="13.6" font-family="Georgia,serif" font-size="14" font-style="italic" font-weight="bold" fill="#9aa7b8" text-anchor="middle">H</text></svg>')
    if "VINFAST" in n or "VF5" in n:
        # 빈패스트: 볼드 V 심볼
        return (f'<svg {S}>'
                '<path d="M 3 3 L 13 16 L 23 3 L 18.3 3 L 13 10.5 L 7.7 3 Z" fill="#2f7bc4"/></svg>')
    if "TAXI" in n:
        # 택시: taxi_logo.png(첨부 이미지)가 있으면 그 로고를, 없으면 기존 옐로우 체커 SVG로 대체
        #   크기는 기존 로고 수준(높이 16px)으로 유지
        if TAXI_LOGO_URI:
            return (f'<img src="{TAXI_LOGO_URI}" alt="TAXI" '
                    f'style="height:16px;width:auto;vertical-align:middle;margin-right:7px"/>')
        return ('<svg width="26" height="16" viewBox="0 0 26 16" style="vertical-align:middle;margin-right:7px">'
                '<rect x="2" y="3" width="22" height="10" rx="2" fill="#f2c200"/>'
                '<rect x="2" y="3" width="3.6" height="3.3" fill="#111"/><rect x="9.2" y="3" width="3.6" height="3.3" fill="#111"/><rect x="16.4" y="3" width="3.6" height="3.3" fill="#111"/>'
                '<rect x="5.6" y="6.3" width="3.6" height="3.3" fill="#111"/><rect x="12.8" y="6.3" width="3.6" height="3.3" fill="#111"/><rect x="20" y="6.3" width="3.6" height="3.3" fill="#111"/></svg>')
    return '🚙 '

def _short_car_name(display_name):
    """예약 카드용 짧은 차량명: 브랜드 접두어와 '(N SEAT)' 접미어를 제거.
    'TOYOTA INNOVA (7 SEAT)'→'INNOVA', 'HYUNDAI SEDONA (6 SEAT)'→'SEDONA', 'TAXI2 (6 SEAT)'→'TAXI2'."""
    s = display_name
    if "(" in s:
        s = s[:s.rindex("(")].strip()
    for brand in ("TOYOTA", "HYUNDAI", "VINFAST", "KIA", "FORD"):
        if s.upper().startswith(brand + " "):
            s = s[len(brand) + 1:].strip()
            break
    return s

# 차량명 프레임 색: 각 차량 외관색 기준(배경 그라디언트 + 대비 텍스트 + 테두리)
CAR_FRAME_STYLE = {  # 배경·테두리는 각 외관색을 20% 어둡게(×0.8) 적용
    "innova": ("linear-gradient(180deg,#c5c6c7,#b1b3b7)", "#14171c", "#9ea2a7"),  # 화이트/실버
    "sedona": ("linear-gradient(180deg,#3d4048,#212328)", "#f2f4f7", "#454950"),  # 블랙
    "vf5":    ("linear-gradient(180deg,#b53a32,#991e17)", "#ffffff", "#7e1610"),  # 레드
    "taxi4":  ("linear-gradient(180deg,#ccab38,#c29204)", "#191b1f", "#a77b00"),  # 옐로우
    "taxi7":  ("linear-gradient(180deg,#ccab38,#c29204)", "#191b1f", "#a77b00"),
}

# ─────────────────────────────────────────────────────────────
# 🚘 차량별 운행 정보(운전자·차량번호·연락처) — 메인 타일과 좌석맵 운전석이 함께 참조한다.
#   키는 네비 라벨(TOYOTA INNOVA / HYUNDAI SEDONA / TAXI1 / TAXI2).
#   차량이 바뀌거나 기사가 교체되면 여기만 고치면 화면 전체에 반영된다.
# ─────────────────────────────────────────────────────────────
CAR_INFO = {
    "TOYOTA INNOVA":  {"driver": "TUAN", "plate": "98H 047 00", "phone": "0983.993.330"},
    "HYUNDAI SEDONA": {"driver": "SON",  "plate": "99A 667 46", "phone": "0977956965"},
    "TAXI1":          {"driver": "LUAN", "plate": "99E 002 46", "phone": "0972.631.361"},
    "TAXI2":          {"driver": "LUAN", "plate": "99E 002 46", "phone": "0972.631.361"},
}

def car_info(name):
    """표시명('TAXI1 (6 SEAT)')·네비 라벨('TAXI1') 어느 쪽으로 물어도 차량 정보를 돌려준다.
    긴 키부터 검사해 TAXI1/TAXI10 처럼 앞부분이 겹치는 이름이 섞이지 않게 한다."""
    n = (name or "").upper()
    for key in sorted(CAR_INFO, key=len, reverse=True):
        if n.startswith(key.upper()):
            return CAR_INFO[key]
    return {}

# 예약 현황 카드 배경색: 차량 로고 바탕색을 20% 알파 투명도로 틴트(어두운 페이지 위 은은한 차량색).
#  낮은 알파라 실효 배경은 어두워지므로 본문 글자색은 밝게. (배경, 텍스트색, 테두리)
#  단, INNOVA만 상단 네이밍 바와 동일한 '밝은 실버' 배경 + 어두운 글자로 예외 처리(세도나와 확실히 구분).
CAR_CARD_STYLE = {
    "innova": ("linear-gradient(180deg,#c9cbce,#b4b7bb)", "#14171c", "#9a9ea3"),  # 밝은 실버(상단 바와 동일 계열·어두운 글자)
    "sedona": ("rgba(120,128,142,0.20)", "#eef1f5", "rgba(120,128,142,0.55)"),  # 블랙(어두운 회색 유지 → INNOVA와 대비)
    "vf5":    ("rgba(214,72,62,0.22)",   "#ffffff", "rgba(214,72,62,0.60)"),    # 레드
    "taxi4":  ("rgba(232,192,70,0.22)",  "#f3f4f6", "rgba(232,192,70,0.60)"),   # 옐로우
    "taxi7":  ("rgba(232,192,70,0.22)",  "#f3f4f6", "rgba(232,192,70,0.60)"),
}

def car_title_frame(mk, inner_html):
    """차량명(로고+이름)을 외관색 배경의 가로 프레임으로 감싼 HTML을 반환.
    (좌석맵 팝업 헤더·배치도 위 제목용 — 메인 화면 선택 타일은 car_nav_tile를 쓴다)"""
    bg, fg, bd = CAR_FRAME_STYLE.get(mk, CAR_FRAME_STYLE["innova"])
    return (f'<div class="car-name-frame" style="background:{bg}; border:1px solid {bd};">'
            f'<span class="car-title-text" style="color:{fg};">{inner_html}</span>'
            f'</div>')

def car_nav_tile(mk, logo_html, label, idx):
    """메인 화면 차량 선택 타일(정사각형). 프레임 전체가 클릭 영역이다.
    가로로 긴 바는 탭 영역이 화면 폭 전체라 옆 차량까지 잘못 눌리기 쉬웠다(오클릭).
    → 정사각형으로 줄이고 2열로 배치해 타일 경계를 분명히 했다.
    class는 car-nav-click을 그대로 유지 — JS 브릿지(숨김 CARNAV 버튼 클릭)가 이 클래스에 걸려 있다."""
    bg, fg, bd = CAR_FRAME_STYLE.get(mk, CAR_FRAME_STYLE["innova"])
    # 운행 정보(운전자·차량번호·연락처)는 반투명 어두운 패널 위 흰 글자로 그린다.
    #  ⚠️ 타일 배경이 차량마다 실버·블랙·옐로우로 달라, '하나의 글자색'으로는 어디서나 읽히게 만들 수 없다.
    #     계산 결과 흰색·검정·연회색 모두 최저 명암비 1.7:1 수준으로 기준(4.5:1) 미달이었다.
    #     → 배경 위에 어두운 스크림(alpha 0.55)을 깔면 실효 배경이 통일되어 흰 글자가 최저 7.0:1로 안전하다.
    info = car_info(label)
    info_html = ""
    if info:
        info_html = (
            '<div class="car-nav-info">'
            f'<div class="cni-driver">{esc(info.get("driver", ""))}</div>'
            f'<div class="cni-line">{esc(info.get("plate", ""))}</div>'
            f'<div class="cni-line">{esc(info.get("phone", ""))}</div>'
            '</div>'
        )
    return (f'<div class="car-nav-click car-nav-tile" data-navidx="{idx}">'
            f'<div class="car-name-frame" style="background:{bg}; border:1px solid {bd};">'
            f'<span class="car-nav-logo">{logo_html}</span>'
            f'<span class="car-title-text" style="color:{fg};">{esc(label)}</span>'
            f'{info_html}'
            f'</div></div>')

# 5·6. 차량별 컬럼: 제목 + 인승 + 좌석 배치도를 한 컬럼에 묶어 렌더링한다.
#       (모바일에서 컬럼이 세로로 쌓여도 각 차량의 이름·인승이 자기 배치도 바로 위에 오도록 병합)

# 좌석 클릭 콜백: rerun 전에 실행되므로 selectbox 위젯 상태를 안전하게 동기화할 수 있다.
def on_seat_click(car_name, seat):
    seat_label = f"좌석 {seat}"
    # 처음 선택했던 같은 빈자리를 다시 누르면 → 선택 해제(빈자리로 리셋), 팝업도 닫힘
    if st.session_state.selected_seat_state.get(car_name) == seat_label:
        st.session_state.selected_seat_state[car_name] = "-- 선택 --"
        st.session_state[f"dropdown_trigger_spec_{car_name}"] = "-- 선택 --"
        st.session_state.editing_booking = None
    else:
        # 빈자리 선택 → 하단 selectbox 상태까지 동기화하고 신청 팝업 대상 차량 지정
        st.session_state.selected_seat_state[car_name] = seat_label
        st.session_state[f"dropdown_trigger_spec_{car_name}"] = seat_label
        st.session_state.active_booking_car = car_name
        st.session_state.editing_booking = None  # 새 예약(수정 아님)
        # 좌석맵 팝업(seatmap_car)은 그대로 두고 → 같은 팝업 안에서 신청 폼으로 전환된다(2개 dialog 전환 제약 회피)
        # 출발 시간 = 실시간(베트남 UTC+7) 기준 '가장 빨리 오는 5분 슬롯'으로 올림(step=5분과 정렬). 예: 19:02 → 19:05.
        _vn_now = now_vn()
        _slot = ((((_vn_now.hour * 60 + _vn_now.minute) + 4) // 5) * 5) % (24 * 60)
        st.session_state.input_user_departure_time_tick = datetime.time(_slot // 60, _slot % 60)
        st.session_state.input_user_arrival_time_tick = datetime.time(0, 0)
    st.session_state.duplicate_error_msg = None

# ── 전 차량 상태(표시명·인승·모델키·로고) 계산 (배치도 렌더와 분리) ──
#  resolved_cars는 아래 예약 현황판 컬럼에서도 쓰이므로 앱/웹 모드와 무관하게 항상 전 차량을 채운다.
resolved_cars = []
for car in cars_data:
    if car["name"] == "TAXI":
        # 1대뿐일 때도 번호를 붙인다(TAXI1, TAXI2 …) — 여러 대를 함께 쓰므로 항상 어느 차인지 구분돼야 한다
        ti = car["taxi_index"]
        nav_label = f"TAXI{ti}"
        prefix = nav_label
        seats_count = car["seats"]
        display_name = f"{prefix} ({seats_count} SEAT)"
        mk, logo_html = "taxi7", brand_logo("TAXI")
    else:
        nav_label = car["name"]
        seats_count = car["seats"]
        display_name = f"{car['name']} ({seats_count} SEAT)"
        mk, logo_html = _model_key(car["name"]), brand_logo(car["name"])
    st.session_state.selected_seat_state.setdefault(display_name, "-- 선택 --")
    resolved_cars.append({
        "display_name": display_name, "layout": car["layout"], "seats": seats_count,
        "is_taxi": car["name"] == "TAXI", "taxi_index": car.get("taxi_index"),
        "nav_label": nav_label, "mk": mk, "logo_html": logo_html,
    })

selected_seat_trigger = None

def _render_car_body(car_rc, show_name=True):
    """차량 1대: (선택) 이름 프레임 + 배치도 + 선택 트리거 + SEATSEL 숨김버튼.
    선택된 빈자리가 있으면 전역 selected_seat_trigger를 세팅한다."""
    global selected_seat_trigger
    if show_name:
        st.markdown(f'<div class="car-header-center">{car_title_frame(car_rc["mk"], car_rc["logo_html"] + car_rc["nav_label"])}</div>', unsafe_allow_html=True)
    # 좌석 배치도 본체
    st.markdown(f'<div class="car-layout-container">{render_car_layout(car_rc["display_name"], car_rc["layout"], st.session_state.bookings)}</div>', unsafe_allow_html=True)
    booked_seats = [s_id for (c_name, s_id) in st.session_state.bookings.keys() if c_name == car_rc["display_name"]]
    available_seats = [f"좌석 {seat}" for seat in range(1, car_rc["seats"] + 1) if seat not in booked_seats]
    # 좌석 선택은 배치도(SVG) 클릭만 사용. 클릭으로 세팅된 selected_seat_state를 읽어 팝업 트리거 구성.
    if available_seats:
        current_sel = st.session_state.selected_seat_state.get(car_rc["display_name"], "-- 선택 --")
        if current_sel != "-- 선택 --" and current_sel not in available_seats:
            current_sel = "-- 선택 --"
            st.session_state.selected_seat_state[car_rc["display_name"]] = "-- 선택 --"
        if current_sel != "-- 선택 --":
            seat_num = int(current_sel.split(" ")[1])
            selected_seat_trigger = (car_rc["display_name"], seat_num)
    else:
        st.error(t("full"))
    # ⚡ SVG 빈 좌석 클릭 시 JS가 대신 눌러줄 숨김 버튼(soft rerun)
    for seat in range(1, car_rc["seats"] + 1):
        if f"좌석 {seat}" in available_seats:
            st.button(
                f"SEATSEL::{car_rc['display_name']}::{seat}",
                key=f"seatsel_{car_rc['display_name']}_{seat}",
                on_click=on_seat_click,
                args=(car_rc["display_name"], seat),
            )

def owner_gate(car, seat, info):
    """값을 바꾸기 '직전'에 통과해야 하는 본인 확인 게이트.
    권한이 있거나 이번 세션에서 이미 확인했으면 True, 아니면 확인 UI를 그리고 False를 돌려준다.
    ⚠️ 버튼을 숨기는 방식이 아니다 — 수정·취소·완료의 각 실행 지점에서 매번 이 관문을 지난다.
    [벤치마킹] Đào Văn Bảo "K-Pulse" — 권한 검사는 서버에서 수행하고, 버튼을 숨기거나
               비활성화한 것에 의존하지 않는다는 원칙. 상세: docs/BENCHMARK.md"""
    if can_manage_booking(info):
        return True
    verified = st.session_state.setdefault("verified_bookings", set())
    if (car, seat) in verified:
        return True
    st.warning(t("owner_warn", name=info.get("name", "")))
    typed = st.text_input(t("owner_ask"), key=f"ownerchk_{car}_{seat}")
    if st.button(t("owner_ok"), type="primary", key=f"ownerbtn_{car}_{seat}", use_container_width=True):
        if typed.strip().casefold() == str(info.get("name", "")).strip().casefold():
            verified.add((car, seat))
            st.rerun()
        else:
            st.error(t("owner_err"))
    return False

def _status_chip(info):
    """예약 카드 헤더에 붙는 승인 상태 배지 HTML.
    대기=호박색, 임박(30분 이내)=주황, 출발시각 지남=빨강, 승인=초록."""
    urg = pending_urgency(info)
    if booking_status(info) == STATUS_APPROVED:
        bg, fg, label = "rgba(64,192,87,0.18)", "#63b365", t("status_approved")
    elif urg == "over":
        bg, fg, label = "rgba(224,49,49,0.22)", "#ff8787", t("status_over")
    elif urg == "soon":
        bg, fg, label = "rgba(253,126,20,0.22)", "#ffa94d", t("status_soon")
    else:
        bg, fg, label = "rgba(250,176,5,0.18)", "#fab005", t("status_pending")
    return (f'<div style="margin-top:3px;"><span style="display:inline-block; background:{bg}; '
            f'color:{fg}; border:1px solid {fg}; border-radius:4px; padding:0 5px; '
            f'font-size:10px; font-weight:700; white-space:nowrap;">{esc(label)}</span></div>')


def _render_pending_approvals():
    """관리자 전용 '승인 대기' 패널 — 출발이 급한 순으로 나열하고 한 건씩 승인한다.
    카드의 버튼 3개(수정·취소·도착완료)는 이미 폭이 빠듯해 4번째 버튼을 넣을 자리가 없으므로,
    승인은 카드가 아니라 이 패널에서 처리한다(0713에 카드 폭 문제로 11번 수정한 전례)."""
    items = pending_bookings()
    with st.expander(t("approve_title", n=len(items)), expanded=bool(items)):
        if not items:
            st.caption(t("approve_none"))
            return
        for (pc_name, pseat), pinfo in items:
            urg = pending_urgency(pinfo)
            mark = "🔴" if urg == "over" else ("🟠" if urg == "soon" else "🟡")
            col_l, col_r = st.columns([3, 1], vertical_alignment="center")
            with col_l:
                st.markdown(
                    f'<div style="font-size:12px; color:#e9ecef; line-height:1.45;">'
                    f'{mark} <strong>{esc(pinfo.get("name", ""))}</strong> · '
                    f'{esc(_short_car_name(pc_name))} {esc(t("seat_n", n=pseat))}<br>'
                    f'<span style="color:#adb5bd;">{esc(pinfo.get("date", ""))} {esc(pinfo.get("time", ""))} '
                    f'→ {esc(pinfo.get("destination", ""))}</span></div>',
                    unsafe_allow_html=True,
                )
            with col_r:
                if st.button(t("approve_btn"), key=f"approve_{pc_name}_{pseat}",
                             type="primary", use_container_width=True):
                    cur = st.session_state.bookings.get((pc_name, pseat))
                    if cur:   # 승인 직전 재확인 — 그 사이 취소·완료됐을 수 있다
                        cur["status"] = STATUS_APPROVED
                        if save_bookings(st.session_state.bookings):
                            log_action("approve", pc_name, pseat, cur)
                            st.toast(t("toast_approved", name=cur.get("name", ""), seat=pseat))
                    st.rerun()


def _render_backup_tools():
    """관리자 전용 백업·복원 — ① 원클릭 내보내기 ② 파일로 복원 ③ 마지막 초기화/복원 되돌리기.
    복원·되돌리기도 파괴적이므로, 실행 직전에 현재 상태를 다시 스냅샷으로 남긴다."""
    with st.expander(t("backup_title")):
        cur_n = len(st.session_state.bookings)
        # ① 내보내기 — 지금 예약 전체를 JSON 파일로
        st.download_button(
            t("backup_export", n=cur_n),
            data=json.dumps({
                "exported_at": now_vn().strftime("%Y-%m-%d %H:%M:%S"),
                "count": cur_n,
                "data": _encode_bookings(st.session_state.bookings),
            }, ensure_ascii=False, indent=2).encode("utf-8"),
            file_name=t("backup_file", date=now_vn().strftime("%Y%m%d_%H%M")),
            mime="application/json", use_container_width=True, key="backup_export_btn",
        )

        # ② 복원 — 내려받은 백업 파일을 올려 현재 예약을 통째로 교체
        st.divider()
        up = st.file_uploader(t("backup_import"), type=["json"], key="backup_import_file")
        if up is not None:
            incoming = None
            try:
                raw = json.loads(up.getvalue().decode("utf-8"))
                # 이 앱이 내보낸 형식({"data": {...}})과 예약 dict 자체 둘 다 받아준다
                incoming = _decode_bookings(raw.get("data") if isinstance(raw, dict) and "data" in raw else raw)
            except Exception:
                incoming = None
            if not incoming:
                st.error(t("backup_bad_file"))
            else:
                st.warning(t("backup_import_warn", n=len(incoming), cur=cur_n))
                if st.button(t("backup_import_do"), type="primary",
                             key="backup_import_btn", use_container_width=True):
                    save_snapshot(st.session_state.bookings, reason="import")   # 교체 전 상태 보관
                    st.session_state.bookings = incoming
                    if save_bookings(st.session_state.bookings):
                        log_action("restore", "", 0, None, note=str(len(incoming)))
                        st.toast(t("backup_restored", n=len(incoming)))
                        st.rerun()

        # ③ 되돌리기 — 마지막 초기화/복원 직전 상태로
        st.divider()
        snap = load_snapshot()
        snap_data = snap.get("data") if isinstance(snap, dict) else None
        if snap_data:
            st.caption(t("backup_snap_info", at=snap.get("at", ""), n=snap.get("count", 0)))
            if st.button(t("backup_undo"), key="backup_undo_btn", use_container_width=True):
                restored = _decode_bookings(snap_data)
                save_snapshot(st.session_state.bookings, reason="undo")   # 되돌리기도 되돌릴 수 있게
                st.session_state.bookings = restored
                if save_bookings(st.session_state.bookings):
                    log_action("undo", "", 0, None, note=str(len(restored)))
                    st.toast(t("backup_restored", n=len(restored)))
                    st.rerun()
        else:
            st.caption(t("backup_no_snap"))


def _render_audit_log():
    """관리자 전용 '최근 활동 기록' — 누가 언제 어떤 예약을 취소·수정·완료했는지 보여준다.
    좌석 신청 현황표(.seat-status-table)의 스타일을 그대로 재사용한다."""
    rows = load_audit(30)
    with st.expander(t("audit_title")):
        if not rows:
            st.caption(t("audit_empty"))
            return
        act_label = {
            "cancel": t("audit_act_cancel"), "edit": t("audit_act_edit"),
            "done": t("audit_act_done"), "reset": t("audit_act_reset"),
            "restore": t("audit_act_restore"), "undo": t("audit_act_undo"),
            "approve": t("audit_act_approve"),
        }
        body = []
        for r in rows:
            actor = str(r.get("actor", "")).strip() or t("audit_unknown")
            seat = r.get("seat") or 0
            where = f'{_short_car_name(str(r.get("car", "")))} {t("seat_n", n=seat)}' if seat else "—"
            target = str(r.get("target", "")).strip() or "—"
            if r.get("note"):
                target += f' ({r["note"]})'
            act = str(r.get("action", ""))
            body.append(
                f'<tr><td class="ss-time">{esc(r.get("at", ""))}</td>'
                f'<td>{esc(act_label.get(act, act))}</td>'
                f'<td>{esc(actor)}</td>'
                f'<td>{esc(target)}</td>'
                f'<td>{esc(where)}</td></tr>'
            )
        st.markdown(
            '<div class="seat-status-wrap"><table class="seat-status-table"><thead><tr>'
            f'<th>{t("audit_at")}</th><th>{t("audit_action")}</th><th>{t("audit_actor")}</th>'
            f'<th>{t("audit_target")}</th><th>{t("st_seat")}</th>'
            f'</tr></thead><tbody>{"".join(body)}</tbody></table></div>',
            unsafe_allow_html=True,
        )


def _booking_form(car_target, seat_target):
    """차량 신청/수정 입력 폼 본체 — 웹 신청 팝업(booking_dialog)과 앱 좌석맵 팝업에서 공용.
    완료/취소 시 앱 좌석맵 팝업(seatmap_car)도 함께 닫는다."""
    # 남의 예약을 수정하려는 경우 → 폼 대신 본인 확인을 먼저 그린다(같은 팝업 안에서 내용 전환).
    #  신규 신청은 대상이 없으므로 게이트 없음.
    _edit_key = st.session_state.editing_booking
    if _edit_key:
        _einfo = st.session_state.bookings.get(_edit_key)
        if _einfo and not owner_gate(_edit_key[0], _edit_key[1], _einfo):
            return

    form_title = t("form_edit", car=car_target, seat=seat_target) if st.session_state.editing_booking else t("form_new", car=car_target, seat=seat_target)
    # 이 차량의 메인 네이밍 바 색(배경 그라디언트/대비 텍스트/테두리)을 그대로 가져와 팝업에도 적용
    _mk = next((c["mk"] for c in resolved_cars if c["display_name"] == car_target), "innova")
    _fbg, _ffg, _fbd = CAR_FRAME_STYLE.get(_mk, CAR_FRAME_STYLE["innova"])
    # 좌석 타이틀 = 메인 차량 네이밍 바처럼 '배지형'(배경=차량색 / 글자=대비색 / 테두리=차량색) → 4종 모두 가독성 확보
    st.markdown(f"""
    <div style="margin-bottom: 12px;">
        <span style="display: inline-block; background: {_fbg}; color: {_ffg}; border: 1px solid {_fbd}; font-size: 14px; font-weight: 700; padding: 5px 12px; border-radius: 8px; box-shadow: 0 2px 6px rgba(0,0,0,0.35);">{form_title}</span>
    </div>
    """, unsafe_allow_html=True)
    # 신청 완료 버튼 배경 = 메인 차량 네이밍 배경색(대비 텍스트/테두리 동일 적용)
    st.markdown(
        f'<style>.st-key-submit_booking_form_btn button {{ background: {_fbg} !important; color: {_ffg} !important; border: 1px solid {_fbd} !important; }}</style>',
        unsafe_allow_html=True,
    )

    if st.session_state.duplicate_error_msg:
        st.markdown(f"""
        <div class="custom-error-box">
            <span class="custom-error-text">{esc(st.session_state.duplicate_error_msg)}</span>
        </div>
        <div style="margin-bottom: 10px;"></div>
        """, unsafe_allow_html=True)

    # '내 정보 기억'으로 저장해 둔 값이 있으면 이름·출발지를 미리 채운다.
    #   · 위젯 생성 '전에' session_state를 세팅해야 기본값으로 들어간다.
    #   · 수정 모드(editing_booking)이거나 이미 입력값이 있으면 건드리지 않는다 → 남의 예약을 덮어쓰지 않음.
    _prof = load_user_profile()
    if _prof and not st.session_state.editing_booking:
        if not st.session_state.get("input_user_real_name"):
            st.session_state.input_user_real_name = _prof.get("name", "")
        if not st.session_state.get("input_user_departure_loc"):
            st.session_state.input_user_departure_loc = _prof.get("departure", "")

    u_name = st.text_input(t("f_name"), placeholder=t("f_name_ph"), key="input_user_real_name")
    u_dep = st.text_input(t("f_dep"), placeholder=t("f_dep_ph"), key="input_user_departure_loc")
    u_dest = st.text_input(t("f_dest"), placeholder=t("f_dest_ph"), key="input_user_destination_loc")
    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        u_date = st.date_input(t("f_date"), key="input_user_departure_date")
    with fc2:
        u_time = st.time_input(t("f_time"), step=300, key="input_user_departure_time_tick")
    with fc3:
        u_arrive = st.time_input(t("f_arrive"), step=300, key="input_user_arrival_time_tick")

    # 공용 PC를 함께 쓰는 경우를 위해 저장 여부를 사용자가 직접 고른다(끄면 저장된 값도 즉시 삭제).
    remember_me = st.checkbox(t("remember_me"), value=True, key="remember_me_cb")

    act_col1, act_col2 = st.columns(2)
    with act_col1:
        btn_label = t("btn_update") if st.session_state.editing_booking else t("btn_submit")
        if st.button(btn_label, type="primary", key="submit_booking_form_btn", use_container_width=True):
            if u_name and u_name.strip() and u_dest and u_dest.strip():
                is_duplicate = False
                for booked_car_seat, booked_info in st.session_state.bookings.items():
                    if st.session_state.editing_booking:
                        if booked_car_seat == st.session_state.editing_booking:
                            continue
                    if booked_info["name"].strip() == u_name.strip():
                        is_duplicate = True
                        break
                if is_duplicate:
                    st.session_state.duplicate_error_msg = t("dup_error", name=u_name)
                    st.rerun()
                else:
                    time_str = u_time.strftime("%H:%M")
                    arrive_str = u_arrive.strftime("%H:%M") if u_arrive else ""
                    date_str = u_date.strftime("%Y-%m-%d") if u_date else datetime.date.today().strftime("%Y-%m-%d")
                    created_at = now_vn().strftime("%Y-%m-%d %H:%M:%S")
                    if st.session_state.editing_booking:
                        old_key = st.session_state.editing_booking
                        if old_key in st.session_state.bookings:
                            prev = st.session_state.bookings[old_key]
                            if prev.get("created_at"):
                                created_at = prev["created_at"]
                            # 값을 바꾸기 직전 권한 재확인(버튼 노출이 아니라 실행 지점에서 검사)
                            if not (can_manage_booking(prev)
                                    or old_key in st.session_state.get("verified_bookings", set())):
                                st.error(t("owner_err"))
                                st.stop()
                            log_action("edit", old_key[0], old_key[1], prev,
                                       note=f"→ {car_target} / {t('seat_n', n=seat_target)}")
                            del st.session_state.bookings[old_key]
                        st.session_state.editing_booking = None
                    st.session_state.bookings[(car_target, seat_target)] = {
                        "name": u_name.strip(),
                        "departure": u_dep.strip() if u_dep else "",
                        "destination": u_dest.strip(),
                        "date": date_str,
                        "time": time_str,
                        "arrive": arrive_str,
                        "created_at": created_at,
                        # 신청은 '승인 대기'로 시작한다. 내용을 수정하면(차량·시간·목적지 변경) 다시 대기로 돌아간다
                        #  — 승인은 '그 내용'에 대한 승인이므로, 내용이 바뀌면 승인도 다시 받아야 한다.
                        "status": STATUS_PENDING,
                    }
                    st.session_state.duplicate_error_msg = None
                    save_bookings(st.session_state.bookings)
                    # '내 정보 기억' — 예약이 실제로 저장된 뒤에만 반영. 체크 해제 시에는 기존 쿠키를 지운다.
                    #  (기록은 아래 8-c의 1회성 컴포넌트가 담당 — 상시 브릿지는 쿠키를 건드리지 않는다)
                    if remember_me:
                        st.session_state.profile_save = {
                            "n": u_name.strip()[:PROFILE_MAX_NAME],
                            "d": (u_dep.strip() if u_dep else "")[:PROFILE_MAX_DEP],
                        }
                    else:
                        st.session_state.profile_clear = True
                    st.session_state.selected_seat_state[car_target] = "-- 선택 --"
                    st.session_state.seatmap_car = None   # 앱 좌석맵 팝업도 함께 닫힘
                    st.toast(t("toast_booked", name=u_name, seat=seat_target))
                    st.rerun()
            else:
                st.error(t("err_name_dest"))
    with act_col2:
        if st.button(t("btn_cancel"), key="cancel_booking_dialog_btn", use_container_width=True):
            st.session_state.selected_seat_state[car_target] = "-- 선택 --"
            st.session_state[f"dropdown_trigger_spec_{car_target}"] = "-- 선택 --"
            st.session_state.editing_booking = None
            st.session_state.duplicate_error_msg = None
            st.session_state.seatmap_car = None   # 앱 좌석맵 팝업도 함께 닫힘
            st.rerun()

def _close_seatmap():
    # X·바깥클릭·ESC로 닫을 때: seatmap_car만 지우면 선택 좌석이 남아 top-level에서 booking_dialog가
    # 다시 열려 '안 닫힘'처럼 보인다 → 선택 좌석·수정상태·에러까지 함께 리셋해 완전히 닫는다.
    car = st.session_state.get("seatmap_car")
    if car and car in st.session_state.get("selected_seat_state", {}):
        st.session_state.selected_seat_state[car] = "-- 선택 --"
        st.session_state[f"dropdown_trigger_spec_{car}"] = "-- 선택 --"
    st.session_state.editing_booking = None
    st.session_state.duplicate_error_msg = None
    st.session_state.admin_login_open = False   # 관리자 로그인 폼도 함께 닫힘
    st.session_state.admin_pin_error = False
    st.session_state.admin_seat_status_open = False  # 좌석 신청 현황 표도 함께 닫힘
    st.session_state.seatmap_car = None

def _open_seatmap(display_name):
    # on_click 콜백 → 위젯 생성 전에 상태 세팅 → 단일 rerun에서 바로 팝업 오픈(이중 rerun 제거로 반응 속도 개선)
    st.session_state.seatmap_car = display_name

def _open_admin_login():
    # INNOVA·SEDONA 운전석 클릭 콜백.
    #   · 이미 관리자 로그인 상태(로그인 유지 복원 포함)면 → 로그인 폼을 건너뛰고 '좌석 신청 현황'을 바로 연다.
    #     (재로그인 무한 요구 방지: 로그인돼 있으면 비밀번호를 다시 묻지 않는다)
    #   · 아직 로그인 전이면 → 관리자 로그인 폼을 띄운다.
    if st.session_state.get("admin_unlocked"):
        st.session_state.admin_seat_status_open = True
        st.session_state.admin_login_open = False
    else:
        st.session_state.admin_login_open = True
        st.session_state.admin_pin_error = False

def _restore_admin():
    # '로그인 유지' 체크 후 재접속 시 → localStorage 흔적을 보고 JS가 대신 눌러 비밀번호 없이 관리자 복원
    st.session_state.admin_unlocked = True
    st.session_state.admin_keep = True   # 복원 후에도 유지 상태 지속
    st.session_state.admin_clear_ls = False   # 복원 = 로그인 상태이므로 삭제 상태 해제
    # 복원은 localStorage가 이미 '1'일 때만 발동하므로 별도 저장(admin_save_ls) 불필요

def _admin_login_form():
    """관리자 로그인 폼 — 좌석맵 팝업 안에서 표시. 입력 PIN의 해시가 Secrets의 pin_hash와 같으면 잠금 해제.
    PIN은 저장하지 않고 매번 입력(공용 PIN 1개). 연속 실패가 ADMIN_MAX_TRIES를 넘으면 이 세션에서 잠근다."""
    st.markdown(f'<div class="dlg-step-title">{t("admin_title")}</div>', unsafe_allow_html=True)
    st.caption(t("admin_hint"))
    # Secrets 미설정이면 PIN이 여전히 소스 기본값으로 동작 중 → 관리자에게 그 사실을 드러낸다.
    if not _admin_pin_config()[2]:
        st.warning(t("admin_pin_unset"))
    pin = st.text_input(t("admin_pw_label"), placeholder=t("admin_pw_ph"), type="password",
                        max_chars=8, key="admin_pin_input")
    if st.session_state.get("admin_pin_error"):
        st.error(t("admin_err"))
    tries = int(st.session_state.get("admin_pin_tries", 0))
    locked = tries >= ADMIN_MAX_TRIES
    if locked:
        st.error(t("admin_locked_out"))
    # 팝업 하단 '로그인 유지' 선택 — 체크 시 재접속해도 비밀번호 없이 자동 로그인(localStorage 저장)
    keep_login = st.checkbox(t("admin_keep_login"), key="admin_keep_login_cb")
    ac1, ac2 = st.columns(2)
    with ac1:
        if st.button(t("admin_ok"), type="primary", key="admin_pin_ok_btn",
                     use_container_width=True, disabled=locked):
            # 해시 비교로 검증(원문 PIN은 소스·세션 어디에도 남기지 않는다)
            if verify_admin_pin(pin):
                st.session_state.admin_unlocked = True
                st.session_state.admin_login_open = False
                st.session_state.admin_pin_error = False
                st.session_state.admin_pin_tries = 0
                # '로그인 유지' 체크 여부 저장
                st.session_state.admin_keep = bool(keep_login)
                # localStorage 저장/삭제는 '전환 시 1회성 컴포넌트'로만 수행(브릿지는 저장 안 함 → 재설정 사고 원천 차단).
                if keep_login:
                    st.session_state.admin_save_ls = True     # 유지 체크 → localStorage='1' 1회 저장 예약
                    st.session_state.admin_clear_ls = False
                else:
                    st.session_state.admin_clear_ls = True    # 미체크 → localStorage 삭제(유지 상태 아님)
                    st.session_state.admin_save_ls = False
                # 로그인 성공 → 같은 팝업 안에서 '좌석 신청 현황' 표로 전환(운전석 제외 전체 좌석)
                st.session_state.admin_seat_status_open = True
                st.toast(t("admin_unlocked_toast"))
                st.rerun()
            else:
                st.session_state.admin_pin_error = True
                st.session_state.admin_pin_tries = tries + 1   # 연속 실패 누적 → 한도 초과 시 잠금
                st.rerun()
    with ac2:
        if st.button(t("btn_cancel"), key="admin_pin_cancel_btn", use_container_width=True):
            st.session_state.admin_login_open = False
            st.session_state.admin_pin_error = False
            st.session_state.seatmap_car = None
            st.rerun()

# 좌석 신청 현황 팝업에 삽입되는 '현재 위치 지도'(무료 OpenStreetMap/Leaflet) — Google 지도 API 키 불필요.
#   GPS 현재위치 버튼 + 주소 검색창 + 지도 클릭 마커. 타일·지오코딩(Nominatim) 모두 무료(워터마크·비용 없음).
#   다크 테마(#0e1117/#1b1f27/#fab005/#38bdf8)로 팝업과 통일. Leaflet은 CDN(unpkg)에서 로드.
_ADMIN_LOCATION_MAP_HTML = """
<!DOCTYPE html>
<html>
<head>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<style>
  html,body{margin:0;padding:0;background:#0e1117;
    font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,'Noto Sans KR',sans-serif;}
  .map-tools{display:flex;flex-wrap:wrap;gap:6px;margin:0 0 6px 0;}
  #addr{flex:1 1 100%;min-width:0;background:#1b1f27;border:1px solid #3a3f4a;border-radius:8px;
    color:#e9ecef;padding:8px 10px;font-size:13px;outline:none;}
  #addr::placeholder{color:#6c757d;}
  #addr:focus{border-color:#38bdf8;}
  .mbtn{flex:1 1 auto;white-space:nowrap;background:#1b1f27;border:1px solid #3a3f4a;border-radius:8px;
    color:#fab005;font-weight:700;font-size:13px;padding:8px 10px;cursor:pointer;}
  .mbtn:hover{background:#242a33;border-color:#4a5160;}
  .mbtn.gps{color:#38bdf8;}
  .mbtn.gg{color:#63b365;}
  #map{width:100%;height:230px;border-radius:10px;border:1px solid #2b2f38;}
  #readout{margin-top:6px;font-size:12px;color:#adb5bd;min-height:16px;line-height:1.45;
    word-break:break-word;}
  #readout b{color:#e9ecef;}
  .leaflet-container{background:#0e1117;}
</style>
</head>
<body>
  <div class="map-tools">
    <input id="addr" type="text" placeholder="주소·상호 검색 후 Enter" />
    <button class="mbtn" id="searchBtn">검색</button>
    <button class="mbtn gps" id="gpsBtn">📍 현재 위치</button>
    <button class="mbtn gg" id="googleBtn">🔗 구글 검색</button>
  </div>
  <div id="map"></div>
  <div id="readout">주소·지명은 지도 검색으로, 가게·회사 상호는 '🔗 구글 검색'으로 찾으세요.</div>
  <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
  <script>
    var DEFAULT = [10.8231, 106.6297]; // 기본 중심(호치민) — GPS/검색 전 초기 화면
    var map = L.map('map', {zoomControl:true}).setView(DEFAULT, 12);
    L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png',
      {maxZoom:19, attribution:'© OpenStreetMap'}).addTo(map);
    var marker = null;
    var myPos = null; // GPS로 잡은 현재 위치([위도,경도]) — 검색 근접 정렬의 기준점
    var readout = document.getElementById('readout');
    function setMarker(lat, lng, label, km){
      if(marker){ marker.setLatLng([lat,lng]); }
      else { marker = L.marker([lat,lng]).addTo(map); }
      var parts = [];
      if(label){ parts.push('<b>'+label+'</b>'); }
      if(typeof km === 'number' && isFinite(km)){
        parts.push('현재 위치에서 약 ' + (km < 1 ? Math.round(km*1000)+'m' : km.toFixed(1)+'km'));
      }
      readout.innerHTML = parts.length ? parts.join('<br>') : '위치를 지정했습니다.';
    }
    // 역지오코딩 — 마커 위치의 주소를 표시(무료 Nominatim)
    function reverse(lat, lng){
      fetch('https://nominatim.openstreetmap.org/reverse?format=json&lat='+lat+'&lon='+lng)
        .then(function(r){ return r.json(); })
        .then(function(d){ if(d && d.display_name){ setMarker(lat,lng,d.display_name); } })
        .catch(function(){});
    }
    // 지도 클릭 → 마커 이동 + 주소 조회
    map.on('click', function(e){
      setMarker(e.latlng.lat, e.latlng.lng, null);
      reverse(e.latlng.lat, e.latlng.lng);
    });
    // 검색 기준점 — GPS로 잡은 현재 위치가 있으면 그 좌표, 없으면 현재 지도 중심.
    function refPoint(){
      if(myPos){ return myPos; }
      var c = map.getCenter();
      return [c.lat, c.lng];
    }
    // 두 좌표 간 거리(km, 하버사인) — 검색 결과를 현재 위치에서 가까운 순으로 정렬하는 데 사용.
    function distKm(a, b){
      var R = 6371, rad = Math.PI/180;
      var dLat = (b[0]-a[0])*rad, dLon = (b[1]-a[1])*rad;
      var s = Math.sin(dLat/2)*Math.sin(dLat/2)
            + Math.cos(a[0]*rad)*Math.cos(b[0]*rad)*Math.sin(dLon/2)*Math.sin(dLon/2);
      return R*2*Math.atan2(Math.sqrt(s), Math.sqrt(1-s));
    }
    // 주소·장소 검색(Nominatim) — 현재 위치 주변을 우선(viewbox)해 받아온 뒤, 현재 위치에서 가장 가까운 결과를 선택.
    function doSearch(){
      var q = document.getElementById('addr').value.trim();
      if(!q){ return; }
      readout.textContent = '검색 중…';
      var ref = refPoint();
      var d = 0.75; // 기준점 주변 약 ±0.75°(≈80km)를 우선 검색 영역(viewbox)으로 지정
      var vb = (ref[1]-d)+','+(ref[0]+d)+','+(ref[1]+d)+','+(ref[0]-d); // left,top,right,bottom
      var url = 'https://nominatim.openstreetmap.org/search?format=json&limit=10'
              + '&countrycodes=vn'  // 검색 결과를 베트남 내로 한정
              + '&viewbox='+vb+'&bounded=0&q='+encodeURIComponent(q);
      fetch(url)
        .then(function(r){ return r.json(); })
        .then(function(list){
          if(list && list.length){
            // 현재 위치에서 가까운 순으로 정렬 → 최근접 결과를 선택
            list.sort(function(a, b){
              return distKm(ref, [parseFloat(a.lat), parseFloat(a.lon)])
                   - distKm(ref, [parseFloat(b.lat), parseFloat(b.lon)]);
            });
            var best = list[0];
            var lat = parseFloat(best.lat), lng = parseFloat(best.lon);
            map.setView([lat,lng], 16);
            setMarker(lat, lng, best.display_name, distKm(ref, [lat,lng]));
          } else {
            readout.innerHTML = '지도 검색 결과가 없습니다. 가게·회사 상호라면 '
              + "<b style='color:#63b365'>🔗 구글 검색</b>을 눌러보세요.";
          }
        })
        .catch(function(){ readout.textContent = '검색에 실패했습니다. 잠시 후 다시 시도하세요.'; });
    }
    // 구글 지도 검색 — 상호(POI) 검색은 Google 데이터가 강하므로 새 탭으로 Google 지도 검색을 연다.
    //   현재 위치(GPS/지도 중심) 기준으로 결과를 보여주도록 좌표를 함께 전달.
    function openGoogle(){
      var q = document.getElementById('addr').value.trim();
      if(!q){ document.getElementById('addr').focus(); return; }
      var ref = refPoint();
      var url = 'https://www.google.com/maps/search/' + encodeURIComponent(q)
              + '/@' + ref[0] + ',' + ref[1] + ',14z';
      var w = window.open(url, '_blank', 'noopener');
      if(!w){ // 팝업 차단 시 직접 누를 수 있는 링크 제공
        readout.innerHTML = "팝업이 차단됐습니다. <a href='" + url
          + "' target='_blank' style='color:#63b365;font-weight:700'>구글 지도에서 열기</a>";
      }
    }
    document.getElementById('searchBtn').addEventListener('click', doSearch);
    document.getElementById('googleBtn').addEventListener('click', openGoogle);
    document.getElementById('addr').addEventListener('keydown', function(e){
      if(e.key === 'Enter'){ e.preventDefault(); doSearch(); }
    });
    // GPS 현재 위치(브라우저 위치권한 필요)
    document.getElementById('gpsBtn').addEventListener('click', function(){
      if(!navigator.geolocation){
        readout.textContent = '이 브라우저는 위치 기능을 지원하지 않습니다.'; return;
      }
      readout.textContent = '현재 위치 확인 중…';
      navigator.geolocation.getCurrentPosition(function(pos){
        var lat = pos.coords.latitude, lng = pos.coords.longitude;
        myPos = [lat, lng]; // 이후 주소 검색의 근접 정렬 기준점으로 사용
        map.setView([lat,lng], 16);
        setMarker(lat, lng, null);
        reverse(lat, lng);
      }, function(err){
        readout.textContent = (err && err.code === 1)
          ? '위치 권한이 거부되었습니다. 브라우저에서 위치 접근을 허용해 주세요.'
          : '현재 위치를 확인하지 못했습니다.';
      }, {enableHighAccuracy:true, timeout:8000, maximumAge:0});
    });
    // iframe 레이아웃 확정 후 타일 재계산(회색 여백 방지)
    setTimeout(function(){ map.invalidateSize(); }, 250);
  </script>
</body>
</html>
"""

def _admin_seat_status_view(car_rc):
    """관리자 로그인 성공 후 같은 팝업 안에서 뜨는 '좌석 신청 현황' 표.
    운전석(0번)을 제외한 해당 차량의 전체 좌석(신청 안 된 빈 좌석 포함)을 좌석 번호 순으로 나열.
    컬럼: 좌석 번호 / 신청자 / 출발지 / 목적지 / 출발 시간 / 도착 시간."""
    car = car_rc["display_name"]
    st.markdown(f'<div class="dlg-step-title">{t("admin_status_title")}</div>', unsafe_allow_html=True)
    st.caption(car)

    # 시트 문구와 좌석 표 사이 — 운전자용 현재 위치 지도(무료 OpenStreetMap/Leaflet, Google 지도 API 키 불필요).
    #   · '📍 현재 위치' 버튼: 브라우저 GPS로 현 좌표를 잡아 마커+지도 이동(HTTPS·위치권한 허용 필요).
    #   · 주소 검색창: Nominatim(OSM 무료 지오코딩)으로 입력한 장소로 이동. 지도 클릭으로도 마커 지정 가능.
    #   Streamlit components.html은 srcdoc iframe(부모와 동일 출처)이라 기본 Permissions-Policy(self)로 iframe 내부에서 GPS가 동작한다.
    _pwa_components.html(_ADMIN_LOCATION_MAP_HTML, height=340)

    empty = t("st_empty")

    # 오늘(00:00~24:00, 베트남 기준) '완료'된 탑승 이력을 좌석별로 묶는다 → 완료 건도 표에 누적 표시(자정 지나면 자동 초기화).
    today_str = now_vn().strftime("%Y-%m-%d")
    done_by_seat = {}
    try:
        for rec in load_history():
            if str(rec.get("status", "")).strip() != "완료":
                continue
            if str(rec.get("car", "")).strip() != car:
                continue
            if str(rec.get("date", "")).strip() != today_str:   # 출발 날짜 기준 '오늘'
                continue
            try:
                s = int(rec.get("seat"))
            except (TypeError, ValueError):
                continue
            done_by_seat.setdefault(s, []).append(rec)
    except Exception:
        done_by_seat = {}

    def _cells(src):
        return (
            html.escape(str(src.get("name", "")).strip() or empty),
            html.escape(str(src.get("departure", "")).strip() or empty),
            html.escape(str(src.get("destination", "")).strip() or empty),
            html.escape(str(src.get("time", "")).strip() or empty),      # 출발 시간
            html.escape(str(src.get("arrive", "")).strip() or empty),    # 도착 시간
        )

    rows_html = []
    for seat in range(1, car_rc["seats"] + 1):
        # 이 좌석의 표시 행 = 오늘 완료 이력(완료시각순) + 현재 예약(있으면 맨 아래)
        done_list = sorted(done_by_seat.get(seat, []), key=lambda r: str(r.get("completed_at", "")))
        cur = st.session_state.bookings.get((car, seat))
        seat_rows = [(rec, True) for rec in done_list]
        if cur:
            seat_rows.append((cur, False))

        if not seat_rows:
            rows_html.append(
                f'<tr class="seat-status-empty"><td class="ss-seat">{seat}</td>'
                f'<td>{empty}</td><td>{empty}</td><td>{empty}</td>'
                f'<td class="ss-time">{empty}</td><td class="ss-time">{empty}</td></tr>'
            )
            continue

        span = len(seat_rows)
        for idx, (src, is_done) in enumerate(seat_rows):
            name, dep, dest, dept, arr = _cells(src)
            rcls = ' class="seat-status-done"' if is_done else ''
            seat_td = f'<td class="ss-seat" rowspan="{span}">{seat}</td>' if idx == 0 else ''
            rows_html.append(
                f'<tr{rcls}>{seat_td}<td>{name}</td><td>{dep}</td>'
                f'<td>{dest}</td><td class="ss-time">{dept}</td><td class="ss-time">{arr}</td></tr>'
            )
    table_html = (
        '<div class="seat-status-wrap"><table class="seat-status-table"><thead><tr>'
        f'<th>{t("st_seat")}</th><th>{t("st_name")}</th><th>{t("st_dep")}</th>'
        f'<th>{t("st_dest")}</th><th>{t("st_deptime")}</th><th>{t("st_arrive")}</th>'
        f'</tr></thead><tbody>{"".join(rows_html)}</tbody></table></div>'
    )
    st.markdown(table_html, unsafe_allow_html=True)
    # 하단 버튼: 닫기(75%) + 로그아웃(25%) 가로 병렬. 다이얼로그 컬럼을 1:1로 강제하는 전역 CSS를
    #   이기기 위해 st-key-admin_status_btns 스코프로 3:1 flex를 덮어쓴다(스타일 블록 참고).
    with st.container(key="admin_status_btns"):
        cbtn, lbtn = st.columns([3, 1])
        with cbtn:
            if st.button(t("btn_close"), type="primary", key="admin_status_close_btn", use_container_width=True):
                st.session_state.admin_seat_status_open = False
                st.session_state.seatmap_car = None
                st.rerun()
        with lbtn:
            # 로그아웃 → 관리자 잠금 + '로그인 유지' 해제 + localStorage 1회성 클리어(재접속 자동복원 방지)
            if st.button(t("btn_logout"), key="admin_status_logout_btn", use_container_width=True):
                st.session_state.admin_unlocked = False
                st.session_state.admin_keep = False
                st.session_state.admin_clear_ls = True
                st.session_state.confirm_reset_all = False
                st.session_state.admin_seat_status_open = False
                st.session_state.seatmap_car = None
                st.toast(t("admin_locked_toast"))
                st.rerun()

@st.dialog(" ", on_dismiss=_close_seatmap)
def seatmap_dialog(car_rc):
    """차량 이름 클릭 시 뜨는 팝업. 좌석 미선택이면 '🚗 좌석 선택'(배치도), 좌석 클릭 시 같은 팝업 안에서
    '📝 신청 정보 입력' 폼으로 전환된다. @st.dialog 크롬 title은 열린 중 못 바꾸므로 공백(' ')으로 두고
    단계별 제목을 본문 최상단(.dlg-step-title)에 직접 그려 단계별로 구분한다."""
    car = car_rc["display_name"]
    # 관리자 로그인 성공 상태면 → 좌석맵 대신 '좌석 신청 현황' 표 표시(같은 팝업 안)
    if st.session_state.get("admin_seat_status_open"):
        _admin_seat_status_view(car_rc)
        return
    # 운전석 클릭으로 관리자 로그인 요청 상태면 → 좌석맵 대신 관리자 로그인 폼 표시(같은 팝업 안, 자체 제목 렌더)
    if st.session_state.get("admin_login_open"):
        _admin_login_form()
        return
    booked = [s_id for (c_name, s_id) in st.session_state.bookings.keys() if c_name == car]
    sel = st.session_state.selected_seat_state.get(car, "-- 선택 --")
    seat_num = None
    if sel != "-- 선택 --":
        try:
            seat_num = int(sel.split(" ")[1])
        except Exception:
            seat_num = None
    is_form = seat_num is not None and seat_num not in booked
    # 단계별 커스텀 제목: 신청 폼일 때만 '📝 신청 정보 입력' 제목을 표시. 좌석 배치도(좌석 선택)는 제목 없이 바로 노출.
    if is_form:
        st.markdown(f'<div class="dlg-step-title">{t("form_step_title")}</div>', unsafe_allow_html=True)
    # 좌석이 선택된 상태면 같은 팝업 안에서 신청 폼을 보여준다.
    if is_form:
        _booking_form(car, seat_num)
        return
    # 아직 미선택 → 좌석 배치도 + 빈좌석 SEATSEL 숨김버튼(클릭 시 on_seat_click이 좌석 선택 → 폼으로 전환)
    st.markdown(f'<div class="car-header-center" style="margin-top:0!important;">{car_title_frame(car_rc["mk"], car_rc["logo_html"] + car_rc["nav_label"])}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="car-layout-container" style="width:100%!important;">{render_car_layout(car, car_rc["layout"], st.session_state.bookings)}</div>', unsafe_allow_html=True)
    available = [f"좌석 {seat}" for seat in range(1, car_rc["seats"] + 1) if seat not in booked]
    if not available:
        st.error(t("full"))
    else:
        st.caption(t("seatmap_hint"))
    for seat in range(1, car_rc["seats"] + 1):
        if f"좌석 {seat}" in available:
            st.button(
                f"SEATSEL::{car}::{seat}",
                key=f"seatsel_{car}_{seat}",
                on_click=on_seat_click,
                args=(car, seat),
            )
    # INNOVA·SEDONA: 운전석 클릭 시 JS가 대신 누를 숨김 ADMINLOGIN 버튼(→ 관리자 로그인 폼)
    if ("INNOVA" in car) or ("SEDONA" in car):
        st.button(f"ADMINLOGIN::{car}", key=f"adminlogin_{car}", on_click=_open_admin_login)

# 차량 이름 바(로고+이름) 4개를 세로로 배치 — 웹(루트)·앱(?m=1) 완전 동일 레이아웃(메모: UI 양쪽 동일).
#  메인 화면엔 차량 네이밍 바만 노출하고, 바를 클릭하면 해당 차량 좌석 배치도가 팝업(seatmap_dialog)으로 뜬다.
if "seatmap_car" not in st.session_state:
    st.session_state.seatmap_car = None
# 차량 선택은 '정사각형 타일 2×2'. 한 줄에 2개씩 끊어 배치하고, 홀수면 마지막 줄 오른쪽 칸은 비워 둔다.
#  (가로로 긴 바 → 오클릭 우려로 변경. 웹·앱 동일 레이아웃이며 크기만 화면 폭에 맞춰 달라진다)
with st.container(key="car_nav_grid"):
    for _row in range(0, len(resolved_cars), 2):
        _cols = st.columns(2)
        for _slot in range(2):
            i = _row + _slot
            if i >= len(resolved_cars):
                break          # 홀수 대수: 오른쪽 칸은 빈 채로 둬 2열 틀 유지
            car_rc = resolved_cars[i]
            with _cols[_slot]:
                # 타일 전체를 클릭 가능하게 렌더 + JS가 대신 눌러줄 숨김 버튼
                st.markdown(
                    car_nav_tile(car_rc["mk"], car_rc["logo_html"], car_rc["nav_label"], i),
                    unsafe_allow_html=True,
                )
                # on_click 콜백으로 즉시 상태 세팅 → 단일 rerun에 바로 팝업 오픈(클릭 후 바로 열림)
                st.button(f"CARNAV::{i}", key=f"carnavclick_{i}",
                          on_click=_open_seatmap, args=(car_rc["display_name"],))
# 이름 클릭 상태면 해당 차량 좌석맵 팝업을 띄운다
if st.session_state.get("seatmap_car"):
    _tgt = next((c for c in resolved_cars if c["display_name"] == st.session_state.seatmap_car), None)
    if _tgt:
        seatmap_dialog(_tgt)
    else:
        st.session_state.seatmap_car = None

# 7. 좌석 선택 시 뜨는 외근 신청 정보 입력 팝업(모달 다이얼로그)
def _reset_booking_selection():
    """팝업을 X·바깥클릭·ESC로 닫을 때 선택했던 빈자리를 다시 빈자리로 리셋."""
    car = st.session_state.get("active_booking_car")
    if car and car in st.session_state.selected_seat_state:
        st.session_state.selected_seat_state[car] = "-- 선택 --"
        st.session_state[f"dropdown_trigger_spec_{car}"] = "-- 선택 --"
    st.session_state.editing_booking = None
    st.session_state.duplicate_error_msg = None

@st.dialog(t("form_step_title"), on_dismiss=_reset_booking_selection)
def booking_dialog(car_target, seat_target):
    # 신청/수정 팝업('📝 신청 정보 입력') — 공용 폼(_booking_form)을 그대로 사용
    _booking_form(car_target, seat_target)

# 앱 좌석맵 팝업 등에서 좌석이 선택되면 selected_seat_state에서 신청 트리거를 도출(웹은 위 _render_car_body에서 세팅됨).
if not selected_seat_trigger:
    for _rc in resolved_cars:
        _sel = st.session_state.selected_seat_state.get(_rc["display_name"], "-- 선택 --")
        if _sel and _sel != "-- 선택 --":
            try:
                _sn = int(_sel.split(" ")[1])
            except Exception:
                continue
            _bk = [s for (c, s) in st.session_state.bookings.keys() if c == _rc["display_name"]]
            if _sn not in _bk:
                selected_seat_trigger = (_rc["display_name"], _sn)
                break

# 예약 수정 중이면(예약된 좌석은 selectbox에 없으므로) 해당 예약으로 팝업을 연다
if st.session_state.editing_booking and not selected_seat_trigger:
    selected_seat_trigger = st.session_state.editing_booking

# 선택된 빈자리(또는 수정 대상)가 있으면 신청/수정 팝업을 띄운다.
#  단, 앱에서 좌석맵 팝업이 열려 있으면 신청 폼은 그 팝업 안에서 처리하므로 top-level 팝업은 건너뛴다(2중 dialog 방지).
if selected_seat_trigger:
    car_target, seat_target = selected_seat_trigger
    st.session_state.active_booking_car = car_target
    # 좌석맵 팝업이 열려 있으면 신청 폼은 그 팝업 안에서 처리(웹·앱 동일) → top-level 팝업 건너뜀(2중 dialog 방지)
    if not st.session_state.get("seatmap_car"):
        booking_dialog(car_target, seat_target)

# ── 엑셀 데이터 내보내기(탑승 이력) 팝업 ─────────────────────────────
def _build_history_xlsx(rows):
    """필터된 탑승 이력 rows를 예약이력 스키마(csv_headers 순서)로 XLSX 바이트 생성.
    openpyxl 미설치 시 CSV 바이트로 폴백."""
    headers = t("csv_headers")
    def _cells(r):
        return [
            r.get("created_at", ""), r.get("car", ""), r.get("seat", ""), r.get("name", ""),
            r.get("date", ""), r.get("departure", ""), r.get("destination", ""),
            r.get("time", ""), r.get("arrive", ""),
        ]
    try:
        import io as _io, openpyxl
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "탑승 이력"
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True); c.alignment = Alignment(horizontal="center", vertical="center")
        for r in rows:
            ws.append(_cells(r))
        for i, w in enumerate([20, 20, 6, 12, 12, 14, 16, 10, 10], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        buf = _io.BytesIO(); wb.save(buf)
        return buf.getvalue(), "xlsx"
    except Exception:
        import io as _io, csv as _csv
        sb = _io.StringIO(); w = _csv.writer(sb); w.writerow(headers)
        for r in rows:
            w.writerow(_cells(r))
        return ("﻿" + sb.getvalue()).encode("utf-8"), "csv"


# ─────────────────────────────────────────────────────────────
# 📊 기간 요약 통계 — 엑셀을 열지 않아도 화면에서 바로 읽히는 지표
#   기존에는 연·월·일을 골라 XLSX로 '내려받는 것'만 가능해, 추이를 보려면 매번 엑셀을 열어야 했다.
#   · 색은 '크기 비교(magnitude)'용 단일 색 하나만 쓴다 — 항목마다 다른 색을 주면
#     색이 의미를 갖는 것처럼 읽혀 오해를 부른다(순위가 색을 바꾸는 것도 금지).
#   · CHART_BAR은 눈대중이 아니라 검증 도구로 골랐다: 다크 배경 대비 3:1 이상, 밝기 밴드 통과.
#   · 막대 끝에 값을 직접 표기(항목이 5개 이하라 전부 라벨링해도 지저분해지지 않는다).
#   · 표 형태의 원본은 같은 팝업의 '엑셀 다운로드'가 담당한다(숫자를 직접 확인할 경로 확보).
#   [벤치마킹] Lê Quang Trung(AE) "SMART REPORT VER 2" — 일·월·년 통계를 차트로 보관 /
#              Lê Khắc Hưng(AE Leader) "Log Analyzer V2" — 차트·통계·필터로 빠른 조회 /
#              Nguyễn Hoàng Qúy(AE) — dashboard·search/filter UI. 상세: docs/BENCHMARK.md
# ─────────────────────────────────────────────────────────────
CHART_BAR = "#1c7ed6"      # 크기 비교용 단일 색(빈 좌석 테두리와 동일 계열 → 앱 전체와 통일)
CHART_TRACK = "rgba(255,255,255,0.05)"

def _bars_html(pairs, total=0):
    """(라벨, 값) 목록을 가로 막대로. 값이 큰 순으로 이미 정렬된 목록을 받는다."""
    if not pairs:
        return ""
    mx = max(v for _l, v in pairs) or 1
    out = []
    for label, v in pairs:
        pct = max(2.0, v * 100.0 / mx)          # 0건이어도 흔적은 남겨 축이 끊겨 보이지 않게
        share = f" · {v * 100.0 / total:.0f}%" if total else ""
        tip = f"{label}: {v}"
        out.append(
            '<div style="display:flex; align-items:center; gap:8px; margin:0 0 2px 0;">'
            f'<div title="{esc(label)}" style="flex:0 0 33%; min-width:0; font-size:11px; color:#adb5bd; '
            'overflow:hidden; text-overflow:ellipsis; white-space:nowrap;">' + esc(label) + '</div>'
            f'<div style="flex:1 1 auto; min-width:0; background:{CHART_TRACK}; border-radius:4px;">'
            f'<div title="{esc(tip)}" style="width:{pct:.1f}%; height:10px; background:{CHART_BAR}; '
            'border-radius:0 4px 4px 0;"></div></div>'
            '<div style="flex:0 0 auto; font-size:11px; color:#e9ecef; font-variant-numeric:tabular-nums;">'
            + esc(str(v)) + esc(share) + '</div>'
            '</div>'
        )
    return "".join(out)

def _stat_tiles(tiles):
    """숫자 하나가 답인 지표는 차트로 만들지 않고 타일로 보여준다."""
    cells = "".join(
        '<div style="flex:1 1 0; min-width:0; background:#1b1f27; border:1px solid #2b2f38; '
        'border-radius:8px; padding:8px 10px;">'
        f'<div style="font-size:10px; color:#868e96; white-space:nowrap; overflow:hidden; '
        f'text-overflow:ellipsis;">{esc(label)}</div>'
        f'<div style="font-size:18px; font-weight:800; color:#e9ecef; white-space:nowrap; '
        f'overflow:hidden; text-overflow:ellipsis;">{esc(value)}</div></div>'
        for label, value in tiles
    )
    return f'<div style="display:flex; gap:6px; margin:0 0 10px 0;">{cells}</div>'

def _render_period_stats(rows):
    """선택한 기간의 탑승 이력 요약 — 타일 3개 + 차량별 / 목적지 TOP5 / 출발 시간대."""
    total = len(rows)

    by_car, by_dest, by_hour = {}, {}, {}
    for r in rows:
        car = _short_car_name(str(r.get("car", "")).strip()) or "—"
        by_car[car] = by_car.get(car, 0) + 1
        dest = str(r.get("destination", "")).strip() or "—"
        by_dest[dest] = by_dest.get(dest, 0) + 1
        try:
            hh = int(str(r.get("time", "")).split(":")[0]) % 24
            by_hour[hh] = by_hour.get(hh, 0) + 1
        except Exception:
            pass

    top_dest = max(by_dest.items(), key=lambda kv: kv[1])[0] if by_dest else "—"
    st.markdown(f'<div class="dlg-step-title" style="margin:0 0 8px 0; font-size:15px !important;">'
                f'{t("stats_title")}</div>', unsafe_allow_html=True)
    st.markdown(_stat_tiles([
        (t("stats_total"), str(total)),
        (t("stats_cars"), str(len(by_car))),
        (t("stats_top_dest"), top_dest),
    ]), unsafe_allow_html=True)

    def _section(title, pairs):
        if not pairs:
            return
        st.markdown(
            f'<div style="font-size:12px; font-weight:700; color:#fab005; margin:6px 0 4px 0;">{esc(title)}</div>'
            + _bars_html(pairs, total),
            unsafe_allow_html=True,
        )

    _section(t("stats_by_car"), sorted(by_car.items(), key=lambda kv: -kv[1]))
    _section(t("stats_top_dests"), sorted(by_dest.items(), key=lambda kv: -kv[1])[:5])
    if by_hour:
        # 시간대는 '순서가 있는 축'이라 값 크기순이 아니라 시간순으로, 빈 시간대도 0으로 채운다
        #  (비어 있는 구간을 빼면 분포 모양이 왜곡된다)
        lo, hi = min(by_hour), max(by_hour)
        _section(t("stats_by_hour"),
                 [(t("stats_hour", h=h), by_hour.get(h, 0)) for h in range(lo, hi + 1)])


def _close_export():
    st.session_state.export_open = False


@st.dialog(t("export_title"), on_dismiss=_close_export)
def excel_export_dialog():
    history = load_history()

    def _ymd(r):
        d = (r.get("date") or str(r.get("created_at", ""))[:10])
        if len(d) >= 10 and d[4] == "-" and d[7] == "-":
            try:
                return int(d[:4]), int(d[5:7]), int(d[8:10])
            except Exception:
                return None
        return None

    vn = now_vn()
    years = sorted({ymd[0] for r in history if (ymd := _ymd(r))} | {vn.year}, reverse=True)

    c1, c2, c3 = st.columns(3)
    with c1:
        sy = st.selectbox(t("export_year"), years, index=years.index(vn.year), key="exp_year")
    with c2:
        sm = st.selectbox(t("export_month"), list(range(1, 13)), index=vn.month - 1, key="exp_month")
    with c3:
        day_opts = [t("export_all")] + list(range(1, 32))
        sd = st.selectbox(t("export_day"), day_opts, index=0, key="exp_day")

    # 선택 연·월(·일) 필터 — 출발날짜 우선, 없으면 신청일시 날짜로 판정
    rows = []
    for r in history:
        ymd = _ymd(r)
        if not ymd:
            continue
        ry, rm, rd = ymd
        if ry != sy or rm != sm:
            continue
        if sd != t("export_all") and rd != sd:
            continue
        rows.append(r)

    ym = f"{sy}_{sm:02d}" + ("" if sd == t("export_all") else f"_{sd:02d}")
    data, ext = _build_history_xlsx(rows)
    fname = t("export_file", ym=ym)
    if ext == "csv":
        fname = fname.rsplit(".", 1)[0] + ".csv"
    mime = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            if ext == "xlsx" else "text/csv")

    st.write("")
    if not rows:
        st.info(t("export_empty"))
    else:
        _render_period_stats(rows)
    with st.container(key="export_dl"):
        st.download_button(
            t("export_btn"),
            data=data, file_name=fname, mime=mime,
            use_container_width=True, key="export_download_btn",
        )
    st.caption(t("export_caption"))


def _close_cancel():
    st.session_state.cancel_target = None


@st.dialog(t("cancel_title"), on_dismiss=_close_cancel)
def cancel_dialog(car, seat):
    """예약 취소 확인 팝업. 예전에는 버튼 한 번에 즉시 삭제돼 오터치로 남의 배차가 사라질 수 있었다.
    남의 예약이면 본인 확인을 먼저 통과해야 하고, 삭제 직전에 권한을 한 번 더 검사한다."""
    info = st.session_state.bookings.get((car, seat))
    if not info:
        _close_cancel()
        return
    st.markdown(
        f'<div style="font-size:13px; color:#c7ccd6; line-height:1.6; margin-bottom:8px; white-space:pre-line;">'
        f'{esc(t("cancel_confirm", car=car, seat=seat, name=info.get("name", "")))}</div>',
        unsafe_allow_html=True,
    )
    if not owner_gate(car, seat, info):
        return
    cc1, cc2 = st.columns(2)
    with cc1:
        if st.button(t("cancel_yes"), type="primary", use_container_width=True, key="cancel_yes_btn"):
            if not (can_manage_booking(info) or (car, seat) in st.session_state.get("verified_bookings", set())):
                st.error(t("owner_err"))
                st.stop()
            log_action("cancel", car, seat, info)      # 삭제 전에 기록 — 지운 뒤엔 근거가 남지 않는다
            del st.session_state.bookings[(car, seat)]
            save_bookings(st.session_state.bookings)
            st.session_state.cancel_target = None
            st.toast(t("toast_cancelled", name=info.get("name", ""), seat=seat))
            st.rerun()
    with cc2:
        if st.button(t("btn_cancel"), use_container_width=True, key="cancel_no_btn"):
            st.session_state.cancel_target = None
            st.rerun()


def _close_arrival():
    st.session_state.arrive_target = None


@st.dialog(t("arrive_title"), on_dismiss=_close_arrival)
def arrival_dialog(car, seat):
    """도착 완료 시 '도착 시간만' 설정하는 팝업. 완료 누르면 그 시간으로 탑승 이력에 기록 + 좌석 해제."""
    info = st.session_state.bookings.get((car, seat))
    if not info:
        _close_arrival()
        return
    # 도착완료도 좌석을 비우는 파괴적 동작 → 남의 예약이면 본인 확인을 먼저 통과해야 한다.
    if not owner_gate(car, seat, info):
        return
    st.markdown(
        f'<div style="font-size:13px; color:#c7ccd6; line-height:1.6; margin-bottom:8px; white-space:pre-line;">'
        f'{esc(t("arrive_desc", car=car, seat=seat, name=info.get("name", "")))}</div>',
        unsafe_allow_html=True,
    )
    # 6. 도착 시간 (기본값은 도착 완료 클릭 시각의 5분 슬롯; 위젯 상태로 유지)
    a_time = st.time_input(t("f_arrive"), step=300, key="arrive_input_tick")
    if st.button(t("arrive_done"), type="primary", use_container_width=True, key="arrive_done_btn"):
        # 값 변경 직전 권한 재확인
        if not (can_manage_booking(info) or (car, seat) in st.session_state.get("verified_bookings", set())):
            st.error(t("owner_err"))
            st.stop()
        arrive_str = a_time.strftime("%H:%M") if a_time else "00:00"
        archive_booking(car, seat, {**info, "arrive": arrive_str}, status="완료")
        log_action("done", car, seat, info, note=t("f_arrive") + " " + arrive_str)
        del st.session_state.bookings[(car, seat)]
        save_bookings(st.session_state.bookings)
        st.session_state.arrive_target = None
        st.toast(t("toast_done", name=info.get("name", ""), seat=seat))
        st.rerun()


# 8. 실시간 배차 예약 현황판 명단 출력 (하단 단독 배치)
num_bookings = len(st.session_state.bookings)
st.write("")

# 저장이 실패했으면 반드시 화면에 알린다 — 조용히 넘기면 사용자는 저장된 줄 알고 자리를 떠난다.
if st.session_state.pop("save_failed", False):
    st.error(t("save_failed"))

search_query = ""
# 현황판 헤더(제목 + 검색 + 예약이력 다운로드)는 예약 유무와 무관하게 항상 렌더.
#  → 예약 신청이 없더라도 '실시간 차량 예약 현황' 제목과 '예약 이력(CSV)' 버튼이 항상 노출된다.
h_title, h_search, h_csv = st.columns([2, 1, 1], vertical_alignment="center")
with h_title:
    st.markdown(f'<div class="board-title">{t("list_title", n=num_bookings)}</div>', unsafe_allow_html=True)
with h_search:
    # 검색창은 예약이 있을 때만 노출(빈 상태에서 빈 검색창 방지). CSV 위치는 컬럼으로 고정 유지.
    if st.session_state.bookings:
        search_query = st.text_input(
            t("search_ph"),
            placeholder=t("search_ph"),
            key="booking_search_query",
            label_visibility="collapsed"
        )
with h_csv:
    # 예약 이력 → 클릭 시 '엑셀 데이터 내보내기' 팝업(연/월/일 선택 후 XLSX 다운로드)을 애니메이션 모달로 오픈.
    #  버튼은 TAXI 박스와 동일하게 80% 가운데 정렬(양쪽 10% 여백) → 박스 오른쪽 끝선과 한 줄.
    with st.container(key="csv_inset"):
        if st.button(t("csv_btn"), use_container_width=True, key="open_export_btn"):
            st.session_state.export_open = True
            st.rerun()

# 승인 대기 요약 — 관리자만이 아니라 전원에게 보인다(내 배차가 확정됐는지 모두가 알아야 한다).
#  임박·초과 건이 있으면 붉은 톤으로 올려 아침에 놓치지 않게 한다.
_pend = pending_bookings()
if _pend:
    _urgent = sum(1 for _k, _v in _pend if pending_urgency(_v))
    _pb_fg, _pb_bg = ("#ffa94d", "rgba(253,126,20,0.14)") if _urgent else ("#fab005", "rgba(250,176,5,0.12)")
    _pb_txt = t("pending_banner_urgent", n=len(_pend), u=_urgent) if _urgent else t("pending_banner", n=len(_pend))
    st.markdown(
        f'<div style="background:{_pb_bg}; border:1px solid {_pb_fg}; color:{_pb_fg}; '
        f'border-radius:8px; padding:6px 10px; margin:2px 0 8px 0; font-size:12px; font-weight:700;">'
        f'{esc(_pb_txt)}</div>',
        unsafe_allow_html=True,
    )

# 예약 이력 버튼이 눌렸으면 엑셀 내보내기 팝업(모달)을 띄운다. 닫으면 on_dismiss로 플래그 해제.
if st.session_state.get("export_open"):
    excel_export_dialog()

# 도착 완료 버튼이 눌렸으면 '도착 시간' 입력 팝업을 띄운다. 완료 시 이력 기록 + 좌석 해제.
if st.session_state.get("arrive_target"):
    _at_car, _at_seat = st.session_state.arrive_target
    arrival_dialog(_at_car, _at_seat)

# 예약 취소 버튼이 눌렸으면 확인 팝업을 띄운다(본인 확인 → 삭제).
if st.session_state.get("cancel_target"):
    _ct_car, _ct_seat = st.session_state.cancel_target
    cancel_dialog(_ct_car, _ct_seat)

if st.session_state.bookings:
    # 검색어에 매칭되는 예약만 필터링 (대소문자 무시, 여러 필드 대상)
    q = (search_query or "").strip().lower()
    filtered_items = []
    for (car_name, seat_num), info in list(st.session_state.bookings.items()):
        haystack = " ".join([
            car_name, info.get("name", ""), info.get("destination", ""),
            info.get("departure", ""), info.get("date", "")
        ]).lower()
        if not q or q in haystack:
            filtered_items.append(((car_name, seat_num), info))

    if q and not filtered_items:
        st.markdown(f'<div style="font-size: 12px; color: #8e929e; text-align: center; padding: 10px;">{esc(t("no_result", q=search_query))}</div>', unsafe_allow_html=True)

    # 예약 수정: 해당 예약을 입력 필드에 로드 후 editing_booking 설정(→ rerun 시 팝업 오픈)
    def _start_edit(bc_name, bseat):
        info = st.session_state.bookings.get((bc_name, bseat), {})
        st.session_state.input_user_real_name = info.get("name", "")
        st.session_state.input_user_departure_loc = info.get("departure", "")
        st.session_state.input_user_destination_loc = info.get("destination", "")
        try:
            y, mo, d = map(int, info.get("date", "").split("-"))
            st.session_state.input_user_departure_date = datetime.date(y, mo, d)
        except Exception:
            pass
        try:
            h, m = map(int, info.get("time", "").split(":"))
            st.session_state.input_user_departure_time_tick = datetime.time(h, m)
        except Exception:
            pass
        try:
            h, m = map(int, info.get("arrive", "").split(":"))
            st.session_state.input_user_arrival_time_tick = datetime.time(h, m)
        except Exception:
            pass
        st.session_state.selected_seat_state[bc_name] = f"좌석 {bseat}"
        st.session_state.editing_booking = (bc_name, bseat)
        st.session_state.active_booking_car = bc_name
        st.session_state.duplicate_error_msg = None

    # 예약 카드 1장 렌더 — PC·앱 공통(동일 UI).
    #  구조: 차량색 배경 컨테이너 안에 헤더(차량명+좌석배지) + 3열×3행 그리드.
    #        · 1열: 신청자 / 출발지 / 목적지   · 2열: 출발날짜 / 출발시간 / 도착시간   · 3열: 예약수정 / 예약취소 / 도착완료 버튼
    #        3열 모두 동일 폭(1:1:1), 각 행 높이도 버튼과 맞춰 가로 정렬.
    def _render_booking_card(bc_name, bseat, binfo):
        # 카드 배경·글자색 = 해당 차량 색. INNOVA는 밝은 실버(어두운 글자), 그 외는 어두운 틴트(밝은 글자).
        mk = _model_key(bc_name)
        c_bg, c_fg, c_bd = CAR_CARD_STYLE.get(mk, CAR_CARD_STYLE["innova"])
        # 카드별 고유 컨테이너 키 → 배경/테두리를 카드 전체(정보+버튼 포함)에 입힌다.
        _safe = "".join(ch for ch in f"{bc_name}{bseat}" if ch.isalnum())
        cardkey = f"bkcard_{mk}_{_safe}"

        # 헤더: 차량 로고 + 짧은 차량명(INNOVA/SEDONA/VF5/TAXI n) + 좌석 배지 + 구분선. 카드 상단 전체폭.
        car_logo = brand_logo(bc_name)          # 브랜드 인라인 SVG/이미지 로고
        car_short = _short_car_name(bc_name)    # 'TOYOTA INNOVA (7 SEAT)' → 'INNOVA'
        header_html = (
            '<div style="font-weight: bold; font-size: 12px; display: flex; justify-content: space-between; align-items: center; gap: 4px;">'
            f'<span style="color: {c_fg}; font-weight: bold; font-size: 15px; flex: 1 1 auto; min-width: 0; display: flex; align-items: center;">'
            f'{car_logo}'
            f'<span style="overflow: hidden; text-overflow: ellipsis; white-space: nowrap;">{car_short}</span>'
            '</span>'
            f'<span style="flex: 0 0 auto; background: {BOOKED_SEAT_LINE}; border: 1px solid {BOOKED_SEAT_LINE}; color: #ffffff; padding: 1px 5px; border-radius: 4px; font-size: 12px; font-weight: bold; white-space: nowrap;">{t("seat_n", n=bseat)}</span>'
            '</div>'
            # 승인 상태 배지 — 대기(호박색·임박하면 붉은색) / 승인(초록). 한눈에 '내 배차가 확정됐는지' 알 수 있게.
            f'{_status_chip(binfo)}'
            f'<hr style="border: 0; border-top: 1px solid {c_bd}; margin: 4px 0;">'
        )

        # 카드가 화면 절반 폭(좁음)이라 정보를 'CSS 2열 그리드'로 배열(제목 + 값 한 줄, 길면 자동 줄바꿈).
        #  중첩 Streamlit 컬럼을 쓰지 않아 열 겹침·가로 오버플로우가 없다. 높이 축소를 위해 인라인·압축 배치.
        #  왼쪽열=신청자·출발지·목적지 / 오른쪽열=출발날짜·출발시간·도착시간 → 행 순서대로 좌우 번갈아 채움.
        #  value는 신청자가 입력한 값(이름·출발지·목적지) → esc()로 감싸 태그가 실행되지 않게 한다.
        def _cell(label, value):
            return (f'<div style="min-width:0; overflow-wrap:anywhere;">'
                    f'<strong>{label}</strong><br>{esc(value)}</div>')
        info_grid = (
            '<div style="display:grid; grid-template-columns:1fr 1fr; gap:4px 8px; '
            f'font-size:12px; color:{c_fg}; line-height:1.2; margin-bottom:5px;">'
            + _cell(t('c_applicant'), binfo.get('name', ''))
            + _cell(t('c_date'), binfo.get('date', ''))
            + _cell(t('c_departure'), binfo.get('departure', ''))
            + _cell(t('c_time'), binfo.get('time', ''))
            + _cell(t('c_destination'), binfo.get('destination', ''))
            + _cell(t('c_arrive'), binfo.get('arrive', ''))
            + '</div>'
        )

        def _btn_edit():
            if st.button(t("btn_edit_bk"), key=f"edit_btn_{bc_name}_{bseat}", use_container_width=True):
                _start_edit(bc_name, bseat)
                st.rerun()
        def _btn_cancel():
            # 즉시 삭제하지 않고 확인 팝업을 연다(오터치로 남의 배차가 사라지던 문제).
            if st.button(t("btn_cancel_bk"), key=f"cancel_btn_{bc_name}_{bseat}", use_container_width=True):
                st.session_state.cancel_target = (bc_name, bseat)
                st.rerun()
        def _btn_done():
            # 도착 완료: 도착 시간 입력 팝업을 연다(완료 눌러야 이력 기록 + 좌석 해제).
            if st.button(t("btn_done_bk"), key=f"done_btn_{bc_name}_{bseat}", use_container_width=True):
                st.session_state.arrive_target = (bc_name, bseat)
                # 도착 시간 기본값 = 출발 시각과 동일. 예: 출발 08:10 → 도착 기본 08:10
                try:
                    _dh, _dm = (int(x) for x in str(binfo.get("time", "")).split(":")[:2])
                    _dh, _dm = _dh % 24, _dm % 60
                except Exception:
                    _dh, _dm = 0, 0
                st.session_state.arrive_input_tick = datetime.time(_dh, _dm)
                st.rerun()

        # 카드 전체(정보 + 버튼)를 감싸는 컨테이너에 차량색 배경을 입힌다(키별 1회성 스타일 주입).
        st.markdown(
            f"<style>.st-key-{cardkey}{{background:{c_bg} !important; border:1px solid {c_bd} !important; "
            f"border-radius:8px !important; padding:7px 8px !important; margin-bottom:5px !important;}}</style>",
            unsafe_allow_html=True,
        )
        with st.container(key=cardkey):
            st.markdown(header_html, unsafe_allow_html=True)
            st.markdown(info_grid, unsafe_allow_html=True)          # 정보 2열 그리드(위)
            # 버튼 3개(예약수정/취소/도착완료)는 정보 아래 '가로 3분할'로 배치 → 카드 높이를 대폭 축소.
            bcol1, bcol2, bcol3 = st.columns(3)
            with bcol1:
                _btn_edit()
            with bcol2:
                _btn_cancel()
            with bcol3:
                _btn_done()

    # 배차 예약 카드를 '한 줄에 2개(가로 2열)'로 배치하되, '같은 차량끼리만' 짝을 짓는다.
    #  → 서로 다른 차량이 한 줄에 섞이지 않는다. 한 차량 카드가 홀수면 그 차량 마지막 줄 오른쪽 칸은 비워 둔다.
    #    (예: VF5가 1대면 [VF5][빈칸], 다음 줄부터 TAXI 시작)
    with st.container(key="booking_board"):
        # 차량 순으로 그룹을 만든다(각 그룹 = 같은 차량의 좌석번호순 예약 목록).
        car_groups = []
        shown = set()
        for rcar in resolved_cars:
            car_items = sorted(
                [it for it in filtered_items if it[0][0] == rcar["display_name"]],
                key=lambda kv: kv[0][1]
            )
            if car_items:
                car_groups.append(car_items)
                for key_, _binfo in car_items:
                    shown.add(key_)
        # 현재 차량 구성에 없는(4/7 설정 변경·삭제된 차량 등) 예약: 표시명별로 묶어 뒤에 추가.
        leftover_names = []
        leftover_map = {}
        for key_, binfo in filtered_items:
            if key_ in shown:
                continue
            nm = key_[0]
            if nm not in leftover_map:
                leftover_map[nm] = []
                leftover_names.append(nm)
            leftover_map[nm].append((key_, binfo))
        for nm in leftover_names:
            car_groups.append(leftover_map[nm])

        # 각 차량 그룹을 2개씩 한 줄로 렌더. 오른쪽 칸이 없으면(홀수) 빈 컬럼으로 남겨 '2열' 틀 유지.
        for group in car_groups:
            for i in range(0, len(group), 2):
                col_l, col_r = st.columns(2)
                with col_l:
                    (bc_name, bseat), binfo = group[i]
                    _render_booking_card(bc_name, bseat, binfo)
                with col_r:
                    if i + 1 < len(group):
                        (bc_name, bseat), binfo = group[i + 1]
                        _render_booking_card(bc_name, bseat, binfo)
                    # 홀수 마지막 줄: 오른쪽 칸 비워 둠(같은 차량 병렬 틀 유지)

else:
    # 제목·CSV는 위 헤더에서 이미 항상 렌더되므로, 빈 상태에서는 안내 문구만 표시.
    st.markdown(f'<div style="font-size: 12px; color: #8e929e; text-align: center; padding: 10px;">{t("no_bookings")}</div>', unsafe_allow_html=True)

# ⚡ [관리자] 전체 초기화 · 백업/복원 · 활동 기록 — 관리자 로그인(admin_unlocked) 후에만 노출/동작
#   ⚠️ 예약이 0건일 때도 반드시 보여야 한다. '초기화 직후'가 바로 되돌리기가 필요한 순간인데,
#      이 블록이 `if 예약이 있으면:` 안에 있으면 초기화하자마자 되돌리기 버튼에 닿을 수 없다.
#      (그래서 예약 유무와 무관하게 항상 렌더되도록 바깥으로 뺐다)
if st.session_state.get("admin_unlocked"):
    st.markdown('<hr style="border: 0; border-top: 1px solid #2d2f34; margin: 12px 0 8px 0;">', unsafe_allow_html=True)
    if not st.session_state.get("confirm_reset_all"):
        rlc1, rlc2 = st.columns(2)
        with rlc1:
            if st.button(t("btn_reset_all"), key="reset_all_btn", use_container_width=True,
                         disabled=not st.session_state.bookings):
                st.session_state.confirm_reset_all = True
                st.rerun()
        with rlc2:
            # 관리자 재잠금(로그아웃) — 세션·유지 플래그 해제 + localStorage 1회성 클리어 예약
            if st.button(t("admin_lock"), key="admin_lock_btn", use_container_width=True):
                st.session_state.admin_unlocked = False
                st.session_state.admin_keep = False
                st.session_state.admin_clear_ls = True   # 다음 렌더에서 localStorage 삭제(재복원 방지)
                st.session_state.confirm_reset_all = False
                st.toast(t("admin_locked_toast"))
                st.rerun()
    else:
        # 몇 건이 사라지는지 숫자로 보여준다(실수 방지) + 되돌릴 수 있음을 함께 안내
        st.warning(t("reset_warn", n=len(st.session_state.bookings)))
        rc1, rc2 = st.columns(2)
        with rc1:
            if st.button(t("btn_reset_yes"), type="primary", key="reset_all_confirm_btn", use_container_width=True):
                # 전량 삭제 전에 ① 건수를 기록하고 ② 직전 상태를 스냅샷으로 남긴다 — 지운 뒤엔 되돌릴 근거가 없다.
                log_action("reset", "", 0, None, note=str(len(st.session_state.bookings)))
                save_snapshot(st.session_state.bookings, reason="reset")
                st.session_state.bookings = {}
                save_bookings(st.session_state.bookings)
                st.session_state.confirm_reset_all = False
                st.toast(t("toast_reset"))
                st.rerun()
        with rc2:
            if st.button(t("btn_cancel"), key="reset_all_cancel_btn", use_container_width=True):
                st.session_state.confirm_reset_all = False
                st.rerun()
    _render_pending_approvals()
    _render_backup_tools()
    _render_audit_log()

# 8-b. 관리자 '로그인 유지' 영속화 브릿지 — 체크 시 재접속해도 비밀번호 없이 자동 로그인.
#   ⚠️ 핵심 원칙: localStorage 저장(setItem)은 '로그인 시 1회성 컴포넌트'에서만 한다.
#      브릿지(매 틱 실행)는 절대 저장하지 않는다 → 로그아웃 후 예전 마커를 읽어 '1'을 되살리는
#      사고(클라우드 리런 지연 시 새로고침 자동 재로그인)를 원천 차단.
#   · #dk-admin-state 마커(data-cmd): hold=로그인+유지(건드리지 않음) / clear=삭제(로그아웃·미체크) / idle=미로그인(복원 대상)
#   · 숨김 RESTORE_ADMIN 버튼: idle+유지흔적('1')이면 JS가 대신 눌러 복원(성공까지 재시도).
st.button("RESTORE_ADMIN", key="restore_admin", on_click=_restore_admin)
if st.session_state.get("admin_unlocked"):
    _cmd = "hold" if st.session_state.get("admin_keep") else "clear"
    st.markdown('<span id="dk-admin-active" style="display:none;"></span>', unsafe_allow_html=True)
elif st.session_state.get("admin_clear_ls"):
    _cmd = "clear"   # 로그아웃/미체크 상태 → 재접속 자동복원 차단
else:
    _cmd = "idle"    # 미로그인 (localStorage 유지 흔적이 있으면 복원)
st.markdown(f'<span id="dk-admin-state" data-cmd="{_cmd}" style="display:none;"></span>', unsafe_allow_html=True)
# 로그인-유지 저장 1회성: 브릿지는 저장하지 않으므로, 유지 로그인 순간 여기서 딱 한 번 localStorage에 기록.
if st.session_state.get("admin_save_ls"):
    st.session_state.admin_save_ls = False
    _pwa_components.html(
        "<script>try{window.parent.localStorage.setItem('dk_admin','1');}catch(e){}</script>",
        height=0,
    )

# 8-c. '내 정보 기억' 1회성 기록/삭제 — 신청 완료를 누른 그 순간에만 쿠키를 쓴다.
#   8-b의 관리자 유지와 같은 원칙: 매 실행 도는 JS 브릿지는 쿠키를 절대 건드리지 않는다.
#   (브릿지가 값을 되살려 사고를 낸 전례 → 쓰기 경로를 '전환 시 1회'로만 좁혀 둔다)
if st.session_state.get("profile_save"):
    _pwa_components.html(_profile_cookie_script(st.session_state.pop("profile_save")), height=0)
elif st.session_state.get("profile_clear"):
    st.session_state.profile_clear = False
    _pwa_components.html(_profile_cookie_script(None), height=0)

# 9. 드래그 앤 드롭 이벤트를 부모 DOM에 강제로 바인딩하는 투명 JS 브릿지 컴포넌트 및 실시간 시계 가동
import streamlit.components.v1 as components
components.html("""
<script>
// ⚠️ 리런으로 이 컴포넌트 iframe이 새로 로드되면, 이전 iframe이 부모 요소에 걸어둔 클릭/드래그 리스너는
//    죽은 컨텍스트가 되어 눌러도 동작하지 않는다(예: 관리자 로그인 후 차량바 클릭 먹통).
//    남아있는 data-*-bound 표식을 iframe 로드 시 1회 초기화해, 현재 살아있는 iframe이 핸들러를 다시 소유하게 한다.
//    (초기화는 스크립트 로드당 1회뿐 → initDragDrop의 setTimeout 루프에선 재초기화 안 되므로 중복 바인딩 없음)
try {
    window.parent.document
        .querySelectorAll('[data-nav-bound],[data-admin-bound],[data-click-bound],[data-drag-bound],[data-drop-bound],[data-logout-bound]')
        .forEach(el => {
            el.removeAttribute('data-nav-bound'); el.removeAttribute('data-admin-bound');
            el.removeAttribute('data-click-bound'); el.removeAttribute('data-drag-bound');
            el.removeAttribute('data-drop-bound'); el.removeAttribute('data-logout-bound');
        });
} catch (e) {}

const initDragDrop = () => {
    const parentDoc = window.parent.document;
    const draggables = parentDoc.querySelectorAll('.seat-draggable');
    const droptargets = parentDoc.querySelectorAll('.seat-droptarget');
    const clickables = parentDoc.querySelectorAll('.seat-clickable');
    const carnavs = parentDoc.querySelectorAll('.car-nav-click');
    const adminseats = parentDoc.querySelectorAll('.admin-login-seat');

    // 요소가 렌더링되지 않았을 경우 대기
    if (draggables.length === 0 && droptargets.length === 0 && clickables.length === 0 && carnavs.length === 0 && adminseats.length === 0) {
        setTimeout(initDragDrop, 100);
        return;
    }

    // ⚡ INNOVA·SEDONA 운전석 클릭 → 대응하는 숨김 ADMINLOGIN 버튼을 대신 눌러 관리자 로그인 폼 오픈
    adminseats.forEach(el => {
        if (el.getAttribute('data-admin-bound') === 'true') return;
        el.setAttribute('data-admin-bound', 'true');
        el.addEventListener('click', () => {
            const car = el.getAttribute('data-car');
            const token = 'ADMINLOGIN::' + car;
            const btns = parentDoc.querySelectorAll('button');
            for (const b of btns) {
                if ((b.innerText || b.textContent || '').trim() === token) { b.click(); return; }
            }
        });
    });

    // ⚡ 로그아웃 버튼(좌석현황 팝업 / 관리자 초기화 영역) 클릭 순간 localStorage를 '동기적으로 즉시' 삭제.
    //   마커 cmd='clear'(리런 후 브릿지 처리)만으로는, 사용자가 로그아웃 직후 곧바로 새로고침하면
    //   삭제 전에 새 세션이 유지 흔적을 보고 복원해버릴 수 있다 → 클릭 시점에 바로 지워 타이밍 경쟁 제거.
    ['.st-key-admin_status_logout_btn button', '.st-key-admin_lock_btn button'].forEach(sel => {
        const lb = parentDoc.querySelector(sel);
        if (lb && lb.getAttribute('data-logout-bound') !== 'true') {
            lb.setAttribute('data-logout-bound', 'true');
            lb.addEventListener('click', () => {
                try {
                    // 로그아웃 플래그를 부모 창에 세워 이 페이지 수명 동안 재저장·재복원을 원천 차단 + 즉시 삭제
                    window.parent.__dkLoggedOut = true;
                    window.parent.localStorage.removeItem('dk_admin');
                } catch (e) {}
            });
        }
    });

    // ⚡ 관리자 '로그인 유지' — localStorage를 이 브릿지 '한 곳'에서 명령(data-cmd)에 따라 순차 제어.
    //   ⚠️ 저장/삭제/복원을 한 블록에서 순서대로 처리 → 로그아웃 삭제와 복원 재시도가 경쟁하지 않는다
    //      (예전엔 로그아웃 삭제를 별도 iframe에서 해서, 복원 재시도가 먼저 실행돼 로그아웃이 즉시 취소됐음).
    //   ⚠️ 반드시 부모(메인 앱) localStorage 사용 — 컴포넌트 iframe 자신의 저장소는 클라우드에서 다를 수 있음.
    try {
        const pwin = window.parent;
        const store = pwin.localStorage;
        const active = parentDoc.getElementById('dk-admin-active');   // 현재 로그인 상태면 존재
        const stEl = parentDoc.getElementById('dk-admin-state');
        const cmd = stEl ? stEl.getAttribute('data-cmd') : 'idle';
        // 로그인 상태가 되면 로그아웃 플래그 해제(재로그인 시 정상 동작 재개)
        if (active) pwin.__dkLoggedOut = false;
        // ⚠️ 브릿지는 절대 setItem 하지 않는다(저장은 로그인 1회성 컴포넌트 전담). 여기선 삭제/복원만.
        if (pwin.__dkLoggedOut || cmd === 'clear') {
            // 로그아웃/미체크 → 유지 흔적 삭제. 저장을 안 하므로 한 번 지우면 절대 되살아나지 않는다.
            store.removeItem('dk_admin');
        } else if (cmd === 'idle') {
            // 미로그인 + 유지 흔적('1') → 관리자 상태가 뜰 때까지 RESTORE_ADMIN 재시도 클릭.
            //   클라우드 콜드스타트에선 첫 클릭이 세션 준비 전이라 유실될 수 있어 '1회'로는 실패 → 성공까지 반복.
            if (store.getItem('dk_admin') === '1') {
                if (!pwin.__dkRestoreTries) pwin.__dkRestoreTries = 0;
                const nowT = Date.now();
                if (pwin.__dkRestoreTries < 30 && (!pwin.__dkRestoreLast || (nowT - pwin.__dkRestoreLast) > 1200)) {
                    pwin.__dkRestoreLast = nowT;
                    pwin.__dkRestoreTries++;
                    const btns = parentDoc.querySelectorAll('button');
                    for (const b of btns) {
                        if ((b.innerText || b.textContent || '').trim() === 'RESTORE_ADMIN') { b.click(); break; }
                    }
                }
            }
        }
        // cmd === 'hold' (로그인+유지): localStorage는 로그인 시 1회성 컴포넌트가 이미 '1'로 설정 → 브릿지는 관여 안 함.
    } catch (e) {}

    // ⚡ 앱: 차량 이름 바 클릭 → 대응하는 숨김 CARNAV 버튼을 대신 눌러 좌석맵 팝업 오픈
    carnavs.forEach(el => {
        if (el.getAttribute('data-nav-bound') === 'true') return;
        el.setAttribute('data-nav-bound', 'true');
        el.addEventListener('click', () => {
            const idx = el.getAttribute('data-navidx');
            const token = 'CARNAV::' + idx;
            const btns = parentDoc.querySelectorAll('button');
            for (const b of btns) {
                if ((b.innerText || b.textContent || '').trim() === token) { b.click(); return; }
            }
        });
    });

    // ⚡ 빈 좌석 클릭 → 대응하는 숨김 Streamlit 버튼을 대신 눌러 soft rerun 유도(전체 새로고침 없음)
    clickables.forEach(el => {
        if (el.getAttribute('data-click-bound') === 'true') return;
        el.setAttribute('data-click-bound', 'true');

        el.addEventListener('click', (e) => {
            const car = el.getAttribute('data-car');
            const seat = el.getAttribute('data-seat');
            const token = 'SEATSEL::' + car + '::' + seat;
            const btns = parentDoc.querySelectorAll('button');
            for (const b of btns) {
                if ((b.innerText || b.textContent || '').trim() === token) {
                    b.click();
                    return;
                }
            }
        });
    });

    // 드래그 가능한 좌석 이벤트 바인딩
    draggables.forEach(el => {
        if (el.getAttribute('data-drag-bound') === 'true') return;
        el.setAttribute('data-drag-bound', 'true');
        
        el.addEventListener('dragstart', (e) => {
            const car = el.getAttribute('data-car');
            const seat = el.getAttribute('data-seat');
            e.dataTransfer.setData('text/plain', car + '||' + seat);
        });
    });
    
    // 드롭 대상 빈 좌석 이벤트 바인딩
    droptargets.forEach(el => {
        if (el.getAttribute('data-drop-bound') === 'true') return;
        el.setAttribute('data-drop-bound', 'true');
        
        el.addEventListener('dragover', (e) => {
            e.preventDefault();
        });
        
        el.addEventListener('drop', (e) => {
            e.preventDefault();
            const sourceData = e.dataTransfer.getData('text/plain');
            if (!sourceData) return;
            
            const parts = sourceData.split('||');
            if (parts.length !== 2) return;
            const sourceCar = parts[0];
            const sourceSeat = parts[1];
            
            const targetCar = el.getAttribute('data-car');
            const targetSeat = el.getAttribute('data-seat');
            
            if (sourceCar === targetCar && sourceSeat === targetSeat) return;
            
            // 부모 창의 URL을 업데이트하여 Rerun 유도
            window.parent.location.href = "?drag_src_car=" + encodeURIComponent(sourceCar) + 
                                         "&drag_src_seat=" + encodeURIComponent(sourceSeat) + 
                                         "&drag_tgt_car=" + encodeURIComponent(targetCar) + 
                                         "&drag_tgt_seat=" + encodeURIComponent(targetSeat);
        });
    });
    
    // Rerun될 때 새로 만들어지는 요소를 계속 감시하기 위해 루핑 감시
    setTimeout(initDragDrop, 300);
};

// 실시간 디지털 시계 제어 루프 추가 (보안 CORS 에러 없는 동일 포트 iframe 상에서 작동)
const updateClock = () => {
    const parentDoc = window.parent.document;
    const clockEl = parentDoc.getElementById('live-digital-clock');
    if (clockEl) {
        const now = new Date();
        const year = now.getFullYear();
        const month = String(now.getMonth() + 1).padStart(2, '0');
        const date = String(now.getDate()).padStart(2, '0');
        const hours = String(now.getHours()).padStart(2, '0');
        const minutes = String(now.getMinutes()).padStart(2, '0');
        const seconds = String(now.getSeconds()).padStart(2, '0');
        clockEl.textContent = year + '-' + month + '-' + date + ' ' + hours + ':' + minutes + ':' + seconds;
    }
    setTimeout(updateClock, 1000);
};

// 최초 트리거 실행
initDragDrop();
updateClock();
</script>
""", height=0, width=0)

# 드래그앤드롭 및 레이아웃 안정화 완료