#!/usr/bin/env python
"""Claude Code Stop 훅.
Claude Code 턴이 끝날 때 실행된다. version_mark.py가 남긴 '증가 대기' 표시
(.version_pending)가 있으면 → 버전을 1 증가시키고 VER 백업을 만든 뒤 표시를 지운다.
표시가 없으면(=이번 턴에 app.py 코드 작업이 없었으면) 아무것도 하지 않는다.

버전 규칙:
  - 같은 날: count = max(저장 count, VER/mmdd 실제 최대번호) + 1  (되감김 없이 단조 증가)
  - 날짜가 바뀐 첫 작업: 그날 mmdd 로 새로 시작(보통 ver.1)
"""
import os
import re
import csv
import json
import shutil
import datetime
import subprocess

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MARKER = os.path.join(BASE_DIR, ".version_pending")
NOTE_FILE = os.path.join(BASE_DIR, ".version_note")   # Claude Code가 남긴 이번 버전 완료 내용(항목별 줄바꿈)
COUNTER_FILE = os.path.join(BASE_DIR, "version_counter.json")
APP_FILE = os.path.join(BASE_DIR, "app.py")
VER_DIR = os.path.join(BASE_DIR, "VER")

LOG_HEADERS = ["버전", "일시", "완료 내용 (Claude Code)"]
CONTENT_COL_WIDTH = 150   # 완료 내용 열너비(엑셀 문자 폭 단위)


def load_counter():
    if os.path.exists(COUNTER_FILE):
        try:
            with open(COUNTER_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def max_backup_count(display_date):
    """VER/{mmdd} 폴더의 app_ver_N.py 중 가장 큰 N(없으면 0).
    JSON이 유실·손상돼도 실제 백업에서 차수를 복구해 번호 재사용을 막는다."""
    try:
        d = os.path.join(VER_DIR, display_date)
        if not os.path.isdir(d):
            return 0
        mx = 0
        for fn in os.listdir(d):
            m = re.match(r"app_ver_(\d+)\.py$", fn)
            if m:
                mx = max(mx, int(m.group(1)))
        return mx
    except Exception:
        return 0


def _clear_note():
    """이번 버전에 소비한 완료 내용 노트를 삭제(다음 버전이 이전 내용을 재사용하지 않도록)."""
    try:
        if os.path.exists(NOTE_FILE):
            os.remove(NOTE_FILE)
    except Exception:
        pass


def _read_note():
    """이번 턴 완료 내용(.version_note, 항목별 줄바꿈)을 읽어 반환. 없으면 미기재 문자열."""
    note = ""
    if os.path.exists(NOTE_FILE):
        try:
            with open(NOTE_FILE, "r", encoding="utf-8") as f:
                note = f.read().strip()
        except Exception:
            note = ""
    return note or "(완료 내용 미기재)"


def _append_xlsx(path, ver_label, row):
    """xlsx에 한 줄 추가 — 완료 내용 셀은 줄바꿈 보존(wrap), 완료 내용 열너비 150, 같은 버전 중복 방지.
    openpyxl 미설치면 ImportError를 그대로 올려 상위에서 CSV 폴백을 타게 한다."""
    import openpyxl
    from openpyxl.styles import Alignment, Font

    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        for (val,) in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if val == ver_label:   # 이미 기록된 버전 → 중복 추가 안 함
                return
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "버전 로그"
        ws.append(LOG_HEADERS)
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(vertical="center", horizontal="center")

    ws.append(row)
    r = ws.max_row
    ws.cell(r, 1).alignment = Alignment(vertical="top", horizontal="center")
    ws.cell(r, 2).alignment = Alignment(vertical="top", horizontal="center")
    ws.cell(r, 3).alignment = Alignment(vertical="top", wrap_text=True)
    lines = str(row[2]).count("\n") + 1
    ws.row_dimensions[r].height = max(16, lines * 16)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = CONTENT_COL_WIDTH
    ws.freeze_panes = "A2"
    wb.save(path)


def _append_csv(path, ver_label, row):
    """openpyxl이 없을 때의 폴백 — CSV(줄바꿈은 셀 안에 따옴표로 보존, 열너비는 미지원)."""
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8-sig") as f:
                if (ver_label + ",") in f.read():
                    return
        except Exception:
            pass
    new_file = not os.path.exists(path)
    enc = "utf-8-sig" if new_file else "utf-8"
    with open(path, "a", encoding=enc, newline="") as f:
        w = csv.writer(f)
        if new_file:
            w.writerow(LOG_HEADERS)
        w.writerow(row)


def append_version_log(display_date, count, now):
    """버전별 완료 내용을 VER/{mmdd}/{mmdd}_버전별_완료내용.xlsx 에 한 줄 자동 추가한다.
    - 완료 내용은 .version_note(항목별 줄바꿈 텍스트)에서 읽어 셀 안에 줄바꿈 보존(wrap)한다.
    - 완료 내용 열너비 150, 같은 버전 줄이 있으면 중복 방지.
    - openpyxl 미설치 시 같은 이름의 .csv 로 폴백(줄바꿈은 유지, 열너비만 미적용).
    - VER/ 는 .gitignore 대상이라 로컬 문서로만 유지된다(배포/커밋 안 됨).
    실패해도 훅을 절대 죽이지 않는다."""
    try:
        day_dir = os.path.join(VER_DIR, display_date)
        os.makedirs(day_dir, exist_ok=True)
        base = os.path.join(day_dir, f"{display_date}_버전별_완료내용")
        ver_label = f"ver.{count}"
        row = [ver_label, f"{display_date} {now.strftime('%H:%M')}", _read_note()]
        try:
            _append_xlsx(base + ".xlsx", ver_label, row)
        except ImportError:
            _append_csv(base + ".csv", ver_label, row)
        _clear_note()
    except Exception:
        pass


def git(*args, timeout=60):
    """BASE_DIR에서 git 명령 실행. (returncode, stdout+stderr) 반환. 실패해도 예외 안 냄."""
    try:
        p = subprocess.run(
            ["git", *args],
            cwd=BASE_DIR,
            capture_output=True,
            text=True,
            timeout=timeout,
        )
        return p.returncode, (p.stdout or "") + (p.stderr or "")
    except Exception as e:
        return 1, str(e)


def auto_deploy(display_date, count):
    """버전이 오른 턴에 한해 변경사항을 자동 커밋 + push 하여 Streamlit 배포판을 즉시 갱신.
    실패(오프라인·인증 등)해도 훅을 죽이지 않는다. 로컬 커밋은 남으므로 다음 성공 push가 따라잡는다.
    안전장치: 현재 브랜치가 main일 때만 push 한다."""
    # 현재 브랜치 확인 (main 이 아니면 자동 push 하지 않음)
    rc, branch = git("rev-parse", "--abbrev-ref", "HEAD")
    if rc != 0 or branch.strip() != "main":
        return

    # 스테이징 (.gitignore가 VER/·bookings.json 등 로컬 전용 파일을 걸러줌)
    git("add", "-A")

    # 스테이징된 변경이 없으면 커밋 생략 (--quiet: 변경 있으면 0, 없으면 1)
    rc, _ = git("diff", "--cached", "--quiet")
    if rc == 0:
        # 커밋할 게 없어도 원격이 뒤처져 있을 수 있으니 push 는 시도
        git("push", "origin", "main")
        return

    git("commit", "-m", f"auto-deploy: {display_date} ver.{count} (app.py 자동 배포)")
    git("push", "origin", "main")


def main():
    # 증가 대기 표시가 없으면 종료 (app.py 코드 작업이 없었던 턴)
    if not os.path.exists(MARKER):
        return

    now = datetime.datetime.now()
    today = now.strftime("%Y-%m-%d")
    display_date = now.strftime("%m%d")

    data = load_counter()
    if data.get("date", "") == today:
        base = max(int(data.get("count", 0) or 0), max_backup_count(display_date))
    else:
        base = max_backup_count(display_date)  # 날짜가 바뀐 첫 작업 → 새로 시작
    count = base + 1

    try:
        with open(COUNTER_FILE, "w", encoding="utf-8") as f:
            json.dump({"date": today, "count": count}, f, ensure_ascii=False, indent=4)
    except Exception:
        pass

    # VER/[mmdd]/app_ver_N.py 백업 (count가 항상 최대+1이라 덮어쓰기 없음)
    try:
        backup_dir = os.path.join(VER_DIR, display_date)
        os.makedirs(backup_dir, exist_ok=True)
        if os.path.exists(APP_FILE):
            shutil.copy2(APP_FILE, os.path.join(backup_dir, f"app_ver_{count}.py"))
    except Exception:
        pass

    # 버전별 완료 내용 CSV에 이번 버전 줄 자동 추가(.version_note 소비 후 삭제)
    try:
        append_version_log(display_date, count, now)
    except Exception:
        pass

    # 표시 제거 (다음 턴에서 다시 코드 작업이 있어야 재증가)
    try:
        os.remove(MARKER)
    except Exception:
        pass

    # 버전이 올랐으니 변경사항을 자동 커밋 + push → Streamlit 배포판 즉시 갱신
    try:
        auto_deploy(display_date, count)
    except Exception:
        pass


if __name__ == "__main__":
    main()
